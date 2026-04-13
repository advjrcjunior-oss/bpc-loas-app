"""LegalMail Service v2 — Clean, correct, automated BPC/LOAS petition filing.

Usage:
    from legalmail_service import LegalMailService
    svc = LegalMailService()
    result = svc.processar_cliente(pasta)  # One call, everything done

API Endpoints (correct per OpenAPI spec 2026-03):
    POST/PUT/GET /api/v1/complaint          — Petition CRUD
    GET /api/v1/complaint/specialties       — Matéria/Especialidade
    GET /api/v1/complaint/district          — Comarcas/Jurisdições
    GET /api/v1/complaint/procedures        — Ritos
    GET /api/v1/complaint/classes           — Classes processuais
    GET /api/v1/complaint/subjects          — Assuntos
    GET /api/v1/complaint/areas             — Competências
    POST /api/v1/complaintsandpleadings/file        — Upload petição PDF
    POST /api/v1/complaintsandpleadings/attachments — Upload anexos
    GET /api/v1/complaintsandpleadings/attachment/types — Tipos de anexo
    GET /api/v1/complaintsandpleadings/courts       — Tribunais/sistemas
    POST /parts                             — Criar parte (legacy, funciona)
    GET /api/v1/party/professions           — Profissões válidas
"""
import os, json, time, re, math, requests, unicodedata
from datetime import datetime as _datetime
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

# ============================================================
# CONFIG
# ============================================================
LEGALMAIL_SITE = "https://app.legalmail.com.br"
LEGALMAIL_BASE = f"{LEGALMAIL_SITE}/api/v1"
VIACEP_BASE = "https://viacep.com.br/ws"
RATE_LIMIT_DELAY = float(os.environ.get('LEGALMAIL_RATE_DELAY', '4.0'))  # Fix #17: 4s minimum between requests (30 req/min limit)

# UF → TRF mapping (complete, verified against CNJ)
UF_TRIBUNAL_MAP = {
    'DF': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'DF'},
    'GO': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'GO'},
    'BA': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'BA'},
    'PI': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'PI'},
    'MA': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'MA'},
    'PA': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'PA'},
    'AM': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'AM'},
    'AC': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'AC'},
    'AP': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'AP'},
    'RR': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'RR'},
    'RO': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'RO'},
    'TO': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'TO'},
    'MT': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'MT'},
    'RJ': {'trf': 'TRF-2', 'sistema': 'eproc_jfrj', 'uf_tribunal': 'RJ'},
    'ES': {'trf': 'TRF-2', 'sistema': 'eproc_jfes', 'uf_tribunal': 'ES'},
    'SP': {'trf': 'TRF-3', 'sistema': 'pje', 'uf_tribunal': 'SP'},
    'MS': {'trf': 'TRF-3', 'sistema': 'pje', 'uf_tribunal': 'MS'},
    'PR': {'trf': 'TRF-4', 'sistema': 'eproc_jfpr', 'uf_tribunal': 'PR'},
    'SC': {'trf': 'TRF-4', 'sistema': 'eproc_jfsc', 'uf_tribunal': 'SC'},
    'RS': {'trf': 'TRF-4', 'sistema': 'eproc_jfrs', 'uf_tribunal': 'RS'},
    'PE': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'PE'},
    'CE': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'CE'},
    'AL': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'AL'},
    'SE': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'SE'},
    'RN': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'RN'},
    'PB': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'PB'},
    'MG': {'trf': 'TRF-6', 'sistema': 'pje', 'uf_tribunal': 'MG'},
}

INSS_DATA = {
    "nome": "INSTITUTO NACIONAL DO SEGURO SOCIAL - INSS",
    "polo": "passivo",
    "documento": "29.979.036/0001-40",
    "personalidade": "Pessoa jurídica",
    "endereco_cep": "70040-902",
    "endereco_logradouro": "SAUS Quadra 2 Bloco O",
    "endereco_numero": "S/N",
    "endereco_bairro": "Asa Sul",
    "endereco_cidade": "Brasilia",
    "endereco_uf": "DF",
}

# Doc prefix → attachment type name mapping
# Classificacao CORRETA no LegalMail (verificado caso Erick Machado 24/03/2026)
DOC_PREFIX_MAP = {
    "2": "Procuração",
    "3": "Contrato de Honorários",
    "4": "Pedido de Gratuidade de Justiça",  # Era "Declaração de Hipossuficiência"
    "5": "Documento de Identificação",        # Era "Certidão de Nascimento"
    "6": "Documento de Identificação",        # Familiares — mesma classificacao
    "7": "Comprovante de Residência",
    "8": "Outros",
    "9": "Outros",
    "10": "Outros",
    "11": "Outros",                           # Requerimento INSS
    "12": "Outros",                           # Carta indeferimento
    "13": "Exame Médico",                     # Era "Laudo Médico"
    "14": "Exame Médico",                     # Era "Receitas Médicas"
    "15": "Exame Médico",                     # Receitas e exames
    "16": "Exame Médico",                     # Relatório médico
    "17": "Exame Médico",                     # Exames de imagem
    "18": "Outros",
    "19": "Outros",
    "20": "Outros",                           # Biometria
    "21": "Documento de Identificação",
    "22": "Outros",
    "23": "Outros",
    "PLANILHA": "Planilha",                   # Planilha de cálculos
    "QUESITOS": "Outros",                     # Quesitos médicos/sociais
}


# Ordem de upload dos documentos (prefixo → prioridade)
# Ordem EXATA de upload (referencia: caso Erick Machado 24/03/2026)
ORDEM_DOCUMENTOS = [
    "1",   # Petição inicial (arquivo principal)
    "2",   # Procuração ad judicia (SOMENTE assinada, excluir administrativa)
    "4",   # Declaração de hipossuficiência / Gratuidade
    "3",   # Contrato de honorários
    "5",   # Documento de identidade do cliente/menor
    "6",   # Documentos de identidade familiares (agrupados)
    "7",   # Comprovante de residência
    "8",   # CadUnico
    "9",   # Autodeclaracao / composicao familiar
    "10",  # Comprovantes de renda
    "11",  # Requerimento INSS
    "12",  # Carta de indeferimento INSS
    "13",  # Laudo médico
    "14",  # Relatório médico
    "15",  # Receitas e exames
    "16",  # Relatório médico 2
    "17",  # Exames de imagem
    "18",  # Quesitos perícia médica (gerado)
    "19",  # Quesitos perícia social (gerado)
    "20",  # Biometria
    "21",  # OAB / Carteira profissional
    "22",  # Outros documentos
    "23",  # Outros documentos 2
    "QUESITOS_MED",  # Quesitos perícia médica (fallback)
    "QUESITOS_SOC",  # Quesitos perícia social (fallback)
    "PLANILHA",  # Planilha de cálculos (último)
]


def validar_pdf(caminho):
    """Valida que o arquivo é um PDF válido, não vazio, não corrompido."""
    if not os.path.exists(caminho):
        return False, "Arquivo nao existe"
    if os.path.getsize(caminho) < 100:
        return False, "Arquivo muito pequeno (possivelmente vazio)"
    if not PyPDF2:
        return True, "PyPDF2 nao disponivel, assumindo valido"
    try:
        with open(caminho, 'rb') as f:
            header = f.read(5)
            if header != b'%PDF-':
                return False, "Nao e um PDF valido (header incorreto, pode ser DOCX disfarçado)"
        with open(caminho, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            if len(reader.pages) == 0:
                return False, "PDF sem paginas"
        return True, "OK"
    except Exception as e:
        return False, f"PDF corrompido: {e}"


def validar_procuracao(caminho):
    """Verifica se a procuração tem assinatura (digital ou imagem)."""
    if not os.path.exists(caminho):
        return False, "Arquivo nao existe"
    if not PyPDF2:
        return True, "PyPDF2 nao disponivel, assumindo assinada"
    try:
        with open(caminho, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            # Verificar assinatura digital
            if '/Sig' in str(reader.trailer) or '/AcroForm' in str(reader.trailer):
                return True, "Assinatura digital detectada"
            # Verificar se tem imagens (possível assinatura escaneada)
            for page in reader.pages:
                resources = page.get('/Resources', {})
                xobject = resources.get('/XObject', {})
                if xobject:
                    return True, "Imagem detectada (possível assinatura)"
            # Se tem mais de 1 página, provavelmente tem assinatura na última
            if len(reader.pages) > 1:
                return True, "Multiplas paginas (assumindo assinatura)"
        return False, "Nenhuma assinatura detectada"
    except Exception as e:
        return False, f"Erro ao verificar: {e}"


def ordenar_documentos(arquivos):
    """Ordena lista de arquivos na sequência correta para upload."""
    def prioridade(nome):
        nome_upper = nome.upper()
        if "PLANILHA" in nome_upper:
            return ORDEM_DOCUMENTOS.index("PLANILHA") if "PLANILHA" in ORDEM_DOCUMENTOS else 100
        if "QUESITOS" in nome_upper and "MEDIC" in nome_upper:
            return ORDEM_DOCUMENTOS.index("QUESITOS_MED") if "QUESITOS_MED" in ORDEM_DOCUMENTOS else 98
        if "QUESITOS" in nome_upper and "SOCIAL" in nome_upper:
            return ORDEM_DOCUMENTOS.index("QUESITOS_SOC") if "QUESITOS_SOC" in ORDEM_DOCUMENTOS else 99
        if "QUESITOS" in nome_upper:
            return 98  # quesitos generico
        # Extrair prefixo numerico
        m = re.match(r'^(\d+)', os.path.basename(nome))
        if m:
            num = int(m.group(1))
            try:
                return ORDEM_DOCUMENTOS.index(str(num))
            except ValueError:
                return 50 + num
        return 97
    return sorted(arquivos, key=prioridade)


# ============================================================
# VIACEP SERVICE
# ============================================================
def validar_cep(cep):
    """Validate CEP via ViaCEP. Returns address dict or None."""
    clean = re.sub(r'\D', '', str(cep))
    if len(clean) != 8:
        return None
    try:
        r = requests.get(f"{VIACEP_BASE}/{clean}/json/", timeout=10)
        if r.status_code == 200:
            data = r.json()
            if not data.get('erro'):
                return data
    except Exception:
        pass
    return None


def buscar_cep_por_endereco(uf, cidade, logradouro):
    """Search CEP by address via ViaCEP. Returns CEP string or None."""
    try:
        rua = re.sub(r'\s+', '+', logradouro.strip()[:40])
        url = f"{VIACEP_BASE}/{uf}/{cidade}/{rua}/json/"
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list) and len(data) > 0:
                return data[0].get('cep', '').replace('-', '')
    except Exception:
        pass
    return None


def completar_endereco_por_cep(cep):
    """Get full address from CEP. Returns dict with logradouro, bairro, cidade, uf."""
    data = validar_cep(cep)
    if not data:
        return None
    return {
        'endereco_cep': data.get('cep', '').replace('-', ''),
        'endereco_logradouro': data.get('logradouro', ''),
        'endereco_bairro': data.get('bairro', ''),
        'endereco_cidade': data.get('localidade', ''),
        'endereco_uf': data.get('uf', ''),
    }


# ============================================================
# LEGALMAIL SERVICE
# ============================================================
def _normalize_text(s):
    """Remove acentos e converte pra uppercase pra comparacao."""
    return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode().upper()


class LegalMailService:
    """Clean LegalMail API client for BPC/LOAS petitions."""

    def __init__(self, api_key=None, cert_id=None):
        self.api_key = api_key or os.environ.get('LEGALMAIL_API_KEY', '')
        self.cert_id = cert_id or int(os.environ.get('LEGALMAIL_CERTIFICADO_ID', '0') or '0')
        self._inss_id = None
        self._last_request = 0
        self._options_cache = {}  # Cache tribunal lookups (specialties, districts, etc.)
        self._session_cookies = None  # For internal proporAcao/update endpoint
        self._http = requests.Session()  # Reutiliza conexoes TCP (pool)
        if not self.api_key:
            raise ValueError("LEGALMAIL_API_KEY não configurada")

    # ---- Low-level API ----

    def _request(self, method, endpoint, **kwargs):
        """Make rate-limited authenticated request with timeout/connection retry."""
        elapsed = time.time() - self._last_request
        if elapsed < RATE_LIMIT_DELAY:
            time.sleep(RATE_LIMIT_DELAY - elapsed)

        sep = '&' if '?' in endpoint else '?'
        url = f"{LEGALMAIL_BASE}{endpoint}{sep}api_key={self.api_key}"
        kwargs.setdefault('timeout', 30)

        for net_retry in range(2):
            try:
                self._last_request = time.time()
                r = getattr(self._http, method)(url, **kwargs)
                break
            except (requests.Timeout, requests.ConnectionError) as e:
                if net_retry == 0:
                    print(f"    [NET] {method.upper()} falhou ({e.__class__.__name__}), retry em 10s...")
                    time.sleep(10)
                else:
                    raise Exception(f"API indisponivel apos 2 tentativas: {e}")

        # Fix #17: Progressive retry on rate limit (up to 3 attempts)
        for retry in range(3):
            if r.status_code != 429:
                break
            wait = 60 * (retry + 1)  # 60s, 120s, 180s
            print(f"    [RATE LIMIT] Aguardando {wait}s... (tentativa {retry+2}/3)")
            time.sleep(wait)
            self._last_request = time.time()
            r = getattr(requests, method)(url, **kwargs)

        return r

    def _get_options(self, endpoint, idpet, use_cache=True):
        """GET options list from a complaint/* endpoint. Caches per endpoint+tribunal."""
        cache_key = f"{endpoint}:{idpet}"
        if use_cache and cache_key in self._options_cache:
            return self._options_cache[cache_key]
        r = self._request("get", f"{endpoint}?idpeticoes={idpet}")
        if r.status_code == 200:
            try:
                data = r.json()
                if isinstance(data, list):
                    if use_cache:
                        self._options_cache[cache_key] = data
                    return data
            except (ValueError, AttributeError):
                pass
        return []

    def _inferir_tribunal(self, dados):
        """Inferir tribunal a partir do sistema ou ufTribunal quando campo tribunal esta vazio."""
        uf = dados.get('ufTribunal', '')
        sistema = dados.get('sistema', '')
        # Pelo sistema
        if 'jfpr' in sistema or 'jfsc' in sistema or 'jfrs' in sistema:
            return 'TRF-4'
        if 'jfrj' in sistema or 'jfes' in sistema:
            return 'TRF-2'
        # Pela UF
        uf_map = {'SP': 'TRF-3', 'MS': 'TRF-3', 'MG': 'TRF-6',
                   'RN': 'TRF-5', 'PE': 'TRF-5', 'CE': 'TRF-5', 'AL': 'TRF-5', 'SE': 'TRF-5', 'PB': 'TRF-5',
                   'PR': 'TRF-4', 'SC': 'TRF-4', 'RS': 'TRF-4',
                   'RJ': 'TRF-2', 'ES': 'TRF-2'}
        if uf in uf_map:
            return uf_map[uf]
        # Default: TRF-1 (cobre DF,GO,BA,MT,PA,AM,etc)
        return 'TRF-1'

    # ---- Internal session for proporAcao/update (bypasses API classe requirement) ----

    def _get_session(self):
        """Login to LegalMail site and get session cookies.
        Uses LEGALMAIL_EMAIL and LEGALMAIL_PASSWORD env vars, or defaults."""
        if self._session_cookies:
            return self._session_cookies
        email = os.environ.get('LEGALMAIL_EMAIL', '')
        password = os.environ.get('LEGALMAIL_PASSWORD', '')
        if not email or not password:
            print("    [SESSION] LEGALMAIL_EMAIL/PASSWORD nao configurados no .env")
            return None
        s = requests.Session()
        r = s.post(f'{LEGALMAIL_SITE}/login', data={
            'email': email, 'password': password
        }, allow_redirects=True, timeout=15)
        if r.status_code == 200 and s.cookies:
            self._session_cookies = dict(s.cookies)
            print("    [SESSION] Login OK")
            return self._session_cookies
        print(f"    [SESSION] Login falhou: {r.status_code}")
        return None

    def _proporAcao_update(self, idpet, idproc, dados):
        """Use internal endpoint POST /api/proporAcao/update to set fields.
        This bypasses the API classe/areas requirement for eProc.
        Requires session cookies (not api_key)."""
        cookies = self._get_session()
        if not cookies:
            return False

        ASSUNTO_BPC = 'DIREITO ASSISTENCIAL (14) | Benefício Assistencial (Art. 203,V CF/88) (1402) | Pessoa com Deficiência (140201)'

        form = {
            'fk_peticao': str(idpet),
            'fk_processo': str(idproc),
            'assuntoPrincipal_proporAcao': ASSUNTO_BPC,
            'classeProcessual_proporAcao': '',
            'ritos_proporAcao': dados.get('rito', 'JUIZADO ESPECIAL FEDERAL'),
            'especialidade_proporAcao': dados.get('competencia', 'Federal'),
            'areas_proporAcao': dados.get('area', ''),
            # OBRIGATORIO: proporAcao/update é destrutivo — campos ausentes sao resetados
            'tribunal_proporAcao': dados.get('tribunal') or self._inferir_tribunal(dados),
            'orgao_julgador_destino': dados.get('tribunal') or self._inferir_tribunal(dados),
            'sistema_tribunal': dados.get('sistema', ''),
            'foro_proporAcao': dados.get('comarca', ''),
            'grau_proporAcao': '1º Grau',
            'tipoDistribuicao_proporAcao': dados.get('distribuicao', 'Por sorteio'),
            'titulo': 'Petição Inicial',
            'justicaGratuita_proporAcao': '1',
            'antecipacao_tutela': '1',
            'juizo_digital': '1',
            'renuncia_60_salarios': '1',
        }
        # Parties
        ativo = dados.get('idpoloativo', [])
        passivo = dados.get('idpolopassivo', [])
        if ativo:
            form['poloAtivo_proporAcao'] = str(ativo[0])
        if passivo:
            form['poloPassivo_proporAcao'] = str(passivo[0])
        if len(ativo) > 1:
            form['processos_clientes[]'] = [str(a) for a in ativo[1:]]
        if dados.get('valorCausa'):
            # proporAcao/update: o campo aceita valor em REAIS inteiros
            # O endpoint NAO multiplica por 100 — ele salva o valor como enviado
            # Mas centavos sao ignorados pelo form submit, entao arredondar pra cima
            vc = dados['valorCausa']
            if isinstance(vc, float):
                vc = math.ceil(vc)  # Arredondar PRA CIMA (nunca pedir menos que o devido)
            form['valorCausa_proporAcao'] = str(int(vc))

        r = requests.post(f'{LEGALMAIL_SITE}/api/proporAcao/update',
                         data=form, cookies=cookies, timeout=15)
        try:
            resp_ok = r.status_code == 200 and r.json().get('status') == 'success'
        except (ValueError, AttributeError):
            resp_ok = r.status_code == 200 and 'success' in r.text
        if resp_ok:
            print(f"    [proporAcao] Assunto+rito OK via endpoint interno")
            return True
        print(f"    [proporAcao] Falhou: {r.status_code} {r.text[:100]}")
        return False

    # ---- Petition CRUD (correct endpoints) ----

    def criar_peticao(self, tribunal, sistema, instancia="1", uf_tribunal="",
                      polo_ativo_ids=None, polo_passivo_ids=None):
        """POST /complaint — Create petition draft WITH parties linked.

        Fix #9: Always include ufTribunal (required by TRF-1 PJe).
        Fix #11: Include parties in POST to avoid eProc deadlock.
        """
        payload = {
            "tribunal": tribunal,
            "instancia": instancia,
            "sistema": sistema,
            "certificado_id": self.cert_id,
        }
        if uf_tribunal:
            payload["ufTribunal"] = uf_tribunal
        if polo_ativo_ids:
            payload["idpoloativo"] = polo_ativo_ids
        if polo_passivo_ids:
            payload["idpolopassivo"] = polo_passivo_ids

        r = self._request("post", "/complaint", json=[payload])
        if r.status_code != 200:
            raise Exception(f"Erro ao criar petição: {r.status_code} {r.text[:300]}")

        try:
            data = r.json()
        except ValueError:
            raise Exception(f"Resposta nao-JSON ao criar petição: {r.text[:300]}")
        if isinstance(data, list):
            data = data[0]
        if data.get("status") == "erro":
            raise Exception(f"Erro: {data.get('mensagem', data)}")

        dados = data.get("dados", {})
        idpet = dados.get("idpeticoes")
        idproc = dados.get("idprocessos")
        if not idpet:
            raise Exception(f"ID não retornado: {data}")
        print(f"    Rascunho criado: id={idpet}")
        return int(idpet), int(idproc)

    def get_peticao(self, idpet):
        """GET /complaint — Get current petition data."""
        r = self._request("get", f"/complaint?idpeticoes={idpet}")
        if r.status_code != 200:
            raise Exception(f"GET petição {idpet}: {r.status_code}")
        try:
            return r.json()["peticao"]["dados"]
        except (ValueError, KeyError, TypeError) as e:
            raise Exception(f"GET petição {idpet}: resposta invalida — {str(r.text)[:200]}")

    def put_peticao(self, idpet, dados):
        """PUT /complaint — Update petition fields. Returns True/False."""
        r = self._request("put", f"/complaint?idpeticoes={idpet}", json=dados)
        if r.status_code == 200:
            return True
        print(f"    PUT ERRO {r.status_code}: {r.text[:200]}")
        return False

    # ---- Field fill using INTERNAL APIs (discovered 25/03/2026) ----

    def _search_internal(self, endpoint, params):
        """GET /api/search* using session cookies (internal frontend API)."""
        cookies = self._get_session()
        if not cookies:
            return []
        try:
            r = requests.get(f'{LEGALMAIL_SITE}/api/{endpoint}',
                           params=params, cookies=cookies, timeout=15)
            if r.status_code == 200:
                data = r.json()
                return data if isinstance(data, list) else []
        except Exception:
            pass
        return []

    def preencher_campos(self, idpet, cidade, uf, tipo='deficiente'):
        """Fill ALL petition fields using INTERNAL APIs (same as frontend).

        Uses /api/search* for lookups and /api/proporAcao/update to save ALL at once.
        This avoids the PUT /complaint dependency chain that causes field failures.

        Returns dict of filled/failed fields.
        """
        result = {'filled': [], 'errors': []}

        # Get current state
        dados = self.get_peticao(idpet)
        tribunal = dados.get('tribunal', '') or self._inferir_tribunal(dados)
        sistema = dados.get('sistema', '')
        idproc = None
        try:
            r = self._request("get", f"/complaint?idpeticoes={idpet}")
            idproc = r.json().get('peticao', {}).get('idprocessos')
        except Exception:
            pass

        def _opt_text(opt):
            """Extract text from search API option (dict or string)."""
            if isinstance(opt, dict):
                return str(opt.get('text', opt.get('nome', '')))
            return str(opt)

        # ===== STEP 1: Resolve foro/comarca =====
        foro = dados.get('comarca', '') or ''
        if not foro and cidade:
            options = self._search_internal('searchMunicipios', {
                'term': cidade[:20], 'tribunal': tribunal,
                'sistema_tribunal': sistema, 'instancia': '1º Grau',
            })
            if options:
                for opt in options:
                    text = _opt_text(opt)
                    if text and _normalize_text(cidade) in _normalize_text(text):
                        foro = text
                        break
                if not foro:
                    cidade_fb = self.COMARCA_FALLBACK.get(cidade.upper(), '')
                    if cidade_fb:
                        for opt in options:
                            text = _opt_text(opt)
                            if text and _normalize_text(cidade_fb) in _normalize_text(text):
                                foro = text
                                break
                if not foro and options:
                    foro = _opt_text(options[0])
                    result['errors'].append(f"comarca: '{cidade}' nao encontrada, usou {foro}")
            if foro:
                result['filled'].append(f"foro={foro[:40]}")
                print(f"    foro={foro[:40]}: OK")

        # ===== STEP 2: Resolve rito =====
        rito = dados.get('rito', '') or ''
        if not rito:
            options = self._search_internal('searchRitos', {
                'term': 'Juizado', 'tribunal': tribunal,
                'grau': '1º Grau', 'sistema_tribunal': sistema, 'foro': foro,
            })
            for opt in (options or []):
                text = _opt_text(opt)
                if 'juizado' in text.lower() and 'federal' in text.lower():
                    rito = text
                    break
            if not rito and options:
                rito = _opt_text(options[0])
            if rito:
                result['filled'].append(f"rito={rito}")
                print(f"    rito={rito}: OK")

        # ===== STEP 3: Resolve classe =====
        classe = dados.get('classe', '') or ''
        if not classe:
            options = self._search_internal('searchClasseProcessual', {
                'term': 'Juizado', 'tribunal': tribunal,
                'especialidade': dados.get('competencia', 'Federal'),
                'grau': '1º Grau', 'foro': foro,
                'sistema_tribunal': sistema, 'ritos': rito,
            })
            for opt in (options or []):
                text = _opt_text(opt)
                if 'juizado' in text.lower() and ('cível' in text.lower() or 'civel' in text.lower()):
                    classe = text
                    break
                if 'juizado' in text.lower():
                    classe = text
                    break
            if not classe and options:
                classe = _opt_text(options[0])
            if classe:
                result['filled'].append(f"classe={classe[:40]}")
                print(f"    classe={classe[:40]}: OK")

        # ===== STEP 4: Resolve assunto =====
        assunto = dados.get('assunto', '') or ''
        if not assunto:
            search_term = 'Assistencial'
            options = self._search_internal('searchAssuntos', {
                'term': search_term, 'tribunal': tribunal,
                'classe_processual': classe, 'grau': '1º Grau',
                'especialidade': dados.get('competencia', ''),
                'foro': foro, 'sistema_tribunal': sistema,
                'ritos': rito, 'complementar': 'nao',
            })
            target = 'defici' if tipo == 'deficiente' else 'idoso'
            for opt in (options or []):
                text = _opt_text(opt)
                if target in text.lower() and 'assistencial' in text.lower():
                    assunto = text
                    break
            if not assunto:
                for opt in (options or []):
                    text = _opt_text(opt)
                    if 'assistencial' in text.lower() and '203' in text:
                        assunto = text
                        break
            if not assunto and options:
                assunto = _opt_text(options[0])
            if assunto:
                result['filled'].append(f"assunto={assunto[:50]}")
                print(f"    assunto={assunto[:50]}: OK")

        # ===== STEP 5: Save ALL via proporAcao/update =====
        cookies = self._get_session()
        if cookies and idproc:
            form = {
                'fk_peticao': str(idpet),
                'fk_processo': str(idproc),
                # OBRIGATORIO enviar tribunal — proporAcao/update é DESTRUTIVO
                # Campos nao enviados sao RESETADOS pra vazio
                'tribunal_proporAcao': tribunal,
                'orgao_julgador_destino': tribunal,
                'sistema_tribunal': sistema,
                'grau_proporAcao': '1º Grau',
                'foro_proporAcao': foro,
                'ritos_proporAcao': rito,
                'classeProcessual_proporAcao': classe,
                'assuntoPrincipal_proporAcao': assunto,
                'especialidade_proporAcao': dados.get('competencia', '') or 'Direito Previdenciário',
                'areas_proporAcao': dados.get('area', ''),
                'tipoDistribuicao_proporAcao': 'Por sorteio',
                'titulo': 'Petição Inicial',
                'justicaGratuita_proporAcao': '1',
                'antecipacao_tutela': '1',
                'juizo_digital': '1',
                'renuncia_60_salarios': '1',
            }
            # Partes
            polo_a = dados.get('idpoloativo', [])
            polo_p = dados.get('idpolopassivo', [])
            if polo_a:
                form['poloAtivo_proporAcao'] = str(polo_a[0])
            if len(polo_a) > 1:
                form['processos_clientes[]'] = [str(a) for a in polo_a[1:]]
            if polo_p:
                form['poloPassivo_proporAcao'] = str(polo_p[0])
            # Valor
            if dados.get('valorCausa') and dados['valorCausa'] > 0:
                form['valorCausa_proporAcao'] = str(int(math.ceil(dados['valorCausa'])))

            r = requests.post(f'{LEGALMAIL_SITE}/api/proporAcao/update',
                             data=form, cookies=cookies, timeout=15)
            try:
                resp_ok = r.status_code == 200 and r.json().get('status') == 'success'
            except (ValueError, AttributeError):
                resp_ok = r.status_code == 200 and 'success' in r.text
            if resp_ok:
                result['filled'].append('proporAcao/update OK')
                print(f"    proporAcao/update: TODOS OS CAMPOS SALVOS")
            else:
                result['errors'].append(f'proporAcao/update: {r.status_code}')
                print(f"    proporAcao/update ERRO: {r.status_code} {r.text[:100]}")
        else:
            result['errors'].append('sem session ou idproc')

        return result

    # Mapa de fallback: cidade → comarca quando busca direta falha
    COMARCA_FALLBACK = {
        # RS (TRF-4 JFRS)
        'FARROUPILHA': 'Caxias do Sul', 'GARIBALDI': 'Caxias do Sul',
        'BENTO GONCALVES': 'Caxias do Sul', 'FLORES DA CUNHA': 'Caxias do Sul',
        'CARLOS BARBOSA': 'Caxias do Sul',
        'BARRA DO RIBEIRO': 'Porto Alegre', 'GUAIBA': 'Porto Alegre',
        'VIAMAO': 'Porto Alegre', 'GRAVATAI': 'Porto Alegre',
        'CANOAS': 'Porto Alegre', 'NOVO HAMBURGO': 'Porto Alegre',
        # RJ (TRF-2 JFRJ)
        'CASIMIRO DE ABREU': 'Macaé', 'RIO DAS OSTRAS': 'Macaé',
        'CACHOEIRAS DE MACACU': 'Itaboraí', 'GUAPIMIRIM': 'Itaboraí',
        'SILVA JARDIM': 'Itaboraí', 'TANGUA': 'Itaboraí',
        'MAGE': 'Magé', 'TERESOPOLIS': 'Petrópolis',
        # ES (TRF-2 JFES)
        'CASTELO': 'Cachoeiro de Itapemirim', 'VARGEM ALTA': 'Cachoeiro de Itapemirim',
        'ICONHA': 'Cachoeiro de Itapemirim', 'ALFREDO CHAVES': 'Cachoeiro de Itapemirim',
        'MUNIZ FREIRE': 'Cachoeiro de Itapemirim', 'IUNA': 'Cachoeiro de Itapemirim',
        # PR (TRF-4 JFPR)
        'SAO JOSE DOS PINHAIS': 'Curitiba', 'COLOMBO': 'Curitiba',
        'PIRAQUARA': 'Curitiba', 'PINHAIS': 'Curitiba',
        'ARAUCARIA': 'Curitiba', 'ALMIRANTE TAMANDARE': 'Curitiba',
        'CAFELANDIA': 'Campo Mourão', 'GOIOERE': 'Campo Mourão',
        'PONTA GROSSA': 'Ponta Grossa', 'CASTRO': 'Ponta Grossa',
        'IRATI': 'Ponta Grossa', 'GUARAPUAVA': 'Guarapuava',
        # SC (TRF-4 JFSC)
        'ARMAZEM': 'Tubarão', 'TUBARAO': 'TUBARÃO', 'LAGUNA': 'Tubarão',
        'CAMPOS NOVOS': 'Lages', 'CURITIBANOS': 'Lages',
        'JARAGUA DO SUL': 'Joinville', 'SAO BENTO DO SUL': 'Joinville',
        # SP (TRF-3 PJe) — cidades que caem em subseções grandes
        'ARACARIGUAMA': 'Americana', 'RAFARD': 'Americana',
        'SANTA CRUZ DAS PALMEIRAS': 'Ribeirão Preto',
        # GO (TRF-1 PJe)
        'GOIANIA': 'Aparecida de Goiânia',
        # MT (TRF-1 PJe)
        'COLNIZA': 'Barra do Garças',
    }

    def _match_comarca(self, districts, cidade, uf):
        """Match city to comarca/subsecao from available options.
        Handles: exact city, Subsecao Judiciaria de CIDADE, JEF preference, fallback map."""
        if not cidade:
            return None
        cidade_upper = cidade.upper().strip()
        cidade_norm = _normalize_text(cidade)
        nomes = [d.get('nome', '') for d in districts]

        # 1. Exact match
        for n in nomes:
            if cidade_upper == n.upper():
                return n

        # 2. City name in subsecao (e.g., "Subsecao Judiciaria de Campinas")
        matches = [n for n in nomes if cidade_norm in _normalize_text(n)]
        if matches:
            # Prefer JEF (Juizado Especial Federal) over regular subsecao
            jef = [m for m in matches if 'JUIZADO' in m.upper() or 'JEF' in m.upper()]
            if jef:
                return jef[0]
            # Prefer Subsecao over Secao (more specific)
            subsecao = [m for m in matches if 'SUBSEC' in _normalize_text(m) or 'SUBSE' in _normalize_text(m)]
            if subsecao:
                return subsecao[0]
            return matches[0]

        # 3. Partial match (first word of city name)
        first_word = cidade_norm.split()[0] if cidade_norm else ''
        if first_word and len(first_word) > 3:
            matches = [n for n in nomes if first_word in _normalize_text(n)]
            jef = [m for m in matches if 'JUIZADO' in m.upper() or 'JEF' in m.upper()]
            if jef:
                return jef[0]
            if matches:
                return matches[0]

        # 4. For TRF-1, try "Secao Judiciaria de UF"
        if uf:
            uf_matches = [n for n in nomes if uf.upper() in _normalize_text(n) and 'SEC' in _normalize_text(n)]
            if uf_matches:
                return uf_matches[0]

        # 5. Fallback map (cidades conhecidas → comarca)
        cidade_fallback = self.COMARCA_FALLBACK.get(cidade_upper, '')
        if cidade_fallback:
            fb_norm = _normalize_text(cidade_fallback)
            for n in nomes:
                if fb_norm in _normalize_text(n):
                    print(f"    [COMARCA] Fallback: {cidade} -> {n}")
                    return n

        # 6. Fallback: busca na internet pela jurisdição correta
        jurisdicao_web = self._buscar_jurisdicao_web(cidade, uf, nomes)
        if jurisdicao_web:
            return jurisdicao_web

        return None

    def _buscar_jurisdicao_web(self, cidade, uf, nomes_disponiveis):
        """Fallback: busca jurisdição pela capital do estado quando match local falha."""
        if not cidade or not uf:
            return None
        try:
            # Fallback: tentar cidade sede da seção judiciária do estado
            capitais = {
                'SP': 'SAO PAULO', 'RJ': 'RIO DE JANEIRO', 'MG': 'BELO HORIZONTE',
                'RS': 'PORTO ALEGRE', 'PR': 'CURITIBA', 'SC': 'FLORIANOPOLIS',
                'BA': 'SALVADOR', 'PE': 'RECIFE', 'CE': 'FORTALEZA',
                'PA': 'BELEM', 'AM': 'MANAUS', 'GO': 'GOIANIA',
                'DF': 'BRASILIA', 'ES': 'VITORIA', 'AL': 'MACEIO',
                'SE': 'ARACAJU', 'RN': 'NATAL', 'PB': 'JOAO PESSOA',
                'PI': 'TERESINA', 'MA': 'SAO LUIS', 'MT': 'CUIABA',
                'MS': 'CAMPO GRANDE', 'RO': 'PORTO VELHO', 'AC': 'RIO BRANCO',
                'AP': 'MACAPA', 'RR': 'BOA VISTA', 'TO': 'PALMAS',
            }
            capital = capitais.get(uf.upper(), '')
            if capital:
                capital_matches = [n for n in nomes_disponiveis if capital in _normalize_text(n)]
                if capital_matches:
                    print(f"    [WEB] Jurisdição fallback para capital: {capital_matches[0]}")
                    return capital_matches[0]

        except Exception as e:
            print(f"    [WEB] Erro buscar jurisdição: {e}")
        return None

    def _match_assunto_bpc(self, subjects, tipo='deficiente'):
        """Find BPC/LOAS subject from list. Never pick previdenciário por incapacidade."""
        search = 'defici' if tipo == 'deficiente' else 'idoso'

        # Priority 1: ASSISTENCIAL + Art. 203 + deficiente/idoso
        for s in subjects:
            nome = s.get('nome', '')
            if 'assistencial' in nome.lower() and '203' in nome and search in nome.lower():
                return nome

        # Priority 2: ASSISTENCIAL + Pessoa com Deficiência
        for s in subjects:
            nome = s.get('nome', '')
            if 'assistencial' in nome.lower() and search in nome.lower():
                return nome

        # Priority 3: Any ASSISTENCIAL (not previdenciário!)
        for s in subjects:
            nome = s.get('nome', '')
            if 'assistencial' in nome.lower() and 'benefício assistencial' in nome.lower():
                return nome

        return None

    # ---- Party management ----

    def criar_parte(self, party_data):
        """POST /parts — Create party (idempotent by CPF). Returns party ID."""
        # Ensure required fields
        if 'personalidade' not in party_data:
            party_data['personalidade'] = 'Pessoa física'
        if 'polo' not in party_data:
            party_data['polo'] = 'ativo'
        if not party_data.get('endereco_numero'):
            party_data['endereco_numero'] = 'S/N'
        if not party_data.get('endereco_bairro'):
            party_data['endereco_bairro'] = 'Centro'  # Fallback — LegalMail requires non-empty bairro
        if not party_data.get('endereco_logradouro'):
            party_data['endereco_logradouro'] = 'Nao informado'

        # Validate and fix CEP — LegalMail REQUIRES hyphen format (XXXXX-XXX)
        cep = party_data.get('endereco_cep', '')
        if cep:
            cep_clean = re.sub(r'\D', '', cep)
            addr = validar_cep(cep_clean)
            if addr:
                # LegalMail requires XXXXX-XXX format (with hyphen!)
                party_data['endereco_cep'] = f"{cep_clean[:5]}-{cep_clean[5:]}"
                # Fill missing address fields from ViaCEP
                if not party_data.get('endereco_cidade'):
                    party_data['endereco_cidade'] = addr.get('localidade', '')
                if not party_data.get('endereco_uf'):
                    party_data['endereco_uf'] = addr.get('uf', '')
                if not party_data.get('endereco_bairro') and addr.get('bairro'):
                    party_data['endereco_bairro'] = addr['bairro']
                if not party_data.get('endereco_logradouro') and addr.get('logradouro'):
                    party_data['endereco_logradouro'] = addr['logradouro']
            else:
                print(f"    [WARN] CEP {cep} inválido, tentando buscar...")
                uf = party_data.get('endereco_uf', '')
                cidade = party_data.get('endereco_cidade', '')
                rua = party_data.get('endereco_logradouro', '')
                if uf and cidade and rua:
                    novo_cep = buscar_cep_por_endereco(uf, cidade, rua)
                    if novo_cep:
                        clean = re.sub(r'\D', '', novo_cep)
                        party_data['endereco_cep'] = f"{clean[:5]}-{clean[5:]}"
                        print(f"    [CEP] Corrigido: {cep} -> {party_data['endereco_cep']}")

        # Fix #7: CEP rural sem logradouro — ViaCEP retorna vazio
        # Checar ViaCEP ANTES de setar defaults (evita logica circular)
        is_rural = False
        cep_check = re.sub(r'\D', '', party_data.get('endereco_cep', ''))
        if cep_check and len(cep_check) == 8:
            addr_check = validar_cep(cep_check)
            if addr_check and not addr_check.get('logradouro'):
                is_rural = True

        if is_rural:
            party_data['endereco_logradouro'] = 'Zona Rural'
            party_data['endereco_bairro'] = 'Zona Rural'
            party_data['endereco_numero'] = 'S/N'
        else:
            if not party_data.get('endereco_logradouro') or party_data['endereco_logradouro'] == 'Nao informado':
                party_data['endereco_logradouro'] = 'Zona Rural'
            if not party_data.get('endereco_bairro'):
                party_data['endereco_bairro'] = 'Centro'

        # Fix #8: TRF-2 eProc requires etnia — always include it
        if 'etnia' not in party_data:
            party_data['etnia'] = 'Nao declarada'

        r = self._request("post", "/parts", json=party_data)
        if r.status_code == 200:
            try:
                pid = int(r.json().get('id', 0))
                print(f"    Parte: {party_data.get('nome', '?')} -> id={pid}")
                return pid
            except (ValueError, AttributeError) as e:
                print(f"    [ERRO] Resposta inesperada ao criar parte: {r.text[:200]}")
                return None
        else:
            print(f"    [ERRO] Parte {party_data.get('nome', '?')}: {r.status_code} {r.text[:200]}")
            return None

    def get_or_create_inss(self):
        """Get INSS party ID (cached)."""
        if self._inss_id:
            return self._inss_id

        # Try to find existing
        r = self._request("get", "/parts")
        if r.status_code == 200:
            for p in r.json():
                doc = str(p.get('documento', ''))
                if '29.979.036' in doc or '29979036' in doc:
                    self._inss_id = int(p.get('id', 0))
                    return self._inss_id

        # Create
        pid = self.criar_parte(INSS_DATA.copy())
        if pid:
            self._inss_id = pid
        return self._inss_id

    def vincular_partes(self, idpet, polo_ativo_ids, polo_passivo_ids=None):
        """Set parties on petition."""
        dados = self.get_peticao(idpet)
        dados['idpoloativo'] = polo_ativo_ids if isinstance(polo_ativo_ids, list) else [polo_ativo_ids]
        if polo_passivo_ids:
            dados['idpolopassivo'] = polo_passivo_ids if isinstance(polo_passivo_ids, list) else [polo_passivo_ids]
        return self.put_peticao(idpet, dados)

    # ---- File uploads ----

    def upload_peticao_pdf(self, idpet, idproc, pdf_path):
        """Upload main petition PDF."""
        with open(pdf_path, 'rb') as f:
            r = self._request("post",
                f"/complaintsandpleadings/file?idpeticoes={idpet}&idprocessos={idproc}",
                files={'file': (os.path.basename(pdf_path), f, 'application/pdf')},
                timeout=60)
        if r.status_code == 200:
            try:
                body = r.json()
                if body.get('success') is False or body.get('status') == 'error':
                    print(f"    PDF principal ERRO (API): {body.get('message', body)}")
                    return False
            except (ValueError, AttributeError):
                pass  # Resposta nao-JSON — aceitar se 200
            print(f"    PDF principal: OK")
            return True
        print(f"    PDF principal ERRO: {r.status_code}")
        return False

    def upload_anexo(self, idpet, pdf_path, doc_type_id):
        """Upload attachment PDF."""
        with open(pdf_path, 'rb') as f:
            r = self._request("post",
                f"/complaintsandpleadings/attachments?idpeticoes={idpet}&fk_documentos_tipos={doc_type_id}",
                files={'file': (os.path.basename(pdf_path), f, 'application/pdf')},
                timeout=60)
        if r.status_code == 200:
            try:
                body = r.json()
                if body.get('success') is False or body.get('status') == 'error':
                    return False
            except (ValueError, AttributeError):
                pass
            return True
        return False

    def get_tipos_anexo(self, idpet):
        """Get available attachment types for this petition."""
        r = self._request("get", f"/complaintsandpleadings/attachment/types?idpeticoes={idpet}")
        if r.status_code == 200:
            return r.json()
        return []

    def resolver_tipo_anexo(self, idpet, prefix):
        """Match file prefix to attachment type ID."""
        tipos = self.get_tipos_anexo(idpet)
        if not tipos:
            return None

        target_name = DOC_PREFIX_MAP.get(prefix, '')
        if not target_name:
            return None

        # Build lookup map
        tipo_map = {}
        for t in tipos:
            nome = t.get('nome', '').upper()
            tid = t.get('iddocumentos_tipos')
            tipo_map[nome] = tid

        # Exact match
        for nome, tid in tipo_map.items():
            if target_name.upper() in nome:
                return tid

        # Partial match
        first_word = target_name.split()[0].upper()
        for nome, tid in tipo_map.items():
            if first_word in nome:
                return tid

        # Fallback: "Outros" (fix #15 — different ID per tribunal)
        for nome, tid in tipo_map.items():
            if 'OUTRO' in nome:
                return tid
        # Last resort: first available type
        return tipos[0].get('iddocumentos_tipos') if tipos else None

    def upload_todos_anexos(self, idpet, idproc, pasta):
        """Upload all documents from organized folder.

        Usa ordem correta (ORDEM_DOCUMENTOS), valida PDFs,
        exclui adjudicação administrativa, e trata planilha/quesitos.
        """
        uploaded = 0
        errors = []
        pendencias = []

        # Listar e ordenar todos os PDFs
        todos_pdfs = sorted([f for f in os.listdir(pasta) if f.endswith('.pdf')])
        todos_pdfs = ordenar_documentos(todos_pdfs)

        # Checkpoint: pular docs ja uploaded (resume apos falha)
        upload_log = os.path.join(pasta, '_upload_progress.json')
        already_uploaded = set()
        if os.path.exists(upload_log):
            try:
                with open(upload_log, encoding='utf-8') as uf:
                    already_uploaded = set(json.load(uf).get('uploaded_files', []))
                if already_uploaded:
                    print(f"    [RESUME] {len(already_uploaded)} docs ja uploaded, pulando")
            except Exception:
                pass

        for f in todos_pdfs:
            if f in already_uploaded:
                uploaded += 1
                continue
            filepath = os.path.join(pasta, f)
            nome_upper = f.upper()

            # Validar PDF antes de subir
            valido, msg_validacao = validar_pdf(filepath)
            if not valido:
                errors.append(f"PDF INVALIDO: {f} — {msg_validacao}")
                pendencias.append({"arquivo": f, "tipo": "PDF_INVALIDO", "detalhe": msg_validacao})
                continue

            # Excluir adjudicação administrativa em casos BPC/LOAS
            if "ADJUDICACAO" in nome_upper or "ADJUDICAÇÃO" in nome_upper:
                print(f"    [SKIP] Adjudicacao excluida: {f}")
                continue

            # Detectar prefixo ANTES de usar (fix bug #5)
            match = re.match(r'^(\d+)', f)
            prefix = match.group(1) if match else None

            # Procuração: só ad judicia (assinada). Excluir administrativa.
            if prefix == "2" and ("ADMINISTRATIVA" in nome_upper or "ADMIN" in nome_upper):
                print(f"    [SKIP] Procuracao administrativa excluida: {f}")
                continue

            # Tratar planilha e quesitos (sem prefixo numerico)
            if "PLANILHA" in nome_upper:
                prefix = "PLANILHA"
            elif "QUESITOS" in nome_upper and "MEDIC" in nome_upper:
                prefix = "QUESITOS"
            elif "QUESITOS" in nome_upper and "SOCIAL" in nome_upper:
                prefix = "QUESITOS"
            elif "QUESITOS" in nome_upper:
                prefix = "QUESITOS"

            if prefix == "1":
                # Petição principal — validação extra
                if self.upload_peticao_pdf(idpet, idproc, filepath):
                    uploaded += 1
                    print(f"    [OK] Petição principal: {f}")
                else:
                    errors.append(f"Petição principal: {f}")
            elif prefix == "2":
                # Procuração — validar assinatura
                tem_assinatura, msg_assinatura = validar_procuracao(filepath)
                if not tem_assinatura:
                    errors.append(f"PROCURACAO SEM ASSINATURA: {f} — {msg_assinatura}")
                    pendencias.append({"arquivo": f, "tipo": "PROCURACAO_SEM_ASSINATURA", "detalhe": msg_assinatura})
                    print(f"    [WARN] Procuração sem assinatura: {f} — subindo mesmo assim")
                # Subir mesmo sem assinatura (flag no pendencias)
                doc_type_id = self.resolver_tipo_anexo(idpet, "2")
                if doc_type_id and self.upload_anexo(idpet, filepath, doc_type_id):
                    uploaded += 1
                    print(f"    [OK] Procuração: {f}")
                else:
                    errors.append(f"Procuração upload falhou: {f}")
            elif prefix:
                doc_type_id = self.resolver_tipo_anexo(idpet, prefix)
                if doc_type_id:
                    if self.upload_anexo(idpet, filepath, doc_type_id):
                        uploaded += 1
                        print(f"    [OK] Anexo: {f} (tipo: {DOC_PREFIX_MAP.get(prefix, '?')})")
                    else:
                        errors.append(f"Anexo upload falhou: {f}")
                else:
                    errors.append(f"Tipo não encontrado: {f} (prefix={prefix})")
            else:
                errors.append(f"Arquivo sem prefixo reconhecido: {f}")

        # Salvar checkpoint de upload (pra resume em caso de falha)
        try:
            uploaded_files = [f for f in todos_pdfs if f not in [e.split(': ')[-1] for e in errors]]
            with open(upload_log, 'w', encoding='utf-8') as uf:
                json.dump({'uploaded_files': list(already_uploaded | set(uploaded_files)),
                           'idpet': idpet, 'total': len(todos_pdfs)}, uf)
        except Exception:
            pass

        # Salvar pendencias se houver
        if pendencias:
            pendencias_file = os.path.join(pasta, "PENDENCIAS_UPLOAD.json")
            try:
                with open(pendencias_file, 'w', encoding='utf-8') as pf:
                    json.dump({"pendencias": pendencias, "data": time.strftime("%Y-%m-%d %H:%M")}, pf, ensure_ascii=False, indent=2)
                print(f"    [WARN] {len(pendencias)} pendencias salvas em PENDENCIAS_UPLOAD.json")
            except Exception as e:
                print(f"    [ERRO] Falha ao salvar pendencias: {e}")

        return uploaded, errors

    # ---- Valor da causa extraction ----

    def extrair_valor_causa(self, pasta):
        """Extract valor da causa from planilha Excel.

        Strategies (in order):
        1. Numeric cell next to "TOTAL DA CONTA" or "TOTAL GERAL"
        2. Value embedded in cell text (e.g. "TOTAL DA CONTA: R$ 27.507,43")
        3. Last large numeric value in any sheet
        """
        if not openpyxl:
            print("    [WARN] openpyxl nao disponivel, valor nao extraido")
            return None

        xlsx_files = sorted([f for f in os.listdir(pasta) if f.endswith('.xlsx') and 'calculo' in f.lower()])
        if not xlsx_files:
            xlsx_files = sorted([f for f in os.listdir(pasta) if f.endswith('.xlsx')])
        if not xlsx_files:
            return None

        # Preferir arquivo mais recente por data de modificacao
        target = max(xlsx_files, key=lambda f: os.path.getmtime(os.path.join(pasta, f)))
        wb = None
        try:
            wb = openpyxl.load_workbook(os.path.join(pasta, target), data_only=True)

            # Search ALL sheets, not just active
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        val = str(cell.value or '').upper()
                        if ('TOTAL' in val and ('CONTA' in val or 'GERAL' in val or 'CAUSA' in val)
                                and 'SUB' not in val):
                            # Strategy 1: Numeric cell to the right
                            for off in range(1, 8):
                                try:
                                    right = ws.cell(row=cell.row, column=cell.column + off).value
                                    if right and self._parse_brl(right):
                                        return self._parse_brl(right)
                                except Exception:
                                    continue
                            # Strategy 2: Value embedded in the cell text itself
                            # e.g. "TOTAL DA CONTA EM 03/2026: R$ 27.507,43"
                            embedded = self._parse_brl(cell.value)
                            if embedded and embedded > 100:
                                return embedded

            # Strategy 3: Last large numeric value in any sheet
            for ws in wb.worksheets:
                for row in reversed(list(ws.iter_rows(values_only=False))):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and 1000 < cell.value < 500000:
                            return round(float(cell.value), 2)

        except Exception as e:
            print(f"    [WARN] Erro extraindo valor: {e}")
        finally:
            if wb:
                try: wb.close()
                except Exception: pass
        return None

    def _parse_brl(self, val):
        """Parse Brazilian Real value from various formats.
        Handles: 27507.43, "R$ 27.507,43", "TOTAL: R$ 27.507,43", etc."""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return round(float(val), 2)
        s = str(val)
        # Extract monetary value with regex (handles embedded values in text)
        m = re.search(r'R\$\s*([\d.,]+)', s)
        if m:
            s = m.group(1)
        else:
            # Try to find a standalone number
            s = s.replace('R$', '').strip()
            # If string has lots of non-numeric chars, try to extract just the number
            m2 = re.search(r'([\d.,]+)', s)
            if m2:
                s = m2.group(1)
            else:
                return None
        s = s.strip()
        if not s:
            return None
        # Parse Brazilian number format: 27.507,43 -> 27507.43
        if ',' in s and '.' in s:
            if s.rindex(',') > s.rindex('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            parts = s.split(',')
            if len(parts[-1]) <= 2:
                s = s.replace(',', '.')
            else:
                s = s.replace(',', '')
        try:
            return round(float(s), 2)
        except ValueError:
            return None

    # ---- DOCX to PDF conversion ----

    def converter_docx_para_pdf(self, docx_path):
        """Convert .docx to .pdf using LibreOffice or Word COM."""
        import subprocess
        pdf_path = os.path.splitext(docx_path)[0] + '.pdf'
        if os.path.exists(pdf_path):
            return pdf_path

        # Try LibreOffice first (more reliable)
        try:
            result = subprocess.run([
                'soffice', '--headless', '--convert-to', 'pdf',
                '--outdir', os.path.dirname(docx_path), docx_path
            ], capture_output=True, text=True, timeout=60)
            if os.path.exists(pdf_path):
                return pdf_path
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

        # Fallback: Word COM via PowerShell
        try:
            abs_path = os.path.abspath(docx_path).replace('/', '\\')
            abs_pdf = os.path.splitext(abs_path)[0] + '.pdf'
            # Escapar aspas simples no path pra evitar injection
            safe_path = abs_path.replace("'", "''")
            safe_pdf = abs_pdf.replace("'", "''")
            ps_script = f"""
$word = New-Object -ComObject Word.Application
$word.Visible = $false
$word.DisplayAlerts = 0
try {{
    $doc = $word.Documents.Open('{safe_path}')
    $doc.SaveAs('{safe_pdf}', 17)
    $doc.Close()
}} finally {{
    $word.Quit()
}}
"""
            import tempfile
            ps_file = os.path.join(tempfile.gettempdir(), '_docx2pdf.ps1')
            with open(ps_file, 'w', encoding='utf-8-sig') as f:
                f.write(ps_script)
            subprocess.run(['powershell', '-ExecutionPolicy', 'Bypass', '-File', ps_file],
                          capture_output=True, timeout=60)
            if os.path.exists(pdf_path):
                return pdf_path
        except Exception:
            pass

        return None

    # ---- MAIN: Process one client (one command, everything) ----

    def processar_cliente(self, pasta, client_data=None, extracted=None, dry_run=False):
        """Process one client folder: create draft, fill fields, upload docs.

        Args:
            pasta: organized client folder path
            client_data: dict with party data (nome, documento/CPF, endereco_*)
                If has 'representante' key, creates 2 parties (mãe + menor)
            extracted: dict from analisar_pasta_internal (optional)
            dry_run: if True, create draft and fill fields but skip document upload

        Returns: dict with status, idpeticoes, url, details
        """
        nome_pasta = os.path.basename(pasta)
        print(f"\n{'='*60}")
        print(f"  {nome_pasta}")
        print(f"{'='*60}")

        result = {'pasta': nome_pasta, 'status': 'erro', 'fill_errors': [], 'filled': []}

        # 1. Detect UF and tribunal
        from app import detect_uf_from_folder
        uf = None
        cidade = ''

        if extracted:
            cidade = extracted.get('cidade', '') or ''
            uf = extracted.get('uf', '')

        if not uf:
            uf = detect_uf_from_folder(pasta)
        if not uf and client_data:
            uf = client_data.get('endereco_uf', '')
        if not uf:
            result['error'] = 'UF não detectada'
            return result

        cfg = UF_TRIBUNAL_MAP.get(uf)
        if not cfg:
            result['error'] = f'UF {uf} sem mapeamento de tribunal'
            return result

        trf = cfg['trf']
        sistema = cfg['sistema']
        print(f"    Tribunal: {trf}/{sistema} (UF={uf})")

        # 2. Get cidade from client data or CEP
        if not cidade and client_data:
            cidade = client_data.get('endereco_cidade', '')
            if not cidade and client_data.get('endereco_cep'):
                addr = completar_endereco_por_cep(client_data['endereco_cep'])
                if addr:
                    cidade = addr.get('endereco_cidade', '')
                    client_data.update({k: v for k, v in addr.items() if v and not client_data.get(k)})

        # 3. Create petition draft
        try:
            idpet, idproc = self.criar_peticao(trf, sistema, "1", cfg['uf_tribunal'])
        except Exception as e:
            result['error'] = str(e)
            return result

        result['idpeticoes'] = idpet
        result['idprocessos'] = idproc
        result['url'] = f"{LEGALMAIL_SITE}/petitions/{idpet}"

        # 4. Extract valor da causa
        valor = self.extrair_valor_causa(pasta)
        if valor:
            print(f"    Valor da causa: R$ {valor:,.2f}")

        # 5. Create/link parties
        inss_id = self.get_or_create_inss()
        polo_ativo_ids = []

        if client_data:
            rep_data = client_data.get('representante')
            menor_data = client_data.get('menor')

            if rep_data and menor_data:
                # MENOR PRIMEIRO, representante SEGUNDO (regra protocolo BPC)
                # O LegalMail usa o primeiro nome do polo ativo como titulo
                # Fix CPF vazio: buscar na peticao docx se nao veio do OCR
                if not rep_data.get('documento') and not rep_data.get('cpf'):
                    cpf_found = self._extrair_cpf_da_peticao(pasta, rep_data.get('nome', ''))
                    if cpf_found:
                        rep_data['documento'] = cpf_found
                        print(f"    [CPF] Representante extraido da peticao: {cpf_found}")
                    else:
                        print(f"    [ERRO] CPF representante nao encontrado! Parte pode falhar.")
                        result['fill_errors'].append('CPF_REP_MISSING')
                if not menor_data.get('documento') and not menor_data.get('cpf'):
                    cpf_found = self._extrair_cpf_da_peticao(pasta, menor_data.get('nome', ''))
                    if cpf_found:
                        menor_data['documento'] = cpf_found
                        print(f"    [CPF] Menor extraido da peticao: {cpf_found}")
                menor_id = self.criar_parte(self._prepare_party(menor_data, 'ativo', 'DESEMPREGADO'))
                rep_id = self.criar_parte(self._prepare_party(rep_data, 'ativo', 'DO LAR\\DONA DE CASA'))
                if menor_id:
                    polo_ativo_ids.append(menor_id)
                else:
                    print(f"    [ERRO] Falha ao criar parte MENOR: {menor_data.get('nome', '?')}")
                    result['fill_errors'].append('MENOR_CREATION_FAILED')
                if rep_id:
                    polo_ativo_ids.append(rep_id)
                else:
                    print(f"    [ERRO] Falha ao criar parte REPRESENTANTE: {rep_data.get('nome', '?')}")
                    result['fill_errors'].append('REP_CREATION_FAILED')
            elif client_data.get('nome') and client_data.get('documento'):
                # Single party (adult like Denise, Roseli)
                pid = self.criar_parte(self._prepare_party(client_data, 'ativo'))
                if pid:
                    polo_ativo_ids.append(pid)

        # 6. Fill all fields (detect tipo: deficiente vs idoso)
        tipo_bpc = 'deficiente'  # default
        if extracted:
            dn = extracted.get('data_nascimento', '')
            if dn:
                try:
                    nasc = _datetime.strptime(dn, '%Y-%m-%d')
                    idade = (_datetime.now() - nasc).days // 365
                    if idade >= 65:  # BPC-LOAS idoso: 65+ anos (Art. 20, Lei 8.742/93)
                        tipo_bpc = 'idoso'
                except (ValueError, TypeError):
                    pass  # Data de nascimento invalida — usar default 'deficiente'
        fill_result = self.preencher_campos(idpet, cidade, uf, tipo=tipo_bpc)
        result['filled'] = fill_result['filled']
        result['fill_errors'] = fill_result['errors']

        # 7. Set valor da causa + parties (separate PUT to avoid field conflicts)
        try:
            dados = self.get_peticao(idpet)
            if valor and valor > 0:
                dados['valorCausa'] = valor
            elif valor is not None and valor <= 0:
                print(f"    [WARN] Valor da causa invalido: {valor} — ignorado")
                result['fill_errors'].append(f'VALOR_INVALIDO: {valor}')
            if polo_ativo_ids:
                dados['idpoloativo'] = polo_ativo_ids
            if inss_id:
                dados['idpolopassivo'] = [inss_id]
            self.put_peticao(idpet, dados)
        except Exception as e:
            result['fill_errors'].append(f"valor/partes: {e}")

        # 8. Convert docx → PDF if needed
        for f in os.listdir(pasta):
            if f.startswith('1- PETICAO') and f.endswith('.docx'):
                pdf_path = os.path.join(pasta, f.replace('.docx', '.pdf'))
                if not os.path.exists(pdf_path):
                    print(f"    Convertendo {f} → PDF...")
                    self.converter_docx_para_pdf(os.path.join(pasta, f))

        # 9. Upload all documents (skip in dry_run mode)
        if dry_run:
            print(f"    [DRY-RUN] Upload pulado — rascunho criado sem documentos")
            result['uploaded'] = 0
        else:
            uploaded, upload_errors = self.upload_todos_anexos(idpet, idproc, pasta)
            result['uploaded'] = uploaded
            if upload_errors:
                result['upload_errors'] = upload_errors

        # 10. VALIDACAO FINAL — checar TODOS os campos obrigatorios
        try:
            final = self.get_peticao(idpet)
            validacao = []
            if not final.get('tribunal'):
                validacao.append("SEM_TRIBUNAL")
            if not final.get('idpoloativo'):
                validacao.append("SEM_POLO_ATIVO")
            elif len(final.get('idpoloativo', [])) < 2 and (client_data and client_data.get('representante')):
                validacao.append("POLO_INCOMPLETO")
            if not final.get('idpolopassivo') or (inss_id and inss_id not in (final.get('idpolopassivo') or [])):
                validacao.append("SEM_INSS")
            if not final.get('valorCausa'):
                validacao.append("SEM_VALOR")
            if not final.get('comarca'):
                validacao.append("SEM_COMARCA")
            if not final.get('classe'):
                validacao.append("SEM_CLASSE")
            if not final.get('assunto'):
                validacao.append("SEM_ASSUNTO")
            if not final.get('gratuidade'):
                validacao.append("SEM_GRATUIDADE")
            is_eproc = 'eproc' in (final.get('sistema', '') or '').lower()
            if is_eproc and not final.get('rito'):
                validacao.append("SEM_RITO")
            if not is_eproc and not final.get('competencia'):
                validacao.append("SEM_COMPETENCIA")
            # Verificar ordem do polo: menor deve ser primeiro
            polo_a = final.get('idpoloativo', [])
            if len(polo_a) >= 2 and polo_ativo_ids and polo_a != polo_ativo_ids:
                print(f"    [WARN] Polo ordem diferente do esperado: {polo_a} vs {polo_ativo_ids}")
                # Forcar ordem correta
                dados_fix = self.get_peticao(idpet)
                dados_fix['idpoloativo'] = polo_ativo_ids
                self.put_peticao(idpet, dados_fix)
                print(f"    [FIX] Polo reordenado para {polo_ativo_ids}")

            if validacao:
                result['validacao_erros'] = validacao
                print(f"    [VALIDACAO] PROBLEMAS: {', '.join(validacao)}")
            else:
                print(f"    [VALIDACAO] PRONTO PARA PROTOCOLAR")
        except Exception as e:
            print(f"    [VALIDACAO] Erro: {e}")

        result['status'] = 'ok'
        print(f"    RESULTADO: {len(fill_result['filled'])} campos | {result.get('uploaded', 0)} docs | {len(fill_result['errors'])} erros")
        return result

    def _extrair_cpf_da_peticao(self, pasta, nome_busca):
        """Buscar CPF de uma pessoa na peticao docx gerada.
        Aceita CPF formatado (XXX.XXX.XXX-XX) e nao formatado (XXXXXXXXXXX).
        Retorna CPF formatado ou None."""
        try:
            from docx import Document
            for f in sorted(os.listdir(pasta)):
                if ('peticao' in f.lower() or 'PETICAO' in f) and f.endswith('.docx'):
                    doc = Document(os.path.join(pasta, f))
                    texto = '\n'.join([p.text for p in doc.paragraphs])
                    # Buscar CPFs formatados E nao formatados
                    cpfs_fmt = re.findall(r'\d{3}\.\d{3}\.\d{3}-\d{2}', texto)
                    cpfs_raw = re.findall(r'(?<!\d)(\d{11})(?!\d)', texto)
                    # Formatar os raw
                    for raw in cpfs_raw:
                        fmt = f"{raw[:3]}.{raw[3:6]}.{raw[6:9]}-{raw[9:]}"
                        if fmt not in cpfs_fmt:
                            cpfs_fmt.append(fmt)
                    cpfs = cpfs_fmt
                    if not cpfs:
                        print(f"    [CPF] Nenhum CPF encontrado em {f}")
                        return None
                    # Se so tem 1 CPF, retorna ele
                    if len(cpfs) == 1:
                        return cpfs[0]
                    # Se tem nome_busca, pegar o CPF mais proximo do nome
                    if nome_busca:
                        nome_upper = nome_busca.upper()
                        # Buscar por primeiro nome tambem (ex: "Renata" em vez de "Renata Laurinda da Costa")
                        primeiro_nome = nome_upper.split()[0] if nome_upper else ''
                        pos_nome = texto.upper().find(nome_upper)
                        if pos_nome < 0 and primeiro_nome:
                            pos_nome = texto.upper().find(primeiro_nome)
                        if pos_nome >= 0:
                            melhor = None
                            melhor_dist = 99999
                            for cpf in cpfs:
                                pos_cpf = texto.find(cpf)
                                if pos_cpf < 0:
                                    # Buscar versao sem formatacao
                                    raw = cpf.replace('.', '').replace('-', '')
                                    pos_cpf = texto.find(raw)
                                dist = abs(pos_cpf - pos_nome) if pos_cpf >= 0 else 99999
                                if dist < melhor_dist:
                                    melhor_dist = dist
                                    melhor = cpf
                            if melhor:
                                return melhor
                    # Fallback: CPFs unicos, retorna o ultimo (geralmente representante)
                    unicos = list(dict.fromkeys(cpfs))
                    return unicos[-1] if len(unicos) > 1 else unicos[0]
        except Exception as e:
            print(f"    [CPF] Erro extraindo da peticao: {e}")
        return None

    def _prepare_party(self, data, polo, profissao_default=None):
        """Prepare party data dict with all required fields."""
        party = {
            'nome': (data.get('nome', '') or '').upper(),
            'documento': data.get('documento', '') or data.get('cpf', '') or '',
            'personalidade': 'Pessoa física',
            'polo': polo,
            'endereco_cep': data.get('endereco_cep', '') or data.get('cep', '') or '',
            'endereco_logradouro': data.get('endereco_logradouro', '') or data.get('logradouro', '') or '',
            'endereco_numero': data.get('endereco_numero', '') or data.get('numero', '') or 'S/N',
            'endereco_bairro': data.get('endereco_bairro', '') or data.get('bairro', '') or '',
            'endereco_cidade': data.get('endereco_cidade', '') or data.get('cidade', '') or '',
            'endereco_uf': data.get('endereco_uf', '') or data.get('uf', '') or '',
        }

        # Profissao — must match exact API values (case sensitive)
        # Mapa COMPLETO de profissoes aceitas pelo LegalMail (case-sensitive!)
        PROF_MAP = {
            'DESEMPREGADO': 'Desempregado',
            'DESEMPREGADA': 'Desempregado',
            'DO LAR': 'DO LAR\\DONA DE CASA',
            'DO LAR\\DONA DE CASA': 'DO LAR\\DONA DE CASA',
            'DONA DE CASA': 'DO LAR\\DONA DE CASA',
            'DOMESTICA': 'DO LAR\\DONA DE CASA',
            'ESTUDANTE': 'Estudante',
            'MENOR': 'Desempregado',  # Menor = Desempregado (cd=548)
            'SEM PROFISSAO': 'Sem Profissão Definida',
            'SEM PROFISSÃO DEFINIDA': 'Sem Profissão Definida',
            'APOSENTADO': 'Aposentado',
            'APOSENTADA': 'Aposentado',
            'AUTONOMO': 'Autônomo',
            'AUTONOMA': 'Autônomo',
            'AGRICULTORA': 'Agricultor',
            'AGRICULTOR': 'Agricultor',
            'LAVRADOR': 'Agricultor',
            'LAVRADORA': 'Agricultor',
            'VENDEDOR': 'Vendedor',
            'VENDEDORA': 'Vendedor',
            'DIARISTA': 'Diarista',
            'SERVENTE': 'Servente',
            'OPERADOR': 'Operador',
            'MOTORISTA': 'Motorista',
            'COZINHEIRA': 'Cozinheiro',
            'COZINHEIRO': 'Cozinheiro',
        }
        prof = data.get('profissao', '') or ''
        if not prof and profissao_default:
            prof = profissao_default
        if prof:
            mapped = PROF_MAP.get(prof.upper().strip())
            if mapped:
                party['profissao'] = mapped
            else:
                # Profissao desconhecida — usar como esta mas avisar
                print(f"    [WARN] Profissao '{prof}' nao mapeada, usando como esta")
                party['profissao'] = prof

        # Etnia (TRF-2 eProc requires it) — padronizado com acento
        party['etnia'] = data.get('etnia', 'Não declarada')
        # Normalizar variantes sem acento
        if party['etnia'] == 'Nao declarada':
            party['etnia'] = 'Não declarada'

        return party


# ============================================================
# STANDALONE: Fix existing drafts
# ============================================================
def fix_rascunho(svc, idpet, cidade=None, uf=None, polo_ativo_ids=None):
    """Fix an existing draft: fill missing fields, link parties."""
    print(f"\n  Corrigindo rascunho {idpet}...")
    dados = svc.get_peticao(idpet)

    # Detect cidade/uf from existing data if not provided
    if not cidade:
        comarca = dados.get('comarca', '') or ''
        # Try to extract city from comarca name
        m = re.search(r'(?:de|DE)\s+(.+?)(?:\s*\(|$|-)', comarca)
        if m:
            cidade = m.group(1).strip()

    result = {'idpet': idpet}

    # Fill missing fields
    if uf or cidade:
        fill = svc.preencher_campos(idpet, cidade, uf or '')
        result['filled'] = fill['filled']
        result['errors'] = fill['errors']

    # Link parties if provided
    if polo_ativo_ids:
        svc.vincular_partes(idpet, polo_ativo_ids)
        result['partes'] = True

    return result
