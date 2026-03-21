import os

# Load .env file (ANTHROPIC_API_KEY always overwritten from .env to avoid stale system key)
env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                k = k.strip()
                if k == "ANTHROPIC_API_KEY" or k not in os.environ:
                    os.environ[k] = v.strip()

# Clean API key (remove leading = or spaces from Railway env)
api_key = os.environ.get("ANTHROPIC_API_KEY", "")
if api_key and not api_key.startswith("sk-"):
    cleaned = api_key.lstrip(" =")
    os.environ["ANTHROPIC_API_KEY"] = cleaned

import re
import json
import glob
import base64
import hashlib
import hmac
import secrets
import traceback
import tempfile
import requests
import time as _time_module
from functools import wraps
from collections import defaultdict
from flask import Flask, render_template, request, jsonify, send_from_directory

import shutil
import zipfile
import anthropic
import httpx as _httpx
import fitz  # PyMuPDF
from PIL import Image
import openpyxl


# ==================== AI FALLBACK SYSTEM ====================
# Tries: Anthropic (Haiku/Sonnet) -> Z.AI GLM -> Google Gemini
# Ensures Ana NEVER stops responding even if Anthropic is down

_GLM_API_KEY = os.environ.get("GLM_API_KEY", "")
_GEMINI_API_KEY = os.environ.get("GOOGLE_API_KEY", "") or os.environ.get("GEMINI_API_KEY", "")

def _call_glm(messages, system=None, max_tokens=1500, model="glm-4-plus"):
    """Z.AI GLM fallback."""
    if not _GLM_API_KEY:
        return None
    try:
        msgs = []
        if system:
            msgs.append({"role": "system", "content": system})
        msgs.extend(messages)
        with _httpx.Client(timeout=30) as client:
            r = client.post(
                "https://open.bigmodel.cn/api/paas/v4/chat/completions",
                headers={"Authorization": f"Bearer {_GLM_API_KEY}", "Content-Type": "application/json"},
                json={"model": model, "messages": msgs, "max_tokens": max_tokens},
            )
            if r.status_code == 200:
                return r.json()["choices"][0]["message"]["content"]
    except Exception:
        pass
    return None

def _call_gemini(messages, system=None, max_tokens=1500):
    """Google Gemini fallback."""
    if not _GEMINI_API_KEY:
        return None
    try:
        # Convert messages to Gemini format
        parts = []
        if system:
            parts.append({"text": f"System: {system}\n\n"})
        for m in messages:
            prefix = "User: " if m["role"] == "user" else "Assistant: "
            parts.append({"text": prefix + m["content"]})
        with _httpx.Client(timeout=30) as client:
            r = client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={_GEMINI_API_KEY}",
                json={"contents": [{"parts": parts}], "generationConfig": {"maxOutputTokens": max_tokens}},
            )
            if r.status_code == 200:
                return r.json()["candidates"][0]["content"]["parts"][0]["text"]
    except Exception:
        pass
    return None

def ai_chat(messages, system=None, model="claude-haiku-4-5-20251001", max_tokens=1500, api_key=None):
    """
    Universal AI call with automatic fallback.
    Tries: Anthropic -> Z.AI GLM -> Google Gemini
    """
    import logging
    logger = logging.getLogger("ai_fallback")

    # 1. Try Anthropic (primary)
    try:
        kwargs = {"timeout": 30.0}
        if api_key:
            kwargs["api_key"] = api_key
        client = anthropic.Anthropic(**kwargs)
        params = {"model": model, "max_tokens": max_tokens, "messages": messages}
        if system:
            params["system"] = system
        response = client.messages.create(**params)
        return response.content[0].text
    except Exception as e:
        err = str(e)
        logger.warning(f"[AI] Anthropic failed ({model}): {err[:100]}")

    # 2. Try Z.AI GLM
    try:
        result = _call_glm(messages, system=system, max_tokens=max_tokens)
        if result:
            logger.info("[AI] Fallback to GLM succeeded")
            return result
    except Exception as e:
        logger.warning(f"[AI] GLM failed: {e}")

    # 3. Try Google Gemini
    try:
        result = _call_gemini(messages, system=system, max_tokens=max_tokens)
        if result:
            logger.info("[AI] Fallback to Gemini succeeded")
            return result
    except Exception as e:
        logger.warning(f"[AI] Gemini failed: {e}")

    logger.error("[AI] ALL PROVIDERS FAILED")
    return None

# Import all helper functions
import helpers

app = Flask(__name__)

# ==================== SECURITY ====================

# Debug/admin token from environment (NEVER hardcoded)
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", secrets.token_urlsafe(32))
if not os.environ.get("ADMIN_TOKEN"):
    print(f"[SECURITY] ADMIN_TOKEN não configurado. Token temporário gerado: {ADMIN_TOKEN}")
    print(f"[SECURITY] Configure ADMIN_TOKEN no Railway para um token fixo.")

# Webhook secret for validating incoming webhooks
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET", "")

def _check_admin_token():
    """Validate admin token from query param or header."""
    token = request.args.get("token", "") or request.headers.get("X-Admin-Token", "")
    if not token or not hmac.compare_digest(token, ADMIN_TOKEN):
        return False
    return True

def require_admin(f):
    """Decorator to require admin token on endpoints."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not _check_admin_token():
            return jsonify({"error": "unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

# Rate limiting (in-memory, per IP)
_rate_limit_store = defaultdict(list)
_rate_limit_lock = __import__('threading').Lock()

def _rate_limit_check(key, max_requests=30, window_seconds=60):
    """Simple rate limiter. Returns True if request is allowed."""
    now = _time_module.time()
    with _rate_limit_lock:
        _rate_limit_store[key] = [t for t in _rate_limit_store[key] if now - t < window_seconds]
        if len(_rate_limit_store[key]) >= max_requests:
            return False
        _rate_limit_store[key].append(now)
        return True

@app.before_request
def _global_rate_limit():
    """Global rate limit: 60 requests/minute per IP for API endpoints."""
    if request.path.startswith("/api/"):
        ip = request.remote_addr or "unknown"
        if not _rate_limit_check(f"global:{ip}", max_requests=60, window_seconds=60):
            return jsonify({"error": "rate limit exceeded"}), 429

# Atomic JSON file operations with backup
def _safe_json_save(filepath, data, lock=None):
    """Save JSON atomically: write to .tmp, then rename. Creates .bak backup."""
    def _do_save():
        tmp_path = filepath + ".tmp"
        bak_path = filepath + ".bak"
        try:
            # Backup existing file
            if os.path.exists(filepath):
                shutil.copy2(filepath, bak_path)
            # Write to temp file first
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            # Atomic rename (on Windows, need to remove first)
            if os.path.exists(filepath):
                os.replace(tmp_path, filepath)
            else:
                os.rename(tmp_path, filepath)
        except Exception as e:
            print(f"[SECURITY] Erro ao salvar {filepath}: {e}")
            # Try to restore from backup
            if os.path.exists(bak_path) and not os.path.exists(filepath):
                shutil.copy2(bak_path, filepath)
            raise
    if lock:
        with lock:
            _do_save()
    else:
        _do_save()

def _safe_json_load(filepath, lock=None):
    """Load JSON with fallback to .bak if corrupted."""
    def _do_load():
        try:
            if os.path.exists(filepath):
                with open(filepath, "r", encoding="utf-8") as f:
                    return json.load(f)
        except (json.JSONDecodeError, Exception) as e:
            print(f"[SECURITY] Arquivo corrompido {filepath}: {e}")
            bak_path = filepath + ".bak"
            if os.path.exists(bak_path):
                print(f"[SECURITY] Restaurando backup {bak_path}")
                try:
                    with open(bak_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    shutil.copy2(bak_path, filepath)
                    return data
                except Exception as e2:
                    print(f"[SECURITY] Backup também corrompido: {e2}")
        return {}
    if lock:
        with lock:
            return _do_load()
    else:
        return _do_load()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
TIMBRADO_PATH = os.path.join(BASE_DIR, "timbrado.docx")
SKILL_PROMPT_PATH = os.path.join(BASE_DIR, "skill_prompt.md")

os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_skill_prompt():
    with open(SKILL_PROMPT_PATH, "r", encoding="utf-8") as f:
        return f.read()


@app.route("/")
def index():
    return render_template("index.html")


def build_exec_globals():
    """Build the exec environment with all helpers pre-loaded."""
    import datetime as dt_module
    import openpyxl as openpyxl_module
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Restricted __import__ — only allow safe modules
    _ALLOWED_MODULES = {
        'datetime', 'decimal', 'math', 'string', 're', 'json', 'os', 'os.path',
        'collections', 'itertools', 'functools', 'textwrap', 'copy',
        'calendar', '_strptime', 'locale', 'time',
        'dateutil', 'dateutil.relativedelta', 'dateutil.parser',
        'openpyxl', 'openpyxl.styles', 'openpyxl.utils',
    }
    _original_import = __builtins__.__import__ if hasattr(__builtins__, '__import__') else __import__

    def _safe_import(name, *args, **kwargs):
        if name.split('.')[0] not in _ALLOWED_MODULES and name not in _ALLOWED_MODULES:
            raise ImportError(f"Import bloqueado por segurança: {name}")
        return _original_import(name, *args, **kwargs)

    g = {
        "__builtins__": {**(__builtins__.__dict__ if hasattr(__builtins__, '__dict__') else __builtins__), "__import__": _safe_import},
        # Paths
        "TIMBRADO_PATH": TIMBRADO_PATH,
        "OUTPUT_DIR": OUTPUT_DIR,
        # Standard modules
        "os": os,
        "re": re,
        "json": json,
        "shutil": shutil,
        "zipfile": zipfile,
        "datetime": dt_module.datetime,  # expose datetime.datetime as "datetime" so datetime.now() works
        "date": dt_module.date,
        "timedelta": dt_module.timedelta,
        "dt_module": dt_module,  # full module available as dt_module if needed
        "Decimal": __import__("decimal").Decimal,
        # openpyxl
        "openpyxl": openpyxl_module,
        "Font": Font,
        "PatternFill": PatternFill,
        "Alignment": Alignment,
        "Border": Border,
        "Side": Side,
        "get_column_letter": get_column_letter,
        # All helper functions
        "esc": helpers.esc,
        "run": helpers.run,
        "para": helpers.para,
        "sec_title": helpers.sec_title,
        "sub_title": helpers.sub_title,
        "bp": helpers.bp,
        "bp_r": helpers.bp_r,
        "ped": helpers.ped,
        "quesito": helpers.quesito,
        "empty": helpers.empty,
        "empty_line": helpers.empty_line,
        "table_row": helpers.table_row,
        "make_table": helpers.make_table,
        "make_total_row": helpers.make_total_row,
        "setup_docx": helpers.setup_docx,
        "save_docx": helpers.save_docx,
        "meses_entre": helpers.meses_entre,
        "data_extenso": helpers.data_extenso,
    }
    return g


@app.route("/api/gerar", methods=["POST"])
@require_admin
def gerar():
    try:
        data = request.get_json()
        client = anthropic.Anthropic(timeout=120.0)
        skill_prompt = load_skill_prompt()

        # Clean output directory
        for f in glob.glob(os.path.join(OUTPUT_DIR, "*")):
            if os.path.isfile(f) and not f.endswith(".gitkeep"):
                os.remove(f)
            elif os.path.isdir(f):
                shutil.rmtree(f, ignore_errors=True)

        # Determine which documents to generate
        docs = data.get("documentos", ["todos"])
        if "todos" in docs:
            doc_list = ["peticao", "planilha", "quesitos_medicos", "quesitos_sociais"]
        else:
            doc_list = docs

        doc_labels = {
            "peticao": "peticao inicial BPC/LOAS (.docx com timbrado)",
            "planilha": "planilha de calculo de atrasados modelo Conta Facil Prev (.xlsx com openpyxl)",
            "quesitos_medicos": "quesitos para pericia medica (.docx com timbrado)",
            "quesitos_sociais": "quesitos para pericia social (.docx com timbrado)",
        }

        all_files = []
        errors = []
        base_data_msg = build_data_summary(data)

        # Pre-built helper instructions to reduce tokens
        helper_instructions = """
AMBIENTE DE EXECUCAO - as seguintes funcoes e modulos ja estao disponiveis (NAO redefinir):
- Modulos: os, re, json, shutil, zipfile, datetime (classe datetime.datetime, use datetime.now(), datetime.strptime()), date, timedelta, Decimal, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
- Variaveis: TIMBRADO_PATH, OUTPUT_DIR
- Funcoes DOCX: esc(t), run(text, bold, sz, color, caps, italic), para(runs_str, jc, fi, li, before, after, line, shd, bdr_top, bdr_bot), sec_title(t), sub_title(t), bp(text, fi), bp_r(runs_str, fi, li), ped(letra, texto), quesito(num, texto), empty(), empty_line()
- Funcoes tabela: make_table(headers, rows), make_total_row(cells)
- Funcoes setup: setup_docx(TIMBRADO_PATH, OUTPUT_DIR) -> (base_dir, sect_pr, ns), save_docx(body_xml, sect_pr, ns, base_dir, output_path)
- Funcao calculo: meses_entre(data_str) -> int (meses entre a data e hoje)
- Funcao data: data_extenso() -> str (retorna data atual em portugues: "6 de marco de 2026")

SALARIO MINIMO ATUAL (2026): R$ 1.621,00 (usar SEMPRE este valor em calculos, planilhas e peticoes).
SM 2025 = R$ 1.518,00 (apenas para parcelas vencidas de 2025).
SM 2024 = R$ 1.412,00 (apenas para parcelas vencidas de 2024).

COMO GERAR UM DOCX:
base_dir, sect_pr, ns = setup_docx(TIMBRADO_PATH, OUTPUT_DIR)
body = ""
body += sec_title("TITULO")
body += bp("Paragrafo normal")
save_docx(body, sect_pr, ns, base_dir, os.path.join(OUTPUT_DIR, "arquivo.docx"))

REGRAS DO CODIGO (CRITICAS - seguir rigorosamente):
- NAO redefinir funcoes helper (esc, run, para, bp, etc.) - ja estao carregadas
- NAO usar emojis ou caracteres especiais unicode no codigo
- NAO usar caminhos /mnt/ ou /tmp/ - usar OUTPUT_DIR
- NAO importar modulos que ja estao disponiveis
- Usar apenas ASCII no codigo (acentos sao OK em strings de texto)
- Para nomes em negrito no meio do texto, SEMPRE usar bp_r() com run(): bp_r(run("NOME", bold=True) + run(" resto do texto"))
- NUNCA passar XML como string de texto para bp() - usar bp_r() com multiplos run()
- Datas SEMPRE em portugues: usar data_extenso() para data atual
- PLANILHA: gerar com 3 abas (Resumo, Calculo Detalhado, Criterios). A aba "Calculo Detalhado" deve ter CADA PARCELA em uma linha separada (mes a mes). Usar SM 2026 = 1621.00 para parcelas vincendas e valor da causa. Para parcelas vencidas, usar o SM do ano correspondente (2024=1412.00, 2025=1518.00, 2026=1621.00). Formatacao profissional: cabecalhos azul escuro (#1F3864 texto branco), zebrado nas linhas, bordas finas, totais em amarelo (#FFF2CC). Configurar page_setup landscape e margens estreitas.
- PETICAO: valor da causa = (meses_vencidos * SM_do_ano) + (12 * 1621.00). Usar SM 2026 = R$ 1.621,00.
- PETICAO: o CEP DEVE aparecer na qualificacao das partes, formato "CEP XXXXX-XXX". Nunca omitir o CEP.
- NUNCA usar f-strings aninhadas com aspas iguais. Exemplo ERRADO: f"{f'algo'}". Usar variaveis intermediarias.
- Testar que todas as variaveis existem antes de usar em f-strings
- TERMINOLOGIA: usar "autor" (masculino) ou "autora" (feminino) em vez de "requerente". NUNCA usar "requerente".
- Nos quesitos, usar "Autor:" ou "Autora:" na qualificacao resumida.
"""

        for doc_type in doc_list:
            label = doc_labels.get(doc_type, doc_type)
            print(f"[GERANDO] {label}...")

            user_msg = f"""Gere APENAS: {label}

{base_data_msg}

{helper_instructions}

Gere um UNICO bloco de codigo Python. NAO gere outros documentos.
"""

            last_error = None
            for attempt in range(3):  # up to 3 attempts
                try:
                    if attempt > 0 and last_error:
                        user_msg_attempt = user_msg + f"\n\nERRO NA TENTATIVA ANTERIOR: {last_error}\nCorrija o erro e gere novamente. Verifique parenteses, f-strings e sintaxe.\n"
                        print(f"[RETRY {attempt+1}] {label}: {last_error[:100]}")
                    else:
                        user_msg_attempt = user_msg

                    response_text = ""
                    with client.messages.stream(
                        model="claude-sonnet-4-20250514",
                        max_tokens=16000,
                        system=skill_prompt,
                        messages=[{"role": "user", "content": user_msg_attempt}],
                    ) as stream:
                        for text in stream.text_stream:
                            response_text += text

                    code = extract_python_code(response_text)
                    if not code:
                        last_error = "Claude nao gerou codigo Python"
                        continue

                    # Sanitize code
                    code = sanitize_code(code)

                    # Execute with pre-loaded helpers
                    exec_globals = build_exec_globals()
                    old_cwd = os.getcwd()
                    os.chdir(BASE_DIR)
                    try:
                        exec(code, exec_globals)
                    finally:
                        os.chdir(old_cwd)

                    # Check generated files
                    for f in os.listdir(OUTPUT_DIR):
                        if f.endswith((".docx", ".xlsx")) and not f.startswith("_") and f not in all_files:
                            all_files.append(f)

                    print(f"[OK] {label}")
                    last_error = None
                    break

                except Exception as e:
                    traceback.print_exc()
                    last_error = str(e)

            if last_error:
                errors.append(f"{label}: {last_error}")

        if not all_files:
            error_msg = "Nenhum documento foi gerado."
            if errors:
                error_msg += " Erros: " + "; ".join(errors)
            return jsonify({"error": error_msg}), 400

        result = {"files": sorted(all_files), "message": f"{len(all_files)} documento(s) gerado(s) com sucesso!"}
        if errors:
            result["warnings"] = errors
        return jsonify(result)

    except anthropic.APIError as e:
        return jsonify({"error": f"Erro na API do Claude: {str(e)}"}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Erro: {str(e)}"}), 500


@app.route("/api/lote", methods=["POST"])
@require_admin
def lote():
    """Process multiple client folders end-to-end.
    For each folder: analyze docs -> extract data -> generate all documents -> copy to folder.
    Returns results via streaming SSE so frontend can show progress.
    """
    from flask import Response, stream_with_context
    data = request.get_json()
    pastas = data.get("pastas", [])

    def generate():
        results = []
        for i, pasta in enumerate(pastas):
            pasta = pasta.strip()
            if not pasta or not os.path.isdir(pasta):
                yield f"data: {json.dumps({'type': 'error', 'pasta': pasta, 'message': 'Pasta nao encontrada'})}\n\n"
                continue

            cliente_nome = os.path.basename(pasta).split("_")[0].strip()
            yield f"data: {json.dumps({'type': 'progress', 'pasta': pasta, 'step': 'analisando', 'index': i, 'total': len(pastas), 'cliente': cliente_nome})}\n\n"

            # Step 1: Analyze folder
            try:
                extracted = analisar_pasta_internal(pasta)
            except Exception as e:
                traceback.print_exc()
                yield f"data: {json.dumps({'type': 'error', 'pasta': pasta, 'cliente': cliente_nome, 'message': f'Erro na analise: {str(e)}'})}\n\n"
                continue

            yield f"data: {json.dumps({'type': 'progress', 'pasta': pasta, 'step': 'gerando', 'index': i, 'total': len(pastas), 'cliente': cliente_nome})}\n\n"

            # Step 2: Generate documents
            try:
                generated_files = gerar_documentos_internal(extracted)
            except Exception as e:
                traceback.print_exc()
                yield f"data: {json.dumps({'type': 'error', 'pasta': pasta, 'cliente': cliente_nome, 'message': f'Erro na geracao: {str(e)}'})}\n\n"
                continue

            # Step 3: Remove old generated docs from client folder before copying new ones
            generated_patterns = ["peticao_bpc", "calculo_atrasados", "quesitos_pericia",
                                  "PETICAO INICIAL", "CALCULO DE ATRASADOS", "QUESITOS PERICIA",
                                  "1- PETICAO", "17- CALCULO", "18- QUESITOS", "19- QUESITOS"]
            for f in os.listdir(pasta):
                f_path = os.path.join(pasta, f)
                if os.path.isfile(f_path):
                    f_lower = f.lower()
                    if any(p.lower() in f_lower for p in generated_patterns):
                        try:
                            os.remove(f_path)
                            print(f"  [REMOVIDO] {f} (antigo)")
                        except Exception:
                            pass

            # Copy new generated files
            copied = []
            for fname in generated_files:
                src = os.path.join(OUTPUT_DIR, fname)
                dst = os.path.join(pasta, fname)
                try:
                    shutil.copy2(src, dst)
                    copied.append(fname)
                except Exception as e:
                    print(f"[WARN] Erro ao copiar {fname}: {e}")

            # Step 5: Detect duplicates, organize, merge multi-part PDFs, convert xlsx to PDF
            yield f"data: {json.dumps({'type': 'progress', 'pasta': pasta, 'step': 'organizando', 'index': i, 'total': len(pastas), 'cliente': cliente_nome})}\n\n"
            try:
                detect_duplicates(pasta)
            except Exception as e:
                print(f"[WARN] Erro ao detectar duplicados: {e}")
            try:
                merge_pdf_parts(pasta)  # First pass: merge original "parte1/parte2" files
            except Exception as e:
                print(f"[WARN] Erro ao juntar PDFs: {e}")
            try:
                organizar_pasta(pasta)
            except Exception as e:
                print(f"[WARN] Erro ao organizar pasta: {e}")
            try:
                merge_pdf_parts(pasta)  # Second pass: merge renamed "3- CONTRATO 1/2/3" files
            except Exception as e:
                print(f"[WARN] Erro ao juntar PDFs renomeados: {e}")

            # Convert xlsx to PDF (after organizing, so file has final name)
            for f in os.listdir(pasta):
                if f.endswith('.xlsx') and os.path.isfile(os.path.join(pasta, f)):
                    pdf_result = xlsx_to_pdf(os.path.join(pasta, f))
                    if pdf_result:
                        copied.append(os.path.basename(pdf_result))

            yield f"data: {json.dumps({'type': 'done', 'pasta': pasta, 'cliente': cliente_nome, 'files': copied, 'index': i, 'total': len(pastas)})}\n\n"
            results.append({"pasta": pasta, "cliente": cliente_nome, "files": copied})

        yield f"data: {json.dumps({'type': 'complete', 'results': results})}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


# ==================== MISTRAL OCR ====================
_MISTRAL_API_KEY = os.environ.get("MISTRAL_API_KEY", "")
_mistral_client = None
_ocr_cache = {}  # file_path -> text (avoid re-OCRing same file)

def _get_mistral():
    global _mistral_client
    if _mistral_client is None and _MISTRAL_API_KEY:
        from mistralai.client import Mistral
        _mistral_client = Mistral(api_key=_MISTRAL_API_KEY)
    return _mistral_client

def mistral_ocr(file_path):
    """Extract text from scanned PDF/image via Mistral OCR SDK.
    Returns markdown text or empty string on failure. Results are cached."""
    # Check cache first
    abs_path = os.path.abspath(file_path)
    if abs_path in _ocr_cache:
        return _ocr_cache[abs_path]

    client = _get_mistral()
    if not client:
        return ""
    try:
        filename = os.path.basename(file_path)
        with open(file_path, "rb") as f:
            uploaded = client.files.upload(file={"file_name": filename, "content": f.read()}, purpose="ocr")

        result = client.ocr.process(
            model="mistral-ocr-latest",
            document={"type": "file", "file_id": uploaded.id}
        )

        text = "\n\n".join(p.markdown for p in result.pages if p.markdown)
        if len(text.strip()) > 20:
            print(f"  [OCR OK] {filename} -> {len(text)} chars")
            _ocr_cache[abs_path] = text.strip()
            return text.strip()
    except Exception as e:
        print(f"  [OCR FAIL] {os.path.basename(file_path)}: {e}")
    _ocr_cache[abs_path] = ""
    return ""


def analisar_pasta_internal(pasta):
    """Internal: analyze a folder and return extracted data dict.

    Handles large folders by:
    1. Extracting text from ALL PDFs first (cheap, small)
    2. Only sending images for scanned docs that lack text
    3. Tracking total base64 size to stay under API limits (~9MB safe)
    4. Reducing DPI for large folders
    5. Skipping non-essential docs (contrato, procuracao, OAB) as images
    """
    extensions = (".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp")
    files_found = []
    for f in sorted(os.listdir(pasta)):
        if f.lower().endswith(extensions) and not f.startswith('.') and f != 'desktop.ini':
            files_found.append(os.path.join(pasta, f))

    if not files_found:
        raise ValueError("Nenhum PDF ou imagem encontrado na pasta.")

    print(f"[ANALISE] {len(files_found)} arquivos em {os.path.basename(pasta)}")

    # Extract text from all PDFs
    file_info = []
    for filepath in files_found:
        filename = os.path.basename(filepath)
        info = {"path": filepath, "name": filename, "text": "", "pages": 0, "has_text": False}
        if filepath.lower().endswith(".pdf"):
            info["type"] = "pdf"
            try:
                doc = fitz.open(filepath)
                info["pages"] = len(doc)
                text = ""
                for page in doc:
                    text += page.get_text() + "\n"
                doc.close()
                info["text"] = text.strip()
                info["has_text"] = len(info["text"]) > 100
            except Exception:
                pass
        else:
            info["type"] = "image"
        file_info.append(info)

    # Build content
    text_files = [f for f in file_info if f["has_text"]]
    scan_files = [f for f in file_info if not f["has_text"]]

    # Truncate very long text to avoid huge requests (e.g. PROCESSO INSS with 50+ pages)
    MAX_TEXT_PER_FILE = 8000  # chars
    text_content = f"Analise os documentos de um cliente BPC/LOAS.\nTotal: {len(file_info)} arquivos.\n\n"
    for f in text_files:
        file_text = f['text']
        if len(file_text) > MAX_TEXT_PER_FILE:
            file_text = file_text[:MAX_TEXT_PER_FILE] + f"\n[... texto truncado, total {len(f['text'])} chars ...]"
        text_content += f"=== {f['name']} ({f['pages']}pg) ===\n{file_text}\n\n"

    content_parts = [{"type": "text", "text": text_content}]

    # Scanned docs: try Mistral OCR first, fallback to image
    def scan_priority(f):
        name = f["name"].lower()
        for i, kw in enumerate(["laudo", "certid", "nascimento", "rg", "cpf", "identif", "sus", "autodecl", "cadunico", "relatorio", "parecer", "receita", "encaminhamento", "grupo_familiar", "comprometimento"]):
            if kw in name:
                return i
        return 99

    scan_files.sort(key=scan_priority)
    ocr_count = 0
    total_images = 0
    total_b64_bytes = 0
    MAX_IMAGES = 15
    MAX_B64_BYTES = 9 * 1024 * 1024

    for f in scan_files:
        # Try Mistral OCR first (converts scan to text - cheaper and no limits)
        ocr_text = mistral_ocr(f["path"])
        if ocr_text:
            ocr_count += 1
            file_text = ocr_text
            if len(file_text) > MAX_TEXT_PER_FILE:
                file_text = file_text[:MAX_TEXT_PER_FILE] + f"\n[... truncado, total {len(ocr_text)} chars ...]"
            text_content += f"=== {f['name']} (OCR) ===\n{file_text}\n\n"
            # Update content_parts[0] with new text_content
            content_parts[0] = {"type": "text", "text": text_content}
            continue

        # OCR failed - fallback to image (old behavior)
        if total_images >= MAX_IMAGES or total_b64_bytes >= MAX_B64_BYTES:
            print(f"  [SKIP] {f['name']} (limite imagens atingido)")
            continue

        SKIP_IMAGE_KEYWORDS = ["contrato", "procura", "oab", "pedido.gratuidade", "fatura", "biometria", "conta.luz"]
        name_lower = f["name"].lower()
        if any(kw in name_lower for kw in SKIP_IMAGE_KEYWORDS):
            print(f"  [SKIP IMG] {f['name']} (nao essencial)")
            continue

        content_parts.append({"type": "text", "text": f"\n--- Escaneado: {f['name']} ---"})
        if f["type"] == "pdf":
            try:
                max_pg = min(2, MAX_IMAGES - total_images)
                images = pdf_to_images(f["path"], max_pages=max_pg, dpi=120)
                for img_b64 in images:
                    b64_size = len(img_b64)
                    if total_b64_bytes + b64_size > MAX_B64_BYTES:
                        break
                    content_parts.append({"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": img_b64}})
                    total_images += 1
                    total_b64_bytes += b64_size
            except Exception:
                pass
        else:
            try:
                img_b64, mt = image_to_base64(f["path"])
                if img_b64:
                    b64_size = len(img_b64)
                    if total_b64_bytes + b64_size <= MAX_B64_BYTES:
                        content_parts.append({"type": "image", "source": {"type": "base64", "media_type": mt, "data": img_b64}})
                        total_images += 1
                        total_b64_bytes += b64_size
            except Exception:
                pass

    print(f"[ANALISE] Enviando: {len(text_files)} docs texto + {ocr_count} OCR + {total_images} imagens ({total_b64_bytes / 1024 / 1024:.1f}MB)")

    content_parts.append({"type": "text", "text": EXTRACTION_PROMPT})

    client = anthropic.Anthropic(timeout=120.0)
    response_text = ""
    with client.messages.stream(
        model="claude-haiku-4-5-20251001",
        max_tokens=4096,
        messages=[{"role": "user", "content": content_parts}],
    ) as stream:
        for text in stream.text_stream:
            response_text += text

    extracted = extract_json(response_text)
    if not extracted:
        raise ValueError("Nao conseguiu extrair dados dos documentos")

    return extracted


def gerar_documentos_internal(data):
    """Internal: generate all 4 documents from extracted data. Returns list of filenames."""
    client = anthropic.Anthropic(timeout=120.0)
    skill_prompt = load_skill_prompt()

    # Clean output
    for f in glob.glob(os.path.join(OUTPUT_DIR, "*")):
        if os.path.isfile(f) and not f.endswith(".gitkeep"):
            os.remove(f)
        elif os.path.isdir(f):
            shutil.rmtree(f, ignore_errors=True)

    doc_list = [
        ("peticao", "peticao inicial BPC/LOAS (.docx com timbrado)"),
        ("planilha", "planilha de calculo de atrasados modelo Conta Facil Prev (.xlsx com openpyxl)"),
        ("quesitos_medicos", "quesitos para pericia medica (.docx com timbrado)"),
        ("quesitos_sociais", "quesitos para pericia social (.docx com timbrado)"),
    ]

    base_data_msg = build_data_summary(data)
    helper_instructions = """
AMBIENTE DE EXECUCAO - funcoes e modulos ja disponiveis (NAO redefinir):
- Modulos: os, re, json, shutil, zipfile, datetime (classe datetime.datetime, use datetime.now(), datetime.strptime()), date, timedelta, Decimal, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
- Variaveis: TIMBRADO_PATH, OUTPUT_DIR
- Funcoes DOCX: esc(t), run(text, bold, sz, color, caps, italic), para(runs_str, jc, fi, li, before, after, line, shd, bdr_top, bdr_bot), sec_title(t), sub_title(t), bp(text, fi), bp_r(runs_str, fi, li), ped(letra, texto), quesito(num, texto), empty(), empty_line()
- Funcoes tabela: make_table(headers, rows), make_total_row(cells)
- Funcoes setup: setup_docx(TIMBRADO_PATH, OUTPUT_DIR) -> (base_dir, sect_pr, ns), save_docx(body_xml, sect_pr, ns, base_dir, output_path)
- Funcao calculo: meses_entre(data_str) -> int
- Funcao data: data_extenso() -> str (data atual em portugues)

SALARIO MINIMO ATUAL (2026): R$ 1.621,00 (usar SEMPRE este valor em calculos, planilhas e peticoes).
SM 2025 = R$ 1.518,00 (apenas para parcelas vencidas de 2025). SM 2024 = R$ 1.412,00.

COMO GERAR DOCX:
base_dir, sect_pr, ns = setup_docx(TIMBRADO_PATH, OUTPUT_DIR)
body = ""
body += sec_title("TITULO")
body += bp("Paragrafo")
save_docx(body, sect_pr, ns, base_dir, os.path.join(OUTPUT_DIR, "arquivo.docx"))

REGRAS CRITICAS:
- NAO redefinir helpers. NAO usar emojis. NAO usar /mnt/ ou /tmp/. Usar OUTPUT_DIR.
- Para nomes em negrito: bp_r(run("NOME", bold=True) + run(" resto"))
- NUNCA passar XML como texto para bp(). Datas em portugues com data_extenso().
- PLANILHA: SM 2026 = 1621.00 para vincendas. Para vencidas usar SM do ano correspondente.
- NUNCA usar f-strings aninhadas. Usar variaveis intermediarias.
- TERMINOLOGIA: usar "autor" (masculino) ou "autora" (feminino) em vez de "requerente". NUNCA usar "requerente".
- Nos quesitos, usar "Autor:" ou "Autora:" na qualificacao resumida.
"""

    all_files = []

    for doc_type, label in doc_list:
        print(f"  [GERANDO] {label}...")
        user_msg = f"Gere APENAS: {label}\n\n{base_data_msg}\n\n{helper_instructions}\n\nGere um UNICO bloco ```python. NAO gere outros documentos."

        last_error = None
        for attempt in range(3):
            try:
                msg = user_msg
                if attempt > 0 and last_error:
                    msg += f"\n\nERRO ANTERIOR: {last_error}\nCorrija e gere novamente."
                    print(f"  [RETRY {attempt+1}] {last_error[:80]}")

                response_text = ""
                with client.messages.stream(
                    model="claude-sonnet-4-20250514",
                    max_tokens=16000,
                    system=[{"type": "text", "text": skill_prompt, "cache_control": {"type": "ephemeral"}}],
                    messages=[{"role": "user", "content": msg}],
                    extra_headers={"anthropic-beta": "prompt-caching-2024-07-31"},
                ) as stream:
                    for text in stream.text_stream:
                        response_text += text

                code = extract_python_code(response_text)
                if not code:
                    last_error = "Sem codigo Python"
                    continue

                code = sanitize_code(code)
                exec_globals = build_exec_globals()
                old_cwd = os.getcwd()
                os.chdir(BASE_DIR)
                try:
                    exec(code, exec_globals)
                finally:
                    os.chdir(old_cwd)

                for f in os.listdir(OUTPUT_DIR):
                    if f.endswith((".docx", ".xlsx")) and not f.startswith("_") and f not in all_files:
                        all_files.append(f)

                print(f"  [OK] {label}")
                last_error = None
                break
            except Exception as e:
                traceback.print_exc()
                last_error = str(e)

        if last_error:
            print(f"  [FALHOU] {label}: {last_error}")

    return all_files


@app.route("/api/analisar-pasta", methods=["POST"])
@require_admin
def analisar_pasta():
    """Analyze documents in a folder and extract client data.

    Strategy for handling large folders:
    1. Extract text from ALL PDFs first (text is tiny vs images)
    2. Classify each file: has_text (>100 chars) or scanned (needs image)
    3. Send all text-based content in one message
    4. For scanned docs, send images in batches if needed
    5. If a second batch is needed, merge results
    """
    try:
        data = request.get_json()
        pasta = data.get("pasta", "").strip()

        if not pasta or not os.path.isdir(pasta):
            return jsonify({"error": f"Pasta nao encontrada: {pasta}"}), 400

        # Collect all PDFs and images
        extensions = (".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp")
        files_found = []
        for f in sorted(os.listdir(pasta)):
            if f.lower().endswith(extensions) and not f.startswith('.') and f != 'desktop.ini':
                files_found.append(os.path.join(pasta, f))

        if not files_found:
            return jsonify({"error": "Nenhum PDF ou imagem encontrado na pasta."}), 400

        print(f"[ANALISE] {len(files_found)} arquivos encontrados em {pasta}")

        # === PASS 1: Extract text from all PDFs ===
        file_info = []  # list of {path, name, type, text, pages, has_text}
        for filepath in files_found:
            filename = os.path.basename(filepath)
            info = {"path": filepath, "name": filename, "text": "", "pages": 0, "has_text": False}

            if filepath.lower().endswith(".pdf"):
                info["type"] = "pdf"
                try:
                    doc = fitz.open(filepath)
                    info["pages"] = len(doc)
                    text = ""
                    for page in doc:
                        text += page.get_text() + "\n"
                    doc.close()
                    info["text"] = text.strip()
                    info["has_text"] = len(info["text"]) > 100
                except Exception as e:
                    print(f"[WARN] Erro ao ler {filename}: {e}")
            else:
                info["type"] = "image"
                info["has_text"] = False

            file_info.append(info)
            print(f"  [{info['type'].upper()}] {filename} - {info['pages']}pg - texto: {len(info['text'])} chars")

        # === PASS 2: Build content for Claude ===
        # Separate files into text-based and image-needed
        text_files = [f for f in file_info if f["has_text"]]
        scan_files = [f for f in file_info if not f["has_text"]]

        # Build the text-only content first (all PDFs with extractable text)
        text_content = f"Analise os documentos abaixo de um cliente BPC/LOAS e extraia todos os dados.\n\n"
        text_content += f"Total de arquivos: {len(file_info)}\n\n"

        for f in text_files:
            text_content += f"=== {f['name']} ({f['pages']} paginas) ===\n"
            text_content += f"{f['text']}\n\n"

        # Build multimodal parts: text summary + scanned images
        content_parts = [{"type": "text", "text": text_content}]

        # Add scanned documents as images (with size control)
        # Target: stay under ~18MB total to avoid 413
        MAX_IMAGES = 20  # conservative limit
        total_images = 0
        DPI_SCAN = 150   # good quality but reasonable size

        # Prioritize scanned files: laudos and key docs first
        def scan_priority(f):
            name = f["name"].lower()
            if "laudo" in name:
                return 0
            if "certid" in name or "nascimento" in name:
                return 1
            if "rg" in name or "cpf" in name or "identif" in name:
                return 2
            if "sus" in name:
                return 3
            if "autodecl" in name or "cadunico" in name:
                return 4
            return 5

        scan_files.sort(key=scan_priority)
        ocr_count = 0

        for f in scan_files:
            # Try Mistral OCR first (converts scan to text)
            ocr_text = mistral_ocr(f["path"])
            if ocr_text:
                ocr_count += 1
                text_content += f"=== {f['name']} (OCR) ===\n{ocr_text}\n\n"
                content_parts[0] = {"type": "text", "text": text_content}
                continue

            # OCR failed - fallback to image
            if total_images >= MAX_IMAGES:
                content_parts.append({
                    "type": "text",
                    "text": f"\n[AVISO: {f['name']} nao incluido - limite atingido.]"
                })
                continue

            content_parts.append({
                "type": "text",
                "text": f"\n--- Documento escaneado: {f['name']} ---"
            })

            if f["type"] == "pdf":
                max_pg = min(3, MAX_IMAGES - total_images)
                try:
                    images = pdf_to_images(f["path"], max_pages=max_pg, dpi=DPI_SCAN)
                    for img_b64 in images:
                        content_parts.append({
                            "type": "image",
                            "source": {"type": "base64", "media_type": "image/png", "data": img_b64}
                        })
                        total_images += 1
                except Exception as e:
                    print(f"[WARN] Erro ao converter {f['name']} para imagem: {e}")
            else:
                try:
                    img_b64, media_type = image_to_base64(f["path"])
                    if img_b64:
                        content_parts.append({
                            "type": "image",
                            "source": {"type": "base64", "media_type": media_type, "data": img_b64}
                        })
                        total_images += 1
                except Exception as e:
                    print(f"[WARN] Erro ao processar imagem {f['name']}: {e}")

        print(f"[ANALISE] Enviando: {len(text_files)} docs texto + {ocr_count} OCR + {total_images} imagens")

        # Add extraction prompt
        content_parts.append({
            "type": "text",
            "text": EXTRACTION_PROMPT
        })

        # === PASS 3: Call Claude API with streaming ===
        client = anthropic.Anthropic(timeout=120.0)
        response_text = ""
        with client.messages.stream(
            model="claude-haiku-4-5-20251001",
            max_tokens=4096,
            messages=[{"role": "user", "content": content_parts}],
        ) as stream:
            for text in stream.text_stream:
                response_text += text

        # Parse JSON response
        extracted = extract_json(response_text)
        if not extracted:
            return jsonify({"error": "Nao foi possivel extrair dados. Resposta:\n" + response_text[:2000]}), 400

        return jsonify({
            "data": extracted,
            "arquivos_analisados": [f["name"] for f in file_info],
            "message": f"{len(file_info)} documento(s) analisado(s) com sucesso!"
        })

    except anthropic.APIError as e:
        return jsonify({"error": f"Erro na API do Claude: {str(e)}"}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Erro: {str(e)}"}), 500


# Extraction prompt constant
EXTRACTION_PROMPT = """
Com base em TODOS os documentos acima, extraia e retorne APENAS um JSON valido (sem markdown, sem ```), com esta estrutura exata:

{
  "nome": "NOME COMPLETO DO REQUERENTE (crianca/pessoa com deficiencia)",
  "cpf": "000.000.000-00",
  "rg": "",
  "data_nascimento": "YYYY-MM-DD",
  "endereco": "endereco completo com rua, numero, bairro",
  "cep": "XXXXX-XXX (extrair do comprovante de residencia ou autodeclaracao)",
  "cidade": "cidade",
  "estado": "UF (sigla de 2 letras)",
  "representante_nome": "nome completo do representante legal",
  "representante_cpf": "CPF do representante",
  "representante_rg": "RG do representante",
  "representante_parentesco": "mae/pai/tutor/curador",
  "cids": "F90.0, F84.0 (todos os CIDs encontrados nos laudos)",
  "descricao_saude": "descricao DETALHADA: diagnosticos, sintomas, limitacoes funcionais, medicamentos em uso, tratamentos necessarios. Extrair dos laudos medicos.",
  "nb": "numero do beneficio (formato XXX.XXX.XXX-X)",
  "der": "YYYY-MM-DD (data de entrada do requerimento)",
  "motivo_indeferimento": "motivo exato do indeferimento pelo INSS",
  "familia": [
    {"nome": "NOME COMPLETO", "parentesco": "mae/pai/irmao/irma/avo", "idade": "XX", "renda": "0,00"}
  ],
  "gastos": [
    {"categoria": "Saude", "item": "nome do gasto", "valor": "", "observacao": "Art. 20-B LOAS"}
  ]
}

REGRAS:
- Se um dado nao foi encontrado, deixe como string vazia ""
- Para familia, incluir TODOS os membros que aparecem nos documentos (CadUnico, certidoes, etc.)
- Para gastos de saude (medicamentos, terapias, consultas), SEMPRE colocar observacao "Art. 20-B LOAS"
- O valor dos gastos pode ficar vazio se nao encontrado nos documentos
- Extrair o MAXIMO de informacoes possiveis
- Retornar APENAS o JSON, sem texto antes ou depois
"""


def pdf_extract_text(pdf_path):
    """Extract text from all pages of a PDF."""
    text = ""
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()
    except Exception:
        pass
    return text


def pdf_to_images(pdf_path, max_pages=20, dpi=200):
    """Convert PDF pages to base64 PNG images."""
    images = []
    doc = fitz.open(pdf_path)
    for i, page in enumerate(doc):
        if i >= max_pages:
            break
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        images.append(base64.b64encode(img_bytes).decode("utf-8"))
    doc.close()
    return images


def image_to_base64(image_path):
    """Convert image file to base64 with appropriate media type."""
    ext = os.path.splitext(image_path)[1].lower()
    media_types = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".bmp": "image/bmp",
        ".tiff": "image/tiff",
        ".webp": "image/webp",
    }
    media_type = media_types.get(ext, "image/png")

    # Resize if too large (Claude has limits)
    try:
        img = Image.open(image_path)
        max_dim = 2048
        if img.width > max_dim or img.height > max_dim:
            img.thumbnail((max_dim, max_dim), Image.LANCZOS)

        # Convert to PNG for consistency
        import io
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        img_bytes = buf.getvalue()
        return base64.b64encode(img_bytes).decode("utf-8"), "image/png"
    except Exception:
        # Fallback: read raw bytes
        with open(image_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8"), media_type


def extract_json(text):
    """Extract JSON object from Claude's response."""
    # Try direct parse
    text = text.strip()
    if text.startswith("{"):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass

    # Try to find JSON block in markdown
    match = re.search(r"```(?:json)?\s*\n(\{.*?\})\s*\n```", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError:
            pass

    # Try to find any JSON object
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass

    return None


@app.route("/api/download/<path:filename>")
def download(filename):
    # Block path traversal attempts (including URL-encoded variants)
    if not filename or ".." in filename or "/" in filename or "\\" in filename:
        return jsonify({"error": "invalid filename"}), 400
    # Verify resolved path stays inside OUTPUT_DIR
    full_path = os.path.abspath(os.path.join(OUTPUT_DIR, filename))
    if not full_path.startswith(os.path.abspath(OUTPUT_DIR)):
        return jsonify({"error": "invalid path"}), 400
    if not os.path.isfile(full_path):
        return jsonify({"error": "file not found"}), 404
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


def build_data_summary(data):
    """Build data summary for individual document generation."""
    msg = f"""DADOS DO CLIENTE:
- Nome: {data.get('nome', '')}
- CPF: {data.get('cpf', '')}
- Data de nascimento: {data.get('data_nascimento', '')}
- RG: {data.get('rg', '')}
- Endereço: {data.get('endereco', '')}
- CEP: {data.get('cep', '')}
- Cidade/Comarca: {data.get('cidade', '')}
- Estado: {data.get('estado', '')}
"""

    if data.get("representante_nome"):
        msg += f"""
REPRESENTANTE LEGAL:
- Nome: {data.get('representante_nome', '')}
- CPF: {data.get('representante_cpf', '')}
- RG: {data.get('representante_rg', '')}
- Parentesco: {data.get('representante_parentesco', '')}
"""

    msg += f"""
CONDIÇÃO DE SAÚDE:
- CID(s): {data.get('cids', '')}
- Descrição da condição: {data.get('descricao_saude', '')}

DADOS DO PROCESSO:
- NB (Número do Benefício): {data.get('nb', '')}
- DER (Data de Entrada do Requerimento): {data.get('der', '')}
- Motivo do indeferimento: {data.get('motivo_indeferimento', '')}
"""

    # Composição familiar
    familia = data.get("familia", [])
    if familia:
        msg += "\nCOMPOSIÇÃO FAMILIAR:\n"
        for m in familia:
            msg += f"- {m.get('nome', '')}, {m.get('parentesco', '')}, idade: {m.get('idade', '')}, renda: R$ {m.get('renda', '0,00')}\n"

    # Gastos
    gastos = data.get("gastos", [])
    if gastos:
        msg += "\nGASTOS MENSAIS:\n"
        for g in gastos:
            obs = f" ({g.get('observacao', '')})" if g.get("observacao") else ""
            msg += f"- {g.get('categoria', '')}: {g.get('item', '')} - R$ {g.get('valor', '0,00')}{obs}\n"

    return msg


def sanitize_code(code):
    """Clean up generated code to avoid common errors and block dangerous patterns."""
    # SECURITY: Block dangerous patterns that could be injected via prompt injection
    _DANGEROUS_PATTERNS = [
        '__import__', 'subprocess', 'os.system', 'os.popen', 'os.exec',
        'eval(', 'exec(', 'compile(', 'globals(', 'locals(',
        'socket', 'urllib', 'requests.', 'httpx.', 'http.client',
        'shutil.rmtree', 'shutil.move', 'open(', 'builtins',
        'importlib', 'sys.exit', 'os.kill', 'signal.',
    ]
    for pattern in _DANGEROUS_PATTERNS:
        if pattern in code:
            # Remove the dangerous line entirely
            code = '\n'.join(line for line in code.split('\n')
                           if pattern not in line)

    # Remove emoji/unicode chars that break Windows charmap
    code = re.sub(r'[\U00010000-\U0010ffff]', '', code)
    code = code.replace('\u2713', 'v').replace('\u2705', '[OK]').replace('\u274c', '[X]')
    code = code.replace('\u2022', '-').replace('\u2014', '-').replace('\u2013', '-')
    code = code.replace('\u2018', "'").replace('\u2019', "'")
    code = code.replace('\u201c', '"').replace('\u201d', '"')
    code = code.replace('\u2026', '...')
    code = code.replace('\u00a0', ' ')  # non-breaking space

    # Remove lines that try to redefine our helpers (common source of errors)
    lines = code.split('\n')
    filtered = []
    skip_until_unindent = False
    skip_func_name = None

    helper_funcs = {'esc', 'run', 'para', 'sec_title', 'sub_title', 'bp', 'bp_r',
                    'ped', 'quesito', 'empty', 'empty_line', 'table_row',
                    'make_table', 'make_total_row', 'setup_docx', 'save_docx', 'meses_entre'}

    for line in lines:
        stripped = line.strip()
        # Skip helper function redefinitions
        if stripped.startswith('def '):
            func_name = stripped.split('(')[0].replace('def ', '').strip()
            if func_name in helper_funcs:
                skip_until_unindent = True
                skip_func_name = func_name
                continue
        if skip_until_unindent:
            if stripped == '' or line[0:1] in (' ', '\t'):
                continue
            else:
                skip_until_unindent = False
                skip_func_name = None

        # Remove redundant imports of already-loaded modules
        if stripped.startswith('import ') and any(m in stripped for m in
            ['import shutil', 'import zipfile', 'import re', 'import os',
             'import json', 'import openpyxl', 'from openpyxl']):
            continue
        if stripped.startswith('from datetime import'):
            continue

        filtered.append(line)

    return '\n'.join(filtered)


def xlsx_to_pdf(xlsx_path):
    """Convert xlsx to PDF using Excel COM automation."""
    try:
        import subprocess
        abs_path = os.path.abspath(xlsx_path).replace('/', '\\')
        pdf_path = os.path.splitext(abs_path)[0] + '.pdf'
        # Write PS script to temp file to avoid encoding issues
        import tempfile
        ps_file = os.path.join(tempfile.gettempdir(), '_xlsx2pdf.ps1')
        with open(ps_file, 'w', encoding='utf-8-sig') as pf:
            pf.write(f'''
$ErrorActionPreference = "Stop"
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
    $wb = $excel.Workbooks.Open("{abs_path}")
    $ws = $wb.Worksheets.Item(1)
    $ws.PageSetup.Orientation = 2
    $ws.PageSetup.FitToPagesWide = 1
    $ws.PageSetup.FitToPagesTall = $false
    $ws.PageSetup.LeftMargin = $excel.InchesToPoints(0.4)
    $ws.PageSetup.RightMargin = $excel.InchesToPoints(0.4)
    $ws.PageSetup.TopMargin = $excel.InchesToPoints(0.5)
    $ws.PageSetup.BottomMargin = $excel.InchesToPoints(0.5)
    $wb.ExportAsFixedFormat(0, "{pdf_path}")
    $wb.Close($false)
    Write-Output "OK"
}} catch {{
    Write-Output ("ERRO: " + $_.Exception.Message)
}} finally {{
    $excel.Quit()
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
''')
        result = subprocess.run(
            ['powershell', '-ExecutionPolicy', 'Bypass', '-File', ps_file],
            capture_output=True, timeout=60)
        stdout = result.stdout.decode('utf-8', errors='replace').strip()
        print(f"  [xlsx_to_pdf] stdout: {stdout}")
        if 'OK' in stdout:
            print(f"  [PDF] {os.path.basename(pdf_path)} gerado")
            return pdf_path
        else:
            stderr = result.stderr.decode('utf-8', errors='replace').strip()
            print(f"  [WARN] xlsx_to_pdf stderr: {stderr}")
    except Exception as e:
        print(f"  [WARN] xlsx_to_pdf falhou: {e}")
    return None


def merge_pdf_parts(pasta):
    """Find and merge multi-part PDFs (e.g. contrato.parte1.pdf, contrato.parte2.pdf)."""
    files = [f for f in os.listdir(pasta) if f.lower().endswith('.pdf') and os.path.isfile(os.path.join(pasta, f))]

    # Group files by base name (detect patterns like "name.parte1.pdf", "name_part1.pdf",
    # "03_contrato_1.pdf", "3- CONTRATO DE HONORARIOS 1.pdf")
    groups = {}
    for f in files:
        f_lower = f.lower()
        # Pattern 1: name.parte1.pdf / name.parte2.pdf
        match = re.match(r'^(.+?)[\._\s]*(parte|part)[\._\s]*(\d+)\.pdf$', f_lower)
        if match:
            base = match.group(1).rstrip('._- ')
            num = int(match.group(3))
            groups.setdefault(base, []).append((num, f))
            continue
        # Pattern 2: "3- CONTRATO DE HONORARIOS 1.pdf" / "03_contrato_1.pdf"
        match = re.match(r'^(\d+[-_]\s*.+?)\s+(\d+)\.pdf$', f_lower)
        if match:
            base = match.group(1).rstrip()
            num = int(match.group(2))
            groups.setdefault(base, []).append((num, f))
            continue

    # Only merge groups that are clearly multi-part documents (contrato, relatorio desenvolvimental)
    # Do NOT merge procuracoes, documentos de ID, certidoes etc. (these are separate docs)
    MERGE_KEYWORDS = ["contrato", "relatorio desenvolvimental", "relatorio.desenvolvimental"]
    SKIP_MERGE = ["procura", "certid", "doc. de ident", "documento do", "doc_menor",
                  "doc_responsavel", "comprovante", "quesitos", "peticao", "cadunico",
                  "laudo", "receita", "exame", "fatura"]

    merged = []
    for base, parts in groups.items():
        if len(parts) < 2:
            continue
        # Check if this group should be merged
        base_lower = base.lower()
        should_merge = any(kw in base_lower for kw in MERGE_KEYWORDS)
        should_skip = any(kw in base_lower for kw in SKIP_MERGE)
        if should_skip or not should_merge:
            continue
        parts.sort(key=lambda x: x[0])
        print(f"  [MERGE] Juntando {len(parts)} partes: {base}")

        try:
            merged_doc = fitz.open()
            for num, fname in parts:
                pdf_path = os.path.join(pasta, fname)
                doc = fitz.open(pdf_path)
                merged_doc.insert_pdf(doc)
                doc.close()

            # Save merged file - use uppercase label from first file if available
            first_file = parts[0][1]
            base_from_file = re.match(r'^(.+?)\s+\d+\.pdf$', first_file, re.IGNORECASE)
            if base_from_file:
                merged_name = base_from_file.group(1).strip() + '.pdf'
            else:
                merged_name = base.replace('.', ' ').strip() + '.pdf'
            merged_path = os.path.join(pasta, merged_name)
            merged_doc.save(merged_path)
            merged_doc.close()
            print(f"  [MERGE] Salvo: {merged_name}")

            # Move parts to subfolder
            parts_dir = os.path.join(pasta, '_partes_originais')
            os.makedirs(parts_dir, exist_ok=True)
            for num, fname in parts:
                src = os.path.join(pasta, fname)
                dst = os.path.join(parts_dir, fname)
                if os.path.exists(src):
                    shutil.move(src, dst)

            merged.append(merged_name)
        except Exception as e:
            print(f"  [WARN] Erro ao juntar {base}: {e}")

    return merged


def detect_duplicates(pasta):
    """Detect duplicate files by content hash and move to subfolder."""
    import hashlib

    files = [f for f in os.listdir(pasta)
             if os.path.isfile(os.path.join(pasta, f)) and not f.startswith('.') and f != 'desktop.ini']

    # Hash all files
    hashes = {}
    for f in files:
        path = os.path.join(pasta, f)
        try:
            with open(path, 'rb') as fh:
                h = hashlib.md5(fh.read()).hexdigest()
            hashes.setdefault(h, []).append(f)
        except Exception:
            pass

    # Find duplicates
    duplicates = []
    dup_dir = os.path.join(pasta, '_duplicados')
    for h, fnames in hashes.items():
        if len(fnames) > 1:
            # Keep the first (alphabetically), move the rest
            fnames.sort()
            keep = fnames[0]
            for dup in fnames[1:]:
                if not os.path.isdir(dup_dir):
                    os.makedirs(dup_dir, exist_ok=True)
                src = os.path.join(pasta, dup)
                dst = os.path.join(dup_dir, dup)
                try:
                    shutil.move(src, dst)
                    duplicates.append((dup, keep))
                    print(f"  [DUPLICADO] {dup} (igual a {keep}) -> _duplicados/")
                except Exception as e:
                    print(f"  [WARN] Erro ao mover duplicado {dup}: {e}")

    return duplicates


def organizar_pasta(pasta):
    """Organize files in client folder with numeric prefixes for easy filing order."""
    # Order categories with keywords to match filenames
    ORDEM = [
        (1, "PETICAO INICIAL", ["peticao_bpc", "peticao_inicial", "01_peticao"]),
        (2, "PROCURACAO", ["procura", "pocuracao", "02_procuracao"]),
        (3, "CONTRATO DE HONORARIOS", ["contrato", "kit.assinado", "03_contrato", "CONTRATO DE HONORARIOS"]),
        (4, "DECLARACAO DE HIPOSSUFICIENCIA", ["pedido.gratuidade", "hipossufici", "declaracao_hipo",
                                                "gratuidade", "termo.gratuidade", "04_declaracao"]),
        (5, "DOCUMENTO DO MENOR", ["DOC. DE IDENTIFICAÇÃO-CERTIDÃO DE NASCIMENTO",
                             "DOC. DE IDENTIFICAÇÃO-RG-AUTORA", "DOC. IDENTIFICAÇÃO - AUTORA",
                             "DOCUMENTO DE IDENTIFICAÇÃO - RG E CPF",
                             "DOCUMENTO DE IDENTIFICAÇÃO-RG (", "IDENTIDADE (", "ID VERSO",
                             "RG - AUTOR", "RG - AUTORA", "rg-autora", "rg.autora",
                             "CPF - AUTORA", "rg.manuelly",
                             "certid", "05_doc_menor", "5- DOCUMENTO DO MENOR"]),
        (6, "DOCUMENTO DO RESPONSAVEL", ["DOC. DE IDENTIFICAÇÃO-RG-REPRESENTANTE", "DOC. DE IDENTIFICAÇÃO-CNH-REPRESENTANTE",
                                   "DOC. IDENTIFICAÇÃO - REPRESENTANTE", "DOC. IDENTIFICAÇÃO - GENITORA",
                                   "DOC. IDENTIFICAÇÃO - GENITOR", "DOCUMENTO DE IDENTIFICAÇÃO - RG - REPRESENTANTE",
                                   "IDENTIDADE - REPRESENTANTE", "RG - GENITORA",
                                   "CNH - REPRESENTANTE", "CNH.", "CHN.", "cnh.",
                                   "rg.representante", "rg.mae", "rg.pai", "rg.renta", "rg.marcia",
                                   "DOC. DE IDENTIFICAÇÃO-RG-PAI", "DOC. DE IDENTIFICAÇÃO-CNH-PAI",
                                   "DOCUMENTO DE IDENTIFICAÇÃO - RG - PAI",
                                   "06_doc_responsavel", "6- DOCUMENTO DO RESPONSAVEL"]),
        (7, "COMPROVANTE DE RESIDENCIA", ["conta.luz", "conta.agua", "conta.jan", "conta.fev",
                                          "comprovante.resid", "comprov.resid", "comprovante.renata",
                                          "COMPROVANTE DE RESIDENCIA", "COMP. DE RESIDENCIA", "COMP ENDE",
                                          "autodecl", "declaracao.aluguel",
                                          "DECLARAÇÃO DE RESIDENCIA", "declaracao.resid",
                                          "07_comprovante_residencia", "7- COMPROVANTE DE RESIDENCIA"]),
        (8, "CADUNICO", ["cadunico", "CADUNICO", "CADASTRO UNICO", "cad.unico", "08_cadunico"]),
        (9, "DOCUMENTOS DO GRUPO FAMILIAR", ["grupo_familiar",
                                       "DOC. DE IDENTIFICAÇÃO-CERTIDÃO DE NASCIMENTO-IRM",
                                       "DOC. DE IDENTIFICAÇÃO-CPF-IRM", "DOC. DE IDENTIFICAÇÃO-RG-IRM",
                                       "DOC. IDENTIFICAÇÃO - IRMAO", "DOC. IDENTIFICAÇÃO - PAULO",
                                       "RG - IRMÃ", "RG - IRMA", "IDENTIDADE - ENZO",
                                       "IDENTIDADE - RAPHAEL", "IDENTIDADE - LAURA",
                                       "IDENTIDADE CARLOS", "IDENTIDADE NATALINO", "IDENTIDADE GEOVANI",
                                       "IDENTIDADE DEIVID", "ID DEIVID",
                                       "CPF - PAULO", "CPF - EMANUEL", "rg.emanuel", "rg.irmao",
                                       "rg.irma", "rg.raissa", "cpf.raissa",
                                       "DOC. DE IDENTIFICAÇÃO-CPF-RAISSA", "DOC. DE IDENTIFICAÇÃO-RG-RAISSA",
                                       "certidao.nascimento.irm", "09_docs"]),
        (10, "COMPROVANTE DE RENDA", ["comprometimento_renda", "comprovante.renda", "extrato",
                                      "RENDA.pdf", "10_comprovante_renda"]),
        (11, "REQUERIMENTO INSS", ["PROCESSO INSS", "processo-", "relatorio_tarefa_",
                                   "requerimento", "protocolo.inss", "11_requerimento"]),
        (12, "CARTA DE INDEFERIMENTO", ["indeferimento", "carta.inss"]),
        (13, "LAUDO MEDICO", ["laudo", "LAUDO", "13_laudo"]),
        (14, "RELATORIO MEDICO", ["relatorio", "RELATÓRIO", "parecer", "ATESTADO",
                                  "AVALIAÇÃO PROFISSIONAL", "14_relatorio"]),
        (15, "RECEITAS E EXAMES", ["receita", "receituario", "receituário", "exame",
                                   "encaminhamento", "ENCAMINHAMENTO",
                                   "COMPROVANTE DE AGENDAMENTO", "COMPROVANTE-LISTA", "GUIA",
                                   "15_receitas"]),
        (16, "COMPROVANTE DE GASTOS", ["fatura", "FATURA", "nota.fiscal", "NOTAS FISCAIS",
                                       "comprovante.gasto", "COMPROVANTE DE GASTOS",
                                       "DEPESAS", "DESPESA", "16_comprovante"]),
        (17, "CALCULO DE ATRASADOS", ["calculo_atrasados", "17_calculo"]),
        (18, "QUESITOS PERICIA MEDICA", ["quesitos_pericia_medica", "quesitos_medic", "18_quesitos_medic"]),
        (19, "QUESITOS PERICIA SOCIAL", ["quesitos_pericia_social", "quesitos_soci", "19_quesitos_soci"]),
        (20, "BIOMETRIA", ["biometria", "BIOMETRIA"]),
        (21, "OAB", ["OAB (", "oab_", "carteira.oab"]),
        (22, "TITULO DE ELEITOR", ["titulo", "TITULO DE ELEITOR"]),
        (23, "TERMOS", ["termo.renuncia", "termo.responsabilidade", "termo.representacao"]),
    ]

    # Files to skip
    SKIP = ["desktop.ini", ".gitkeep"]

    files = [f for f in os.listdir(pasta)
             if os.path.isfile(os.path.join(pasta, f)) and f not in SKIP and not f.startswith('.')]

    renamed = []
    used_files = set()

    # Pre-match files already renamed with "N- LABEL" pattern
    already_renamed_re = re.compile(r'^(\d+)-\s+(.+?)(?:\s+\d+)?(\.[^.]+)$')

    for num, label, keywords in ORDEM:
        matches = []
        for f in files:
            if f in used_files:
                continue
            # Check if already renamed with correct prefix
            m = already_renamed_re.match(f)
            if m and int(m.group(1)) == num:
                used_files.add(f)  # Already correctly named, skip
                continue
            f_lower = f.lower()
            for kw in keywords:
                if kw.lower() in f_lower:
                    matches.append(f)
                    break

        for idx, f in enumerate(sorted(matches)):
            used_files.add(f)
            ext = os.path.splitext(f)[1]
            if len(matches) > 1:
                new_name = f"{num}- {label} {idx+1}{ext}"
            else:
                new_name = f"{num}- {label}{ext}"
            src = os.path.join(pasta, f)
            dst = os.path.join(pasta, new_name)
            try:
                if src != dst:
                    if os.path.exists(dst):
                        counter = 2
                        base, ext2 = os.path.splitext(new_name)
                        while os.path.exists(os.path.join(pasta, f"{base} {counter}{ext2}")):
                            counter += 1
                        new_name = f"{base} {counter}{ext2}"
                        dst = os.path.join(pasta, new_name)
                    os.rename(src, dst)
                    renamed.append((f, new_name))
                    print(f"  [RENOMEADO] {f} -> {new_name}")
            except Exception as e:
                print(f"  [WARN] Erro ao renomear {f}: {e}")

    # Files that didn't match any category - leave them as is
    unmatched = [f for f in files if f not in used_files]
    if unmatched:
        print(f"  [INFO] Arquivos nao categorizados: {', '.join(unmatched)}")

    return renamed


def extract_python_code(text):
    """Extract Python code block from Claude's response."""
    # Try ```python ... ``` first (greedy to get the largest block)
    match = re.search(r"```python\s*\n(.*?)```", text, re.DOTALL)
    if match:
        return match.group(1).strip()

    # Try ``` ... ```
    match = re.search(r"```\s*\n(.*?)```", text, re.DOTALL)
    if match:
        code = match.group(1).strip()
        if "import" in code or "def " in code:
            return code

    # If code block was truncated (no closing ```), extract everything after ```python
    match = re.search(r"```python\s*\n(.*)", text, re.DOTALL)
    if match:
        code = match.group(1).strip()
        # Remove trailing ``` if present at the very end
        code = re.sub(r"```\s*$", "", code).strip()
        if "import" in code:
            return code

    return None


# ==================== LEGALMAIL INTEGRATION ====================

LEGALMAIL_API_KEY = os.environ.get("LEGALMAIL_API_KEY", "")
LEGALMAIL_CERT_ID = int(os.environ.get("LEGALMAIL_CERTIFICADO_ID", "0"))
LEGALMAIL_BASE = "https://app.legalmail.com.br/api/v1"

# Map organized file prefixes to LegalMail attachment type IDs
# Map file prefix -> attachment type NAME (resolved to ID dynamically per tribunal)
LEGALMAIL_DOC_TYPE_NAMES = {
    "2": "Procuração",
    "3": "Contrato de Honorários",
    "4": "Declaração de Hipossuficiência",
    "5": "Certidão de Nascimento",
    "6": "Identidade",
    "7": "Comprovante de Residência",
    "8": "Comprovantes",
    "9": "Identidade",
    "10": "Comprovantes",
    "11": "Comprovantes",
    "12": "Carta de Indeferimento",
    "13": "Laudo",
    "14": "Laudo",
    "15": "Receituário",
    "16": "Comprovantes",
    "17": "Planilha",
    "18": "Quesitos Perícia",
    "19": "Quesitos Perícia",
    "20": "Identidade",
    "21": "Identidade",
    "23": "Outros",
}

# Alternative search terms for each prefix (used when exact name doesn't match)
_DOC_TYPE_SEARCH_TERMS = {
    "2": ["procura", "mandato", "substabelec", "instrumento"],
    "3": ["honor", "contrato"],
    "4": ["hipossufic", "pobreza", "declara"],
    "5": ["certid", "nascimento"],
    "6": ["identidade", "rg", "documento", "cnh"],
    "7": ["resid", "endereco", "comprovante"],
    "8": ["cadastro", "cadunico", "cad", "comprovante"],
    "9": ["identidade", "rg", "documento", "familiar"],
    "10": ["renda", "comprovante", "extrato"],
    "11": ["requerimento", "inss", "protocolo", "comprovante"],
    "12": ["indeferimento", "carta", "comunica"],
    "13": ["laudo", "medico", "atestado", "perica"],
    "14": ["relatorio", "laudo", "medico"],
    "15": ["receita", "prescri", "exame"],
    "16": ["comprovante", "extrato"],
    "17": ["planilha", "calculo", "atrasado"],
    "18": ["quesito", "perica", "medic"],
    "19": ["quesito", "perica", "social"],
    "20": ["biometria", "identidade", "foto"],
    "21": ["oab", "identidade", "carteira"],
    "23": ["termo", "outro", "declaracao"],
}

# Cache: petition_id -> {type_name_lower: type_id}
_attachment_type_cache = {}


def legalmail_resolve_doc_type(idpeticoes, prefix):
    """Get the correct attachment type ID for a file prefix, fetching from API if needed.
    Uses fuzzy matching with alternative search terms for PJe compatibility."""
    global _attachment_type_cache

    if idpeticoes not in _attachment_type_cache:
        import time as _time_att
        r = legalmail_request("get", f"/petition/attachment/types?idpeticoes={idpeticoes}")
        if r.status_code == 429:
            print(f"  [LEGALMAIL] Rate limit nos tipos de anexo, aguardando 60s...")
            _time_att.sleep(60)
            r = legalmail_request("get", f"/petition/attachment/types?idpeticoes={idpeticoes}")
        if r.status_code != 200:
            print(f"  [WARN] Falha ao buscar tipos de anexo: {r.status_code}")
            return None
        types = r.json() if r.text.strip().startswith('[') else []
        type_map = {}
        for t in types:
            name = t.get('nome', '').strip()
            tid = str(t.get('iddocumentos_tipos', ''))
            type_map[name.lower()] = tid
        _attachment_type_cache[idpeticoes] = type_map

    type_map = _attachment_type_cache[idpeticoes]
    target_name = LEGALMAIL_DOC_TYPE_NAMES.get(prefix, "Outros")
    target_lower = _normalize_text(target_name) if '_normalize_text' in dir() else target_name.lower()

    # 1. Exact match (accent-insensitive)
    for name, tid in type_map.items():
        if _normalize_text(name) == target_lower:
            return tid

    # 2. Partial match: target in name or name in target
    for name, tid in type_map.items():
        name_n = _normalize_text(name)
        if target_lower in name_n or name_n in target_lower:
            return tid

    # 3. Alternative search terms for this prefix
    search_terms = _DOC_TYPE_SEARCH_TERMS.get(prefix, [])
    for term in search_terms:
        for name, tid in type_map.items():
            if term in _normalize_text(name):
                return tid

    # 4. Common fallback type names
    for fallback in ['outros', 'outro', 'documento', 'documentos', 'comprovante',
                     'comprovantes', 'anexo', 'anexos', 'peticao', 'petição']:
        for name, tid in type_map.items():
            if fallback in _normalize_text(name):
                return tid

    # 5. Absolute last resort: "petição inicial" or first type
    if type_map:
        # Prefer "petição" types
        for name, tid in type_map.items():
            if 'peti' in _normalize_text(name) and 'inicial' in _normalize_text(name):
                return tid
        first_name = next(iter(type_map))
        print(f"  [LEGALMAIL] Tipo '{target_name}' nao encontrado, usando '{first_name}'")
        return type_map[first_name]

    print(f"  [WARN] Nenhum tipo de anexo disponivel para '{target_name}'")
    return None

# BPC/LOAS defaults per system
# ATENÇÃO: NUNCA classificar como benefício previdenciário por incapacidade (remessa indevida à Central de Perícias)
BPC_DEFAULTS = {
    'eproc': {
        'rito': 'RITO ORDINÁRIO (COMUM)',
        'classe': 'PROCEDIMENTO COMUM',
        'competencia': 'Federal',
        'assunto_search': 'Benefício Assistencial',
    },
    'pje': {
        'rito': 'JUIZADO ESPECIAL FEDERAL',
        'classe': 'PROCEDIMENTO DO JUIZADO ESPECIAL CÍVEL (436)',
        'area': 'DIREITO PREVIDENCIÁRIO',
        'assunto_search': 'Pessoa com Deficiência (11946)',
    },
}
# eProc TRF-4: classes endpoint returns EMPTY - classe/assunto must be filled manually in LegalMail
# Legacy constant kept for reference
LEGALMAIL_ASSUNTO_BPC_PJE = "DIREITO ADMINISTRATIVO E OUTRAS MATÉRIAS DE DIREITO PÚBLICO (9985) | Garantias Constitucionais (9986) | Assistência Social (11847)"

# ==================== UF -> TRIBUNAL MAPPING ====================
# Maps Brazilian state to correct TRF, sistema, and default comarca
UF_TRIBUNAL_MAP = {
    # TRF-1 (DF, GO, BA, PI, MA, PA, AM, AC, AP, RR, RO, TO, MT) - sistema: pje
    'DF': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'DF'},
    'GO': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'GO'},
    'BA': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'BA'},
    'PI': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'PI'},
    'MA': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'MA'},
    'PA': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'PA'},
    'AM': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'AM'},
    'AC': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'AC'},
    'AP': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'AP'},
    'RR': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'RR'},
    'RO': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'RO'},
    'TO': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'TO'},
    'MT': {'trf': 'TRF-1', 'sistema': 'pje', 'uf_tribunal': 'MT'},
    # TRF-2 (RJ, ES) - eProc with UF-specific sistema
    'RJ': {'trf': 'TRF-2', 'sistema': 'eproc_jfrj', 'uf_tribunal': 'RJ'},
    'ES': {'trf': 'TRF-2', 'sistema': 'eproc_jfes', 'uf_tribunal': 'ES'},
    # TRF-3 (SP, MS) - PJe
    'SP': {'trf': 'TRF-3', 'sistema': 'pje', 'uf_tribunal': 'SP'},
    'MS': {'trf': 'TRF-3', 'sistema': 'pje', 'uf_tribunal': 'MS'},
    # TRF-4 (PR, SC, RS) - eProc with UF-specific sistema
    'PR': {'trf': 'TRF-4', 'sistema': 'eproc_jfpr', 'uf_tribunal': 'PR'},
    'SC': {'trf': 'TRF-4', 'sistema': 'eproc_jfsc', 'uf_tribunal': 'SC'},
    'RS': {'trf': 'TRF-4', 'sistema': 'eproc_jfrs', 'uf_tribunal': 'RS'},
    # TRF-5 (PE, CE, AL, SE, RN, PB) - sistema: pje
    'PE': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'PE'},
    'CE': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'CE'},
    'AL': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'AL'},
    'SE': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'SE'},
    'RN': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'RN'},
    'PB': {'trf': 'TRF-5', 'sistema': 'pje', 'uf_tribunal': 'PB'},
    # TRF-6 (MG) - eProc with UF-specific sistema
    'MG': {'trf': 'TRF-6', 'sistema': 'eproc_jfmg', 'uf_tribunal': 'MG'},
}

def detect_uf_from_folder(pasta):
    """Detect client's UF from documents in organized folder.
    Reads comprovante de residencia or CadUnico via OCR."""
    import re as _re
    files = sorted(os.listdir(pasta))

    # Priority: comprovante de residencia, then cadunico, then autodeclaracao
    for f in files:
        fl = f.lower()
        if '7- comprovante de residencia' in fl or '8- cadunico' in fl:
            text = mistral_ocr(os.path.join(pasta, f))
            if not text:
                continue
            # Search for UF patterns
            ufs = ['SP', 'RJ', 'MG', 'PR', 'SC', 'RS', 'GO', 'MT', 'MS', 'BA', 'PE',
                   'CE', 'PA', 'AM', 'MA', 'PI', 'RN', 'PB', 'SE', 'AL', 'TO', 'RO',
                   'AC', 'AP', 'RR', 'ES', 'DF']
            # Pattern: city/UF, city-UF, city - UF, CEP XXXXX-XXX CITY UF
            for uf in ufs:
                if _re.search(rf'(?:[-/]\s*{uf}\b|\b{uf}\s*[-/]|\b{uf}\s+\d{{5}}|\d{{5}}-?\d{{3}}\s+\w+\s+{uf}\b|\b{uf}\s*CEP)', text):
                    return uf
            # Try state names
            state_map = {'SÃO PAULO': 'SP', 'SAO PAULO': 'SP', 'MINAS GERAIS': 'MG',
                         'PARANÁ': 'PR', 'PARANA': 'PR', 'SANTA CATARINA': 'SC',
                         'RIO GRANDE DO SUL': 'RS', 'GOIÁS': 'GO', 'GOIAS': 'GO',
                         'MATO GROSSO DO SUL': 'MS', 'MATO GROSSO': 'MT',
                         'ESPÍRITO SANTO': 'ES', 'ESPIRITO SANTO': 'ES',
                         'RIO DE JANEIRO': 'RJ', 'BAHIA': 'BA',
                         'PERNAMBUCO': 'PE', 'CEARÁ': 'CE', 'CEARA': 'CE',
                         'PARÁ': 'PA', 'PARA': 'PA', 'AMAZONAS': 'AM',
                         'MARANHÃO': 'MA', 'MARANHAO': 'MA', 'PIAUÍ': 'PI', 'PIAUI': 'PI',
                         'RIO GRANDE DO NORTE': 'RN', 'PARAÍBA': 'PB', 'PARAIBA': 'PB',
                         'SERGIPE': 'SE', 'ALAGOAS': 'AL', 'TOCANTINS': 'TO',
                         'RONDÔNIA': 'RO', 'RONDONIA': 'RO', 'ACRE': 'AC',
                         'AMAPÁ': 'AP', 'AMAPA': 'AP', 'RORAIMA': 'RR',
                         'DISTRITO FEDERAL': 'DF'}
            text_upper = text.upper()
            for state, uf in state_map.items():
                if state in text_upper:
                    return uf
    return None

def detect_cidade_from_folder(pasta):
    """Detect client's city from address documents via OCR."""
    import re as _re
    files = sorted(os.listdir(pasta))
    for f in files:
        fl = f.lower()
        if '7- comprovante de residencia' in fl or '8- cadunico' in fl:
            text = mistral_ocr(os.path.join(pasta, f))
            if not text:
                continue
            # Pattern: CIDADE/UF or CIDADE - UF or CIDADE-UF
            ufs = ['SP', 'RJ', 'MG', 'PR', 'SC', 'RS', 'GO', 'MT', 'MS', 'BA', 'PE',
                   'CE', 'PA', 'AM', 'MA', 'PI', 'RN', 'PB', 'SE', 'AL', 'TO', 'RO',
                   'AC', 'AP', 'RR', 'ES', 'DF']
            for uf in ufs:
                # Match: "Cidade/UF" or "Cidade - UF" or "Cidade-UF"
                m = _re.search(rf'([A-ZÀ-Ú][a-zà-ú]+(?:\s+[a-zà-ú]+)*(?:\s+[A-ZÀ-Ú][a-zà-ú]+)*)\s*[-/]\s*{uf}\b', text)
                if m:
                    cidade = m.group(1).strip()
                    if len(cidade) > 2:
                        return cidade
            # Try CEP-based: "XXXXX-XXX Cidade UF"
            m = _re.search(r'\d{5}-?\d{3}\s+([A-ZÀ-Ú][a-zà-ú]+(?:\s+[a-zà-ú]+)*(?:\s+[A-ZÀ-Ú][a-zà-ú]+)*)', text)
            if m:
                return m.group(1).strip()
    return None

def get_tribunal_config(uf):
    """Get tribunal config from UF. Returns dict with trf, sistema, uf_tribunal."""
    return UF_TRIBUNAL_MAP.get(uf, {'trf': 'TRF-3', 'sistema': 'pje', 'uf_tribunal': uf or 'SP'})

# INSS data (polo passivo) - fixed for all BPC cases
INSS_DATA = {
    "nome": "INSTITUTO NACIONAL DO SEGURO SOCIAL - INSS",
    "polo": "passivo",
    "documento": "29.979.036/0001-40",
    "personalidade": "Pessoa jur\u00eddica",
    "endereco_cep": "70040-902",
    "endereco_logradouro": "SAUS Quadra 2 Bloco O",
    "endereco_numero": "S/N",
    "endereco_bairro": "Asa Sul",
    "endereco_cidade": "Brasilia",
    "endereco_uf": "DF",
}
# Cached INSS party ID (set on first use)
_inss_party_id = None


def _parse_brl(val):
    """Parse a BRL value string or number to float. Handles 'R$ 29.231,70' and '29231.70'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).strip()
    s = s.replace('R$', '').replace(' ', '').strip()
    if not s:
        return None
    # Detect format: "29.231,70" (BR) vs "29231.70" (US) vs "29,231.70" (US with commas)
    if ',' in s and '.' in s:
        if s.rindex(',') > s.rindex('.'):
            # BR format: 29.231,70
            s = s.replace('.', '').replace(',', '.')
        else:
            # US format: 29,231.70
            s = s.replace(',', '')
    elif ',' in s and '.' not in s:
        # Could be "29231,70" (BR decimal) or "29,231" (US thousands)
        # If comma is near end (1-2 digits after), treat as decimal
        parts = s.split(',')
        if len(parts[-1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def extract_valor_causa(pasta):
    """Extract valor da causa from the xlsx spreadsheet in client folder."""
    try:
        import openpyxl
        xlsx_files = [f for f in os.listdir(pasta) if f.endswith('.xlsx') and 'CALCULO' in f.upper()]
        if not xlsx_files:
            return None
        xlsx_path = os.path.join(pasta, xlsx_files[0])
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # Strategy 1: Find "TOTAL DA CONTA" (not SUBTOTAL) - check embedded and columns
        total_candidates = []
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=6, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    text_upper = cell.value.upper()
                    # Match "TOTAL DA CONTA" but NOT "SUBTOTAL"
                    if 'TOTAL DA CONTA' in text_upper and 'SUBTOTAL' not in text_upper:
                        # Check embedded value in same cell
                        if 'R$' in cell.value:
                            val = _parse_brl(cell.value.split('R$')[-1])
                            if val and val > 100:
                                total_candidates.append(val)
                        # Check columns B-E of same row
                        for col in [5, 4, 3, 2]:
                            val = _parse_brl(ws.cell(row=cell.row, column=col).value)
                            if val and val > 100:
                                total_candidates.append(val)

        if total_candidates:
            wb.close()
            return max(total_candidates)  # Take the largest (total > subtotal)

        # Strategy 2: Find any "TOTAL" cell with embedded R$ value (skip SUBTOTAL)
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=6, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    text_upper = cell.value.upper()
                    if 'TOTAL' in text_upper and 'SUBTOTAL' not in text_upper and 'R$' in cell.value:
                        val = _parse_brl(cell.value.split('R$')[-1])
                        if val and val > 100:
                            wb.close()
                            return val

        # Strategy 3: Check common positions (E13, E12, E11, B11, B12)
        for r, c in [(13, 5), (12, 5), (11, 5), (12, 2), (11, 2)]:
            val = _parse_brl(ws.cell(row=r, column=c).value)
            if val and val > 100:
                wb.close()
                return val

        # Strategy 4: Scan all cells for the largest monetary value (likely total)
        max_val = None
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=6, values_only=False):
            for cell in row:
                val = _parse_brl(cell.value)
                if val and val > 100:
                    if max_val is None or val > max_val:
                        max_val = val
        wb.close()
        return max_val

    except Exception as e:
        print(f"  [WARN] extract_valor_causa: {e}")
        try:
            wb.close()
        except Exception:
            pass
    return None


def legalmail_get_or_create_inss():
    """Find or create INSS party on LegalMail. Returns party ID."""
    global _inss_party_id
    if _inss_party_id is not None:
        return _inss_party_id

    import requests
    # Search existing parties for INSS
    try:
        r = legalmail_request("get", "/parts")
        if r.status_code == 200:
            parts = r.json()
            for p in parts:
                if '29.979.036' in str(p.get('documento', '')) or 'INSS' in str(p.get('nome', '')):
                    _inss_party_id = int(p.get('id'))
                    print(f"  [LEGALMAIL] INSS encontrado: id={_inss_party_id}")
                    return _inss_party_id
    except Exception:
        pass

    # Create INSS party
    r = legalmail_request("post", "/parts", json=INSS_DATA)
    if r.status_code == 200:
        data = r.json()
        _inss_party_id = data.get('id')
        print(f"  [LEGALMAIL] INSS criado: id={_inss_party_id}")
        return _inss_party_id
    else:
        print(f"  [WARN] Falha ao criar INSS: {r.status_code} {r.text[:200]}")
    return None


def _detect_genero(nome):
    """Detect gender from Brazilian first name."""
    if not nome:
        return 'MASCULINO'
    primeiro = nome.strip().split()[0].upper()
    # Common feminine endings and names
    fem_names = {'ALICE', 'ISABELLA', 'MARIA', 'ANA', 'JULIA', 'JULIANA', 'LUCIANA',
                 'LUCIENE', 'LUCELIA', 'REGIANE', 'REGIALVA', 'ANGELA', 'ADRIANA',
                 'BEATRIZ', 'CAMILA', 'DANIELA', 'EDUARDA', 'FERNANDA', 'GABRIELA',
                 'HELENA', 'ISABELA', 'JESSICA', 'LARISSA', 'LETICIA', 'LUANA',
                 'MARIANA', 'NATALIA', 'PATRICIA', 'RAFAELA', 'SABRINA', 'TATIANA',
                 'VALENTINA', 'VITORIA', 'YASMIN'}
    if primeiro in fem_names:
        return 'FEMININO'
    # Feminine endings
    if primeiro.endswith(('A', 'IA', 'NA', 'NE', 'LA', 'LIA', 'RIA', 'INA', 'ELA', 'ICA')):
        # Exclude common masculine names ending in A
        masc_a = {'LUCA', 'JOSUA', 'NOA', 'EZA'}
        if primeiro not in masc_a:
            return 'FEMININO'
    return 'MASCULINO'


def legalmail_create_party(party_data):
    """Create a party on LegalMail. Returns party ID (int) or None.
    Auto-fills required fields: personalidade, endereco_logradouro, genero."""
    data = dict(party_data)

    # Ensure personalidade is set with correct accent
    if 'personalidade' not in data or not data['personalidade']:
        doc = data.get('documento', '')
        if '/' in doc and len(doc.replace('.','').replace('-','').replace('/','')) == 14:
            data['personalidade'] = 'Pessoa jur\u00eddica'
        else:
            data['personalidade'] = 'Pessoa f\u00edsica'
    elif data['personalidade'].lower().startswith('pessoa f'):
        data['personalidade'] = 'Pessoa f\u00edsica'
    elif data['personalidade'].lower().startswith('pessoa j'):
        data['personalidade'] = 'Pessoa jur\u00eddica'

    # Ensure required address fields
    if not data.get('endereco_logradouro'):
        data['endereco_logradouro'] = 'A informar'
    if not data.get('endereco_numero'):
        data['endereco_numero'] = 'S/N'
    if not data.get('endereco_bairro'):
        data['endereco_bairro'] = 'A informar'
    if not data.get('endereco_cidade'):
        data['endereco_cidade'] = 'A informar'
    if not data.get('endereco_uf'):
        data['endereco_uf'] = 'SP'
    if not data.get('endereco_cep'):
        data['endereco_cep'] = '01000-000'

    # Auto-detect genero if not set
    if not data.get('genero'):
        data['genero'] = _detect_genero(data.get('nome', ''))

    r = legalmail_request("post", "/parts", json=data)
    if r.status_code == 200:
        resp = r.json()
        raw_id = resp.get('id')
        if raw_id is None:
            print(f"  [WARN] API retornou sem 'id': {resp}")
            return None
        pid = int(raw_id)
        print(f"  [LEGALMAIL] Parte criada: {data.get('nome', '?')} id={pid}")
        return pid
    else:
        print(f"  [WARN] Falha ao criar parte: {r.status_code} {r.text[:200]}")
    return None


def legalmail_find_party_by_doc(documento):
    """Find existing party by CPF/CNPJ using search endpoint. Returns party ID or None."""
    doc_clean = documento.replace('.', '').replace('-', '').replace('/', '')
    try:
        # Try search endpoint first (more efficient)
        r = legalmail_request("get", f"/parts/search?documento={doc_clean}")
        if r.status_code == 200:
            parts = r.json() if isinstance(r.json(), list) else r.json().get('data', [])
            if parts:
                return parts[0].get('id')
        # Fallback: paginated listing
        offset = 0
        while True:
            r = legalmail_request("get", f"/parts?limit=50&offset={offset}")
            if r.status_code != 200:
                break
            parts = r.json() if isinstance(r.json(), list) else []
            if not parts:
                break
            for p in parts:
                p_doc = str(p.get('documento', '')).replace('.', '').replace('-', '').replace('/', '')
                if p_doc == doc_clean:
                    return p.get('id')
            if len(parts) < 50:
                break
            offset += 50
    except Exception:
        pass
    return None


def _normalize_text(s):
    """Remove accents for comparison."""
    import unicodedata
    return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii').lower().strip()


def _match_comarca(target, options):
    """Fuzzy match comarca name against available options, ignoring accents."""
    target_n = _normalize_text(target)
    # Exact match (ignoring accents)
    for opt in options:
        if _normalize_text(opt) == target_n:
            return opt
    # Contains match (either direction)
    for opt in options:
        opt_n = _normalize_text(opt)
        if target_n in opt_n or opt_n in target_n:
            return opt
    # Word overlap
    target_words = set(target_n.split())
    best_score, best_opt = 0, None
    for opt in options:
        opt_words = set(_normalize_text(opt).split())
        score = len(target_words & opt_words)
        if score > best_score:
            best_score, best_opt = score, opt
    if best_opt and best_score > 0:
        return best_opt
    return None


def legalmail_fill_fields(idpeticoes, sistema, comarca_name, valor_causa=None,
                           id_polo_ativo=None, id_polo_passivo=None,
                           tipo_beneficio='deficiente',
                           uf_tribunal='', tribunal='', instancia='1'):
    """Fill all petition fields via sequential PUT/GET flow.

    Handles both eProc and PJe systems with their different field order requirements.
    eProc: comarca -> rito -> competencia -> classe -> assunto -> area
    PJe: competencia (unavailable via API) -> comarca -> classe -> assunto -> area

    tipo_beneficio: 'deficiente' or 'idoso' (for assunto selection)
    Returns dict with status.
    """
    resolved = {}
    errors = []
    is_pje = 'pje' in sistema.lower() and 'eproc' not in sistema.lower()

    put_endpoint = f"/petition/initial?idpeticoes={idpeticoes}"

    def safe_get(endpoint, label):
        """GET with error handling. Returns list or empty list."""
        r = legalmail_request("get", endpoint)
        if r.status_code == 422:
            return []
        if r.status_code == 200:
            try:
                data = r.json()
                if isinstance(data, list):
                    return data
            except Exception:
                pass
        return []

    def find_best_match(options, targets, key='nome'):
        """Find best match from options list by trying targets in priority order."""
        for target in targets:
            target_n = _normalize_text(target)
            match = [o for o in options if target_n in _normalize_text(o.get(key, ''))]
            if match:
                return match[0].get(key, '')
        return options[0].get(key, '') if options else None

    def do_put(payload, label):
        """PUT partial update with retry on 429."""
        r = legalmail_request("put", put_endpoint, json=payload)
        if r.status_code == 429:
            import time as _t
            _t.sleep(65)
            r = legalmail_request("put", put_endpoint, json=payload)
        if r.status_code == 200:
            print(f"  [LEGALMAIL] PUT {label}: OK")
        else:
            msg = f"PUT {label}: {r.status_code} - {r.text[:200]}"
            print(f"  [LEGALMAIL] {msg}")
            errors.append(msg)
        return r

    print(f"  [LEGALMAIL] === Preenchendo campos para peticao {idpeticoes} ({('PJe' if is_pje else 'eProc')}) ===")

    # ========== eProc flow: comarca -> rito -> competencia -> classe -> assunto -> area ==========
    if not is_pje:
        # STEP 1: Comarca
        if comarca_name:
            comarcas = safe_get(f"/petition/county?idpeticoes={idpeticoes}", 'comarcas')
            if comarcas:
                comarca_names = [c.get('nome', '') for c in comarcas]
                best = _match_comarca(comarca_name, comarca_names)
                resolved['comarca'] = best or comarca_names[0]
                if not best:
                    print(f"  [WARN] Comarca '{comarca_name}' nao encontrada, usando '{comarca_names[0]}'")
            else:
                resolved['comarca'] = comarca_name
            print(f"  [LEGALMAIL] -> comarca: {resolved['comarca']}")
            do_put({'comarca': resolved['comarca']}, 'comarca')

        # STEP 2: Rito
        ritos = safe_get(f"/petition/ritos?idpeticoes={idpeticoes}", 'ritos')
        if ritos:
            rito = find_best_match(ritos, ['JUIZADO ESPECIAL FEDERAL', 'JUIZADO ESPECIAL', 'ORDINÁRIO', 'COMUM'])
            if rito:
                resolved['rito'] = rito
                print(f"  [LEGALMAIL] -> rito: {rito}")
                do_put({'rito': rito}, 'rito')

        # STEP 3: Competencia
        specialties = safe_get(f"/petition/specialties?idpeticoes={idpeticoes}", 'specialties')
        if specialties:
            comp = find_best_match(specialties, ['FEDERAL', 'CÍVEL', 'PREVIDENCIÁRIO'])
            if comp:
                resolved['competencia'] = comp
                print(f"  [LEGALMAIL] -> competencia: {comp}")
                do_put({'competencia': comp}, 'competencia')

        # STEP 4: Classe
        classes = safe_get(f"/petition/classes?idpeticoes={idpeticoes}", 'classes')
        if classes:
            classe = find_best_match(classes, ['JUIZADO ESPECIAL', 'PROCEDIMENTO DO JUIZADO', 'PROCEDIMENTO COMUM'])
            if classe:
                resolved['classe'] = classe
                print(f"  [LEGALMAIL] -> classe: {classe}")
                do_put({'classe': classe}, 'classe')

        # STEP 5: Assunto (may fail if classe not set — eProc returns empty classes)
        subjects = safe_get(f"/petition/subjects?idpeticoes={idpeticoes}", 'subjects')
        if subjects:
            assunto = _find_bpc_assunto(subjects, tipo_beneficio)
            if assunto:
                resolved['assunto'] = assunto
                print(f"  [LEGALMAIL] -> assunto: {assunto[:60]}")
                do_put({'assunto': assunto}, 'assunto')
        elif 'classe' not in resolved:
            # eProc doesn't expose classes via API; assunto depends on classe.
            # Both must be filled manually in LegalMail UI.
            print(f"  [LEGALMAIL] eProc: classe/assunto nao disponiveis via API, preencher manualmente")

        # STEP 6: Area
        areas = safe_get(f"/petition/areas?idpeticoes={idpeticoes}", 'areas')
        if areas:
            area = find_best_match(areas, ['CÍVEL', 'PREVIDENCIÁRIO'])
            if area:
                resolved['area'] = area
                print(f"  [LEGALMAIL] -> area: {area}")
                do_put({'area': area}, 'area')

    # ========== PJe flow ==========
    # PJe has a circular dependency: competencia requires comarca, comarca requires
    # competencia. The specialties endpoint returns 422 without comarca set.
    # Solution: resolve classe + assunto from complaintsandpleadings endpoints
    # (available without prerequisites), set comarca as text, and include all in
    # final PUT. competencia/area/rito must be completed manually in LegalMail UI.
    else:
        # STEP 1: Resolve classe from complaintsandpleadings (always available)
        classes = safe_get(f"/complaintsandpleadings/classes?idpeticoes={idpeticoes}", 'classes')
        if not classes:
            classes = safe_get(f"/petition/classes?idpeticoes={idpeticoes}", 'classes')
        if classes:
            classe = find_best_match(classes,
                ['PROCEDIMENTO DO JUIZADO ESPECIAL CÍVEL',
                 'PROCEDIMENTO DO JUIZADO ESPECIAL',
                 'PROCEDIMENTO COMUM CÍVEL',
                 'PROCEDIMENTO COMUM'])
            if classe:
                resolved['classe'] = classe
                print(f"  [LEGALMAIL] -> classe: {classe}")

        # STEP 2: Resolve assunto from complaintsandpleadings (always available)
        subjects = safe_get(f"/complaintsandpleadings/subjects?idpeticoes={idpeticoes}", 'subjects')
        if not subjects:
            subjects = safe_get(f"/petition/subjects?idpeticoes={idpeticoes}", 'subjects')
        if subjects:
            assunto = _find_bpc_assunto(subjects, tipo_beneficio)
            if assunto:
                resolved['assunto'] = assunto
                print(f"  [LEGALMAIL] -> assunto: {assunto[:60]}")

        # STEP 3: Set comarca as text
        if comarca_name:
            resolved['comarca'] = comarca_name

        # NOTE: competencia, area, rito are blocked by API circular dependency.
        # They MUST be completed manually in LegalMail UI after draft creation.
        print(f"  [LEGALMAIL] PJe: competencia/area/rito precisam ser preenchidos manualmente no LegalMail")

    # ========== FINAL payload with ALL required fields ==========
    resolved['gratuidade'] = True
    resolved['liminar'] = False
    resolved['100digital'] = True
    resolved['renuncia60Salarios'] = True
    resolved['distribuicao'] = 'Por sorteio'

    if valor_causa:
        resolved['valorCausa'] = f'{valor_causa:.2f}' if isinstance(valor_causa, float) else str(valor_causa)
    if uf_tribunal:
        resolved['ufTribunal'] = uf_tribunal
    if tribunal:
        resolved['tribunal'] = tribunal
    if sistema:
        resolved['sistema'] = sistema
    if instancia:
        resolved['instancia'] = instancia

    if id_polo_ativo:
        resolved['idpoloativo'] = [id_polo_ativo] if isinstance(id_polo_ativo, int) else id_polo_ativo
    if id_polo_passivo:
        resolved['idpolopassivo'] = [id_polo_passivo] if isinstance(id_polo_passivo, int) else id_polo_passivo

    # FINAL PUT
    print(f"  [LEGALMAIL] PUT final com {len(resolved)} campos: {list(resolved.keys())}")
    r = legalmail_request("put", put_endpoint, json=resolved)
    if r.status_code == 429:
        import time as _t
        _t.sleep(65)
        r = legalmail_request("put", put_endpoint, json=resolved)
    if r.status_code == 200:
        print(f"  [LEGALMAIL] PUT final: OK")
    else:
        msg = f"PUT final: {r.status_code} {r.text[:200]}"
        errors.append(msg)
        print(f"  [ERRO] {msg}")

    return {"filled": resolved, "errors": errors}


def _find_bpc_assunto(subjects, tipo_beneficio='deficiente'):
    """Find the best BPC/LOAS assunto from subjects list.
    NEVER use 'beneficio previdenciario por incapacidade' (causes wrong routing)."""
    search_term = 'defici' if tipo_beneficio == 'deficiente' else 'idoso'
    # Priority 1: "assistencial" + "203" + deficiente/idoso
    for s in subjects:
        nome = s.get('nome', '')
        nl = nome.lower()
        if 'assistencial' in nl and '203' in nome and search_term in nl:
            return nome
    # Priority 2: "assistencial" + deficiente/idoso
    for s in subjects:
        nome = s.get('nome', '')
        nl = nome.lower()
        if 'assistencial' in nl and search_term in nl:
            return nome
    # Priority 3: "Pessoa com Deficiência" (PJe specific)
    for s in subjects:
        nome = s.get('nome', '')
        if 'Pessoa com Defici' in nome or '11946' in nome:
            return nome
    # Priority 4: any "assistencial" (not incapacidade/previdenciario)
    for s in subjects:
        nome = s.get('nome', '')
        nl = nome.lower()
        if 'assistencial' in nl and 'incapacidade' not in nl:
            return nome
    # Priority 5: "DIREITO ASSISTENCIAL" category
    for s in subjects:
        nome = s.get('nome', '')
        if 'ASSISTENCIAL' in nome.upper():
            return nome
    return None


def docx_to_pdf(docx_path):
    """Convert docx to PDF using Word COM automation."""
    try:
        import subprocess
        abs_path = os.path.abspath(docx_path).replace('/', '\\')
        pdf_path = os.path.splitext(abs_path)[0] + '.pdf'
        import tempfile
        ps_file = os.path.join(tempfile.gettempdir(), '_docx2pdf.ps1')
        with open(ps_file, 'w', encoding='utf-8-sig') as pf:
            pf.write(f'''
$ErrorActionPreference = "Stop"
$word = New-Object -ComObject Word.Application
$word.Visible = $false
$word.DisplayAlerts = 0
try {{
    $doc = $word.Documents.Open("{abs_path}")
    $doc.SaveAs2("{pdf_path}", 17)
    $doc.Close($false)
    Write-Output "OK"
}} catch {{
    Write-Output ("ERRO: " + $_.Exception.Message)
}} finally {{
    $word.Quit()
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($word) | Out-Null
}}
''')
        result = subprocess.run(
            ['powershell', '-ExecutionPolicy', 'Bypass', '-File', ps_file],
            capture_output=True, timeout=60)
        stdout = result.stdout.decode('utf-8', errors='replace').strip()
        if 'OK' in stdout:
            print(f"  [PDF] {os.path.basename(pdf_path)} gerado (docx)")
            return pdf_path
        else:
            print(f"  [WARN] docx_to_pdf: {stdout}")
    except Exception as e:
        print(f"  [WARN] docx_to_pdf falhou: {e}")
    return None


_last_legalmail_call = [0.0]  # Track last API call time

def legalmail_request(method, endpoint, **kwargs):
    """Make authenticated request to LegalMail API. Respects 30 req/min limit."""
    import requests as _requests
    import time as _t
    # Enforce minimum 2.1s between calls (30 req/min = 1 req/2s)
    elapsed = _t.time() - _last_legalmail_call[0]
    if elapsed < 2.1:
        _t.sleep(2.1 - elapsed)
    _last_legalmail_call[0] = _t.time()

    sep = '&' if '?' in endpoint else '?'
    url = f"{LEGALMAIL_BASE}{endpoint}{sep}api_key={LEGALMAIL_API_KEY}"
    try:
        resp = getattr(_requests, method)(url, **kwargs, timeout=30)
        # Handle rate limiting with auto-retry
        if resp.status_code == 429:
            print(f"  [LEGALMAIL] Rate limit em {method.upper()} {endpoint}, aguardando 65s...")
            _t.sleep(65)
            resp = getattr(_requests, method)(url, **kwargs, timeout=30)
        # Wrap json() to handle empty responses
        _orig_json = resp.json
        def _safe_json(**kw):
            try:
                return _orig_json(**kw)
            except Exception:
                return {} if method.lower() in ('put', 'post', 'delete') else []
        resp.json = _safe_json
        return resp
    except _requests.exceptions.RequestException as e:
        print(f"  [LEGALMAIL] Erro de rede em {method.upper()} {endpoint}: {e}")
        class FakeResp:
            status_code = 503
            text = str(e)
            def json(self): return {}
        return FakeResp()


def resolve_folder_path(pasta):
    """Resolve folder path handling Windows encoding issues with accented chars.
    Returns the actual path that exists on disk, or the original if not found."""
    if os.path.isdir(pasta):
        return pasta
    parent = os.path.dirname(pasta)
    target_base = os.path.basename(pasta)
    if not os.path.isdir(parent):
        return pasta
    # Try matching by normalized name
    target_n = _normalize_text(target_base)
    for d in os.listdir(parent):
        if os.path.isdir(os.path.join(parent, d)):
            if _normalize_text(d) == target_n:
                return os.path.join(parent, d)
    # Try matching by first part (before underscore)
    target_first = target_n.split('_')[0].strip()
    for d in os.listdir(parent):
        if os.path.isdir(os.path.join(parent, d)):
            d_first = _normalize_text(d).split('_')[0].strip()
            if d_first == target_first:
                return os.path.join(parent, d)
    return pasta


def _extract_client_data_from_folder(pasta):
    """Extract client name, CPF, address from ALL documents in folder via OCR (cached)."""
    import re as _re
    data = {}
    data['nome'] = os.path.basename(pasta).split("_")[0].strip()
    files = sorted(os.listdir(pasta))

    def _find_cep(text):
        """Extract CEP from text - handles all OCR formats: 87308475, 87.308-475, 80.215-900"""
        patterns = [
            r'(?:CEP|Cep|cep)[:\s]*(\d{2})\.?(\d{3})[-.]?(\d{3})',
            r'\b(\d{2})\.(\d{3})[-.](\d{3})\b',
            r'(?:CEP|Cep|cep)[:\s]*(\d{5})[-.]?(\d{3})',
            r'\b(\d{5})[-](\d{3})\b',
        ]
        for p in patterns:
            m = _re.search(p, text)
            if m:
                digits = _re.sub(r'[^\d]', '', m.group(0))
                # Extract only last 8 digits (CEP)
                cep_m = _re.search(r'(\d{8})', digits)
                if cep_m:
                    cep = cep_m.group(1)
                    if cep != '00000000':
                        return f"{cep[:5]}-{cep[5:]}"
        return None

    # === Search all relevant docs for CPF, CEP, address ===
    search_order = []
    for f in files:
        fl = f.lower()
        if '8- cad' in fl or 'cadunico' in fl: search_order.insert(0, f)  # CadUnico first (best structured)
        elif '2- procura' in fl or 'procura' in fl: search_order.append(f)
        elif '7- comprov' in fl or 'residen' in fl: search_order.append(f)
        elif '5- ' in fl or '6- ' in fl or 'rg' in fl or 'identif' in fl: search_order.append(f)

    for f in search_order:
        path = os.path.join(pasta, f)
        if not os.path.isfile(path) or not path.lower().endswith('.pdf'):
            continue
        text = mistral_ocr(path)
        if not text:
            continue

        # CPF — prefer CPF near client name, skip CadUnico (has multiple CPFs)
        if not data.get('documento') and 'cadunico' not in f.lower() and 'cad' not in f.lower():
            cpf_m = _re.search(r'(?:CPF|cpf)[:\s]*[n°]*\s*(\d{3}[.\s]?\d{3}[.\s]?\d{3}[-.\s]?\d{2})', text)
            if not cpf_m:
                cpf_m = _re.search(r'\b(\d{3}\.\d{3}\.\d{3}-\d{2})\b', text)
            if cpf_m:
                cpf_clean = _re.sub(r'[^\d]', '', cpf_m.group(1))
                if len(cpf_clean) == 11:
                    data['documento'] = f"{cpf_clean[:3]}.{cpf_clean[3:6]}.{cpf_clean[6:9]}-{cpf_clean[9:]}"
                    print(f"  [EXTRACT] CPF em {f}: {data['documento']}")

        # CEP
        if not data.get('endereco_cep'):
            cep = _find_cep(text)
            if cep:
                data['endereco_cep'] = cep
                print(f"  [EXTRACT] CEP em {f}: {cep}")

        # Cidade/UF
        if not data.get('endereco_cidade'):
            for uf in ['SP','RJ','MG','PR','SC','RS','GO','MT','MS','BA','PE','CE','PA','AM','MA','PI','RN','PB','SE','AL','TO','RO','AC','AP','RR','ES','DF']:
                m = _re.search(rf'([A-ZÀ-Úa-zà-ú]{{3,}}(?:\s+[A-Za-zÀ-ú]+)*)\s*[-/]\s*{uf}\b', text)
                if m:
                    cidade = m.group(1).strip()
                    # Filter out false positives
                    bad = ('CEP', 'RUA', 'AV', 'LOCAL', 'CNPJ', 'OAB', 'INSCRIT', 'FONE', 'FAX', 'EMAIL', 'HTTP', 'WWW')
                    if cidade.upper() not in bad and not any(b in cidade.upper() for b in bad) and len(cidade) < 30:
                        data['endereco_cidade'] = cidade
                        data['endereco_uf'] = uf
                        print(f"  [EXTRACT] Cidade em {f}: {cidade}/{uf}")
                        break

        # Logradouro
        if not data.get('endereco_logradouro'):
            m = _re.search(r'(?:Endere[çc]o|Rua|Av\.|Avenida|Travessa)[:\s]*([^\n,]{5,60})', text, _re.IGNORECASE)
            if m:
                logr = m.group(1).strip()
                num_m = _re.search(r'n[°º]?\s*(\d+)', logr, _re.IGNORECASE)
                if num_m:
                    data['endereco_numero'] = num_m.group(1)
                data['endereco_logradouro'] = _re.sub(r'\s*n[°º]?\s*\d+', '', logr).strip()

        # Stop early if we have everything
        if all(data.get(k) for k in ['documento', 'endereco_cep', 'endereco_cidade']):
            break

    # Defaults for missing fields
    if data.get('documento'):
        if not data.get('endereco_cep'):
            _cep_uf = {'SP':'01000-000','RJ':'20000-000','MG':'30000-000','PR':'80000-000',
                       'SC':'88000-000','RS':'90000-000','BA':'40000-000','PE':'50000-000',
                       'CE':'60000-000','GO':'74000-000','DF':'70000-000','AM':'69000-000',
                       'MA':'65000-000','ES':'29000-000','MT':'78000-000','MS':'79000-000'}
            data['endereco_cep'] = _cep_uf.get(data.get('endereco_uf', ''), '01000-000')
        data.setdefault('endereco_logradouro', 'A informar')
        data.setdefault('endereco_numero', 'S/N')
        data.setdefault('endereco_bairro', 'A informar')
        data.setdefault('endereco_cidade', 'A informar')
        data.setdefault('endereco_uf', 'SP')
        return data
    return None


def legalmail_criar_rascunho(pasta, tribunal, sistema, comarca, assunto=None,
                             instancia="1", uf_tribunal="", client_data=None):
    """Create a draft petition on LegalMail with ALL fields filled and documents uploaded.

    Args:
        pasta: client folder path (organized with numbered files)
        tribunal: tribunal code (e.g. "TRF-4", "TRT-9")
        sistema: system code (e.g. "eproc_jfpr", "pje", "pje_trt")
        comarca: comarca name (e.g. "Curitiba", "Guarulhos")
        assunto: subject string (optional, auto-detected from API)
        instancia: "1" or "2" (default "1")
        uf_tribunal: state code when needed (e.g. "SP" for TRF-3/pje)
        client_data: optional dict with polo ativo party data:
            {nome, documento (CPF), endereco_cep, endereco_logradouro,
             endereco_numero, endereco_bairro, endereco_cidade, endereco_uf}

    Returns:
        dict with idpeticoes, idprocessos, status, filled fields, etc.
    """
    import requests, time

    if not LEGALMAIL_API_KEY:
        return {"error": "LEGALMAIL_API_KEY não configurada"}
    if not LEGALMAIL_CERT_ID:
        return {"error": "LEGALMAIL_CERTIFICADO_ID não configurado ou inválido"}

    # Step 1: Create petition initial (draft)
    print(f"  [LEGALMAIL] Criando rascunho: {tribunal}/{sistema} - {comarca}...")
    payload = {
        "tribunal": tribunal,
        "instancia": instancia,
        "sistema": sistema,
        "certificado_id": LEGALMAIL_CERT_ID,
    }
    if uf_tribunal:
        payload["ufTribunal"] = uf_tribunal

    resp = legalmail_request("post", "/petition/initial", json=payload)
    if resp.status_code == 429:
        print("  [LEGALMAIL] Rate limit, aguardando 60s...")
        time.sleep(60)
        resp = legalmail_request("post", "/petition/initial", json=payload)
    if resp.status_code != 200:
        return {"status": "erro", "error": f"HTTP {resp.status_code}: {resp.text[:200]}"}
    try:
        data = resp.json()
    except Exception:
        return {"status": "erro", "error": f"Resposta inválida: {resp.text[:200]}"}
    if isinstance(data, list):
        data = data[0]
    if data.get("status") != "sucesso":
        return {"status": "erro", "error": f"Erro ao criar petição: {data}"}

    dados = data.get("dados", {})
    idpeticoes = dados.get("idpeticoes")
    idprocessos = dados.get("idprocessos")
    if not idpeticoes:
        return {"status": "erro", "error": f"ID da petição não retornado: {data}"}
    print(f"  [LEGALMAIL] Rascunho criado: idpeticoes={idpeticoes}")

    # Step 1.5: Auto-detect comarca from client address if comarca is just a UF code
    if comarca and len(comarca) <= 2:
        detected_city = detect_cidade_from_folder(pasta)
        if detected_city:
            print(f"  [LEGALMAIL] Cidade detectada do endereco: {detected_city}")
            comarca = detected_city
        else:
            print(f"  [WARN] Comarca e apenas UF ({comarca}), cidade nao detectada")

    # Step 2: Extract valor da causa from xlsx
    valor_causa = extract_valor_causa(pasta)
    if valor_causa:
        print(f"  [LEGALMAIL] Valor da causa extraído: R$ {valor_causa:,.2f}")
    else:
        print(f"  [WARN] Valor da causa não encontrado no xlsx")

    # Step 3: Auto-extract client data if not provided
    if not client_data or not client_data.get('documento'):
        print(f"  [LEGALMAIL] Extraindo dados do cliente dos documentos...")
        client_data = _extract_client_data_from_folder(pasta)
        if client_data:
            print(f"  [LEGALMAIL] Cliente: {client_data.get('nome', '?')} CPF: {client_data.get('documento', '?')}")

    # Create/find parties
    id_polo_ativo = None
    id_polo_passivo = None

    # INSS (polo passivo)
    id_polo_passivo = legalmail_get_or_create_inss()

    # Polo ativo (client)
    if client_data and client_data.get('documento'):
        party_data = {**client_data, 'polo': 'ativo'}
        if 'etnia' not in party_data:
            party_data['etnia'] = 'Não declarada'
        id_polo_ativo = legalmail_create_party(party_data)
        if id_polo_ativo:
            print(f"  [LEGALMAIL] Polo ativo criado: id={id_polo_ativo}")
    else:
        print(f"  [WARN] Dados do cliente nao encontrados - polo ativo nao preenchido")

    # Step 4: Fill all petition fields (comarca -> rito -> competencia -> classe -> assunto -> flags -> parties)
    fill_result = legalmail_fill_fields(
        idpeticoes, sistema, comarca,
        valor_causa=valor_causa,
        id_polo_ativo=id_polo_ativo,
        id_polo_passivo=id_polo_passivo,
        uf_tribunal=uf_tribunal,
        tribunal=tribunal,
        instancia=instancia,
    )
    print(f"  [LEGALMAIL] Campos preenchidos: {list(fill_result['filled'].keys())}")
    if fill_result['errors']:
        for err in fill_result['errors']:
            print(f"  [WARN] {err}")

    # Step 5: Collect and process files from organized folder (sorted by numeric prefix)
    all_files = [f for f in os.listdir(pasta)
                 if os.path.isfile(os.path.join(pasta, f)) and not f.startswith('_')]

    def sort_key(fname):
        m = re.match(r'^(\d+)-', fname)
        return (int(m.group(1)), fname) if m else (9999, fname)

    files = sorted(all_files, key=sort_key)

    peticao_pdf = None
    attachments = []

    for f in files:
        filepath = os.path.join(pasta, f)
        match = re.match(r'^(\d+)-\s+', f)
        if not match:
            continue
        prefix = match.group(1)

        if prefix == "1":
            if f.endswith('.docx'):
                peticao_pdf = docx_to_pdf(filepath)
            elif f.endswith('.pdf'):
                peticao_pdf = filepath
        else:
            # Resolve doc type dynamically from API (per tribunal)
            doc_type_id = legalmail_resolve_doc_type(idpeticoes, prefix)
            if doc_type_id:
                if f.endswith('.docx'):
                    pdf_path = docx_to_pdf(filepath)
                    if pdf_path:
                        attachments.append((pdf_path, doc_type_id, f))
                elif f.endswith('.pdf'):
                    attachments.append((filepath, doc_type_id, f))
                elif f.endswith('.xlsx'):
                    pdf_version = os.path.splitext(filepath)[0] + '.pdf'
                    if os.path.exists(pdf_version):
                        continue
                    pdf_path = xlsx_to_pdf(filepath)
                    if pdf_path:
                        attachments.append((pdf_path, doc_type_id, f))
            else:
                print(f"  [WARN] Tipo de anexo não encontrado para prefixo {prefix}: {f}")

    # Step 6: Upload main petition PDF
    if peticao_pdf and os.path.exists(peticao_pdf):
        print(f"  [LEGALMAIL] Enviando petição principal...")
        with open(peticao_pdf, 'rb') as pf:
            resp = legalmail_request("post", f"/petition/file?idpeticoes={idpeticoes}&idprocessos={idprocessos}",
                files={"file": (os.path.basename(peticao_pdf), pf, "application/pdf")})
        if resp.status_code == 429:
            print(f"  [LEGALMAIL] Rate limit na petição, aguardando 60s...")
            time.sleep(60)
            with open(peticao_pdf, 'rb') as pf:
                resp = legalmail_request("post", f"/petition/file?idpeticoes={idpeticoes}&idprocessos={idprocessos}",
                    files={"file": (os.path.basename(peticao_pdf), pf, "application/pdf")})
        print(f"  [LEGALMAIL] Petição enviada: {resp.status_code}")
    else:
        print(f"  [WARN] Petição PDF não encontrada/gerada")

    # Step 7: Upload attachments
    uploaded = 0
    for filepath, doc_type_id, original_name in attachments:
        try:
            time.sleep(1)
            with open(filepath, 'rb') as af:
                resp = legalmail_request("post",
                    f"/petition/attachments?idpeticoes={idpeticoes}&fk_documentos_tipos={doc_type_id}",
                    files={"file": (os.path.basename(filepath), af, "application/pdf")})
            if resp.status_code == 200:
                uploaded += 1
                print(f"  [LEGALMAIL] Anexo enviado: {original_name}")
            elif resp.status_code == 429:
                print(f"  [LEGALMAIL] Rate limit, aguardando 60s...")
                time.sleep(60)
                with open(filepath, 'rb') as af:
                    resp = legalmail_request("post",
                        f"/petition/attachments?idpeticoes={idpeticoes}&fk_documentos_tipos={doc_type_id}",
                        files={"file": (os.path.basename(filepath), af, "application/pdf")})
                if resp.status_code == 200:
                    uploaded += 1
                    print(f"  [LEGALMAIL] Anexo enviado (retry): {original_name}")
                else:
                    print(f"  [WARN] Erro ao enviar {original_name} (retry): {resp.status_code} {resp.text[:100]}")
            else:
                print(f"  [WARN] Erro ao enviar {original_name}: {resp.status_code} {resp.text[:100]}")
        except Exception as e:
            print(f"  [WARN] Erro ao enviar {original_name}: {e}")

    print(f"  [LEGALMAIL] Rascunho completo: {uploaded} anexos, campos={list(fill_result['filled'].keys())}")

    return {
        "status": "ok",
        "idpeticoes": idpeticoes,
        "idprocessos": idprocessos,
        "uploaded_attachments": uploaded,
        "filled_fields": list(fill_result['filled'].keys()),
        "fill_errors": fill_result['errors'],
        "valor_causa": valor_causa,
        "polo_ativo_id": id_polo_ativo,
        "polo_passivo_id": id_polo_passivo,
        "url": f"https://app.legalmail.com.br/petitions/{idpeticoes}"
    }


@app.route("/api/debug/key", methods=["GET"])
@require_admin
def debug_key():
    key = LEGALMAIL_API_KEY
    return jsonify({"key_configured": bool(key), "key_length": len(key)})


@app.route("/api/legalmail/tribunais", methods=["GET"])
def legalmail_tribunais():
    """List all available tribunals and their systems."""
    resp = legalmail_request("get", "/petition/tribunals")
    return jsonify(resp.json())


@app.route("/api/legalmail/comarcas", methods=["GET"])
def legalmail_comarcas():
    """List available comarcas for a given tribunal/sistema."""
    tribunal = request.args.get("tribunal", "TRF-4")
    sistema = request.args.get("sistema", "eproc_jfpr")
    uf_tribunal = request.args.get("uf", "")

    # Create temp petition to get comarcas
    payload = {"tribunal": tribunal, "instancia": "1", "sistema": sistema,
               "certificado_id": LEGALMAIL_CERT_ID, "gratuidade": True}
    if uf_tribunal:
        payload["ufTribunal"] = uf_tribunal
    resp = legalmail_request("post", "/petition/initial", json=payload)
    data = resp.json()
    if isinstance(data, list):
        data = data[0]
    if data.get("status") != "sucesso":
        return jsonify({"error": str(data)}), 400
    idpeticoes = data["dados"]["idpeticoes"]

    # Get comarcas, subjects, classes
    resp_comarcas = legalmail_request("get", f"/petition/county?idpeticoes={idpeticoes}")
    resp_assuntos = legalmail_request("get", f"/petition/subjects?idpeticoes={idpeticoes}")

    result = {
        "comarcas": resp_comarcas.json(),
        "assuntos": resp_assuntos.json(),
    }

    # Delete temp petition
    legalmail_request("delete", f"/petition/initial?idpeticoes={idpeticoes}")

    return jsonify(result)


@app.route("/api/legalmail/rascunho", methods=["POST"])
def legalmail_rascunho():
    """Create a draft petition on LegalMail for a client folder."""
    data = request.get_json()
    pasta = data.get("pasta", "").strip()
    tribunal = data.get("tribunal", "TRF-4").strip()
    sistema = data.get("sistema", "eproc_jfpr").strip()
    comarca = data.get("comarca", "").strip()
    assunto = data.get("assunto", "").strip() or None
    uf_tribunal = data.get("uf", "").strip()
    client_data = data.get("client_data")  # optional: {nome, documento, endereco_*}

    if not pasta or not os.path.isdir(pasta):
        return jsonify({"error": "Pasta não encontrada"}), 400

    result = legalmail_criar_rascunho(pasta, tribunal, sistema, comarca,
                                       assunto=assunto, uf_tribunal=uf_tribunal,
                                       client_data=client_data)
    if "error" in result:
        return jsonify(result), 400
    return jsonify(result)


@app.route("/api/legalmail/rascunho-lote", methods=["POST"])
def legalmail_rascunho_lote():
    """Create draft petitions for multiple client folders."""
    from flask import Response, stream_with_context
    data = request.get_json()
    pastas = data.get("pastas", [])
    tribunal = data.get("tribunal", "TRF-4")
    sistema = data.get("sistema", "eproc_jfpr")
    comarca = data.get("comarca", "")
    assunto = data.get("assunto", "") or None

    uf_tribunal = data.get("uf", "")

    def generate():
        results = []
        for i, pasta in enumerate(pastas):
            pasta = pasta.strip()
            cliente = os.path.basename(pasta).split("_")[0].strip()
            yield f"data: {json.dumps({'type': 'progress', 'index': i, 'total': len(pastas), 'cliente': cliente, 'step': 'enviando'})}\n\n"

            if not os.path.isdir(pasta):
                yield f"data: {json.dumps({'type': 'error', 'index': i, 'cliente': cliente, 'message': 'Pasta nao encontrada'})}\n\n"
                continue

            try:
                result = legalmail_criar_rascunho(pasta, tribunal, sistema, comarca,
                                                   assunto=assunto, uf_tribunal=uf_tribunal)
                if "error" in result:
                    yield f"data: {json.dumps({'type': 'error', 'index': i, 'cliente': cliente, 'message': result['error']})}\n\n"
                else:
                    result["cliente"] = cliente
                    results.append(result)
                    yield f"data: {json.dumps({'type': 'done', 'index': i, 'total': len(pastas), 'cliente': cliente, 'idpeticoes': result['idpeticoes'], 'anexos': result['uploaded_attachments']})}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'index': i, 'cliente': cliente, 'message': str(e)})}\n\n"

        yield f"data: {json.dumps({'type': 'complete', 'results': results})}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


###############################################################################
# MONITORAMENTO DE PROCESSOS E ANÁLISE DE INTIMAÇÕES
###############################################################################

import time
import threading
import datetime as _dt

# Persistent storage - PostgreSQL on Railway, JSON files locally
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_DB = bool(DATABASE_URL)

# Local file fallback
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
NOTIFICATIONS_FILE = os.path.join(DATA_DIR, "notifications.json")
PROCESSES_CACHE_FILE = os.path.join(DATA_DIR, "processes_cache.json")
MONITOR_STATE_FILE = os.path.join(DATA_DIR, "monitor_state.json")

# Initialize PostgreSQL if available
if USE_DB:
    import psycopg2
    import psycopg2.extras
    print(f"[DB] DATABASE_URL encontrada, tentando conectar...")
    print(f"[DB] URL prefix: {DATABASE_URL[:40]}...")

    def _get_db():
        """Connect to PostgreSQL with SSL fallback."""
        try:
            return psycopg2.connect(DATABASE_URL, connect_timeout=10)
        except Exception:
            # Try with sslmode=require (Railway may need it)
            return psycopg2.connect(DATABASE_URL, connect_timeout=10, sslmode='require')

    # Create tables on startup - retry up to 3 times
    _db_connected = False
    for _attempt in range(3):
        try:
            conn = _get_db()
            cur = conn.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS kv_store (
                    key TEXT PRIMARY KEY,
                    value JSONB NOT NULL DEFAULT '[]'::jsonb,
                    updated_at TIMESTAMP DEFAULT NOW()
                )
            """)
            conn.commit()
            cur.close()
            conn.close()
            print(f"[DB] PostgreSQL conectado e tabelas criadas (tentativa {_attempt+1})")
            _db_connected = True
            break
        except Exception as e:
            print(f"[DB] Tentativa {_attempt+1}/3 falhou: {type(e).__name__}: {e}")
            import time as _time_mod
            _time_mod.sleep(2)

    if not _db_connected:
        print(f"[DB] ERRO CRITICO: Nao conseguiu conectar ao PostgreSQL apos 3 tentativas!")
        print(f"[DB] FALLBACK para JSON local - dados NAO persistem entre deploys!")
        USE_DB = False
    else:
        print(f"[DB] USE_DB = True - dados persistem no PostgreSQL")

# Monitor settings
MONITOR_INTERVAL_MINUTES = int(os.environ.get("MONITOR_INTERVAL_MINUTES", "360"))  # Check every 6h (survives Railway restarts)
_monitor_thread = None
_monitor_running = False


def _db_load(key, default=None):
    """Load JSON data from PostgreSQL."""
    if default is None:
        default = []
    try:
        conn = _get_db()
        cur = conn.cursor()
        cur.execute("SELECT value FROM kv_store WHERE key = %s", (key,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return row[0] if row else default
    except Exception as e:
        print(f"[DB] Erro ao ler {key}: {e}")
        return default


def _db_save(key, data):
    """Save JSON data to PostgreSQL."""
    try:
        conn = _get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO kv_store (key, value, updated_at) VALUES (%s, %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = %s, updated_at = NOW()
        """, (key, json.dumps(data, ensure_ascii=False), json.dumps(data, ensure_ascii=False)))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[DB] Erro ao salvar {key}: {e}")


def _load_json_file(filepath, default=None):
    if default is None:
        default = []
    # Use DB if available, with filepath as key
    if USE_DB:
        key = os.path.basename(filepath).replace(".json", "")
        return _db_load(key, default)
    if os.path.exists(filepath):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return default


def _generate_petition_pdf(texto, pdf_path, numero_processo="", tipo_peticao=""):
    """Generate a well-formatted petition PDF from text."""
    doc = fitz.open()
    lines = texto.split('\n')
    page = None
    y = 0
    margin_left = 72
    margin_right = 72
    margin_top = 80
    margin_bottom = 72
    line_height_body = 16
    line_height_title = 20
    page_width = 595  # A4
    page_height = 842
    usable_width = page_width - margin_left - margin_right

    def new_page():
        nonlocal page, y
        page = doc.new_page(width=page_width, height=page_height)
        y = margin_top
        # Header line
        page.draw_line(fitz.Point(margin_left, 56), fitz.Point(page_width - margin_right, 56),
                       color=(0.122, 0.216, 0.388), width=2)
        # Footer
        footer_y = page_height - 36
        page.insert_text(fitz.Point(margin_left, footer_y),
                         f"Dr. José Roberto da Costa Junior — OAB/SP 378.163",
                         fontsize=7, fontname="helv", color=(0.5, 0.5, 0.5))
        page.insert_text(fitz.Point(page_width - margin_right - 30, footer_y),
                         f"Pág. {len(doc)}",
                         fontsize=7, fontname="helv", color=(0.5, 0.5, 0.5))
        return page

    new_page()

    for line in lines:
        stripped = line.strip()
        if not stripped:
            y += 10
            if y > page_height - margin_bottom:
                new_page()
            continue

        # Detect section titles (all caps lines or lines starting with roman numerals)
        is_title = False
        is_section = False
        fontsize = 11
        fontname = "helv"
        color = (0, 0, 0)
        indent = 0

        # Main title detection (RECURSO ORDINÁRIO, APELAÇÃO, EMBARGOS, etc.)
        upper_stripped = stripped.upper()
        title_keywords = ['RECURSO', 'APELAÇÃO', 'AGRAVO', 'EMBARGOS', 'MANIFESTAÇÃO',
                          'CONTRARRAZÕES', 'IMPUGNAÇÃO', 'PETIÇÃO', 'EXMO', 'EXMA']
        if stripped == stripped.upper() and len(stripped) > 5 and len(stripped) < 120:
            is_title = True
            fontsize = 12
            fontname = "hebo"  # Helvetica Bold
            color = (0.122, 0.216, 0.388)  # #1F3763

        # Section headers (I – , II – , III – , etc.)
        elif re.match(r'^[IVX]+[\s\.\-–—]+', stripped) or re.match(r'^[IVX]+\s*$', stripped):
            is_section = True
            fontsize = 11
            fontname = "hebo"
            color = (0.122, 0.216, 0.388)

        # Sub-section (1., 2., a), b), etc.)
        elif re.match(r'^(\d+[\.\)]\s|[a-z]\)\s)', stripped):
            indent = 20

        # Check if need new page
        needed = line_height_title if (is_title or is_section) else line_height_body
        if y + needed > page_height - margin_bottom:
            new_page()

        # Add spacing before titles/sections
        if is_title:
            y += 8
        elif is_section:
            y += 6

        # Wrap long lines
        max_chars = int((usable_width - indent) / (fontsize * 0.45))
        if max_chars < 20:
            max_chars = 60

        wrapped = []
        words = stripped.split()
        current_line = ""
        for word in words:
            test = current_line + (" " if current_line else "") + word
            if len(test) > max_chars:
                if current_line:
                    wrapped.append(current_line)
                current_line = word
            else:
                current_line = test
        if current_line:
            wrapped.append(current_line)

        if not wrapped:
            wrapped = [stripped]

        for wline in wrapped:
            if y + line_height_body > page_height - margin_bottom:
                new_page()

            x = margin_left + indent
            if is_title:
                # Center titles
                text_width = fitz.get_text_length(wline, fontname=fontname, fontsize=fontsize)
                x = (page_width - text_width) / 2

            page.insert_text(fitz.Point(x, y), wline,
                             fontsize=fontsize, fontname=fontname, color=color)
            y += line_height_title if (is_title or is_section) else line_height_body

        # Spacing after title
        if is_title:
            y += 6

    doc.save(pdf_path)
    doc.close()


def _save_json_file(filepath, data):
    if USE_DB:
        key = os.path.basename(filepath).replace(".json", "")
        _db_save(key, data)
        return
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _load_monitor_state():
    return _load_json_file(MONITOR_STATE_FILE, {
        "last_check": None,
        "known_movements": {},  # idprocesso -> set of idmovimentacoes
        "auto_analyze": True,
    })


def _save_monitor_state(state):
    _save_json_file(MONITOR_STATE_FILE, state)


def monitor_fetch_all_processes():
    """Fetch all processes from workspace. Returns list."""
    import requests
    all_procs = []
    offset = 0
    while True:
        try:
            r = legalmail_request("get", f"/process/all?offset={offset}&limit=50")
            if r.status_code == 429:
                time.sleep(60)
                r = legalmail_request("get", f"/process/all?offset={offset}&limit=50")
            if r.status_code != 200:
                break
            data = r.json()
            if not isinstance(data, list) or len(data) == 0:
                break
            all_procs.extend(data)
            if len(data) < 50:
                break
            offset += 50
            time.sleep(2)  # Be gentle with rate limits
        except Exception as e:
            print(f"  [MONITOR] Erro ao buscar processos offset={offset}: {e}")
            break
    return all_procs


def monitor_fetch_autos(idprocesso):
    """Fetch autos/movements for a process. Returns list of movements."""
    try:
        r = legalmail_request("get", f"/process/autos?idprocesso={idprocesso}")
        if r.status_code == 429:
            time.sleep(60)
            r = legalmail_request("get", f"/process/autos?idprocesso={idprocesso}")
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list):
                return data
    except Exception as e:
        print(f"  [MONITOR] Erro ao buscar autos processo {idprocesso}: {e}")
    return []


def monitor_fetch_movement_text(idmovimentacoes):
    """Download movement document and extract text."""
    try:
        r = legalmail_request("get", f"/process/movement/url?idmovimentacoes={idmovimentacoes}")
        if r.status_code == 429:
            time.sleep(60)
            r = legalmail_request("get", f"/process/movement/url?idmovimentacoes={idmovimentacoes}")
        if r.status_code != 200:
            return None
        data = r.json()
        s3_url = data.get("s3_url", "")
        if not s3_url:
            return None

        # Download PDF and extract text
        import requests as req_lib
        pdf_resp = req_lib.get(s3_url, timeout=30)
        if pdf_resp.status_code != 200:
            return None

        pdf_doc = fitz.open(stream=pdf_resp.content, filetype="pdf")
        text = ""
        for page in pdf_doc:
            text += page.get_text()
        pdf_doc.close()
        return text.strip() if text.strip() else None
    except Exception as e:
        print(f"  [MONITOR] Erro ao baixar movimentação {idmovimentacoes}: {e}")
    return None


def monitor_analyze_movement(numero_processo, tribunal, titulo, texto):
    """Use Claude to analyze a movement/intimation."""
    try:
        client_ai = anthropic.Anthropic(timeout=120.0)
        prompt = f"""Você é um assistente jurídico. Analise esta movimentação processual:

Processo: {numero_processo} | Tribunal: {tribunal}
Título: {titulo}

TEXTO:
{texto[:6000]}

Responda APENAS com JSON válido:
{{"resumo": "resumo claro e objetivo (2-3 frases)", "tipo_movimentacao": "intimação|despacho|sentença|decisão|citação|audiência|perícia|expediente|outro", "prazo_dias": 0, "data_prazo": null, "urgencia": "alta|media|baixa", "acao_necessaria": "ação específica do advogado", "tipo_peticao_sugerida": "contestação|recurso|apelação|agravo|manifestação|cumprimento|embargos|impugnação|outro|nenhuma", "resultado_merito": null, "observacoes": "pontos relevantes"}}"""

        response = client_ai.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}]
        )
        response_text = response.content[0].text.strip()
        if "```json" in response_text:
            response_text = response_text.split("```json")[1].split("```")[0].strip()
        elif "```" in response_text:
            response_text = response_text.split("```")[1].split("```")[0].strip()
        return json.loads(response_text)
    except Exception as e:
        print(f"  [MONITOR] Erro na análise IA: {e}")
    return None


def monitor_check_updates():
    """Main monitoring function: check all processes for new movements.

    How it works:
    - Fetches the full process list from LegalMail (/process/all)
    - Compares each process's `last_import` date with our stored date
    - Only fetches autos for processes that have been updated since last check
    - On FIRST RUN: stores all current movement IDs as "known" (no flood of old notifications)
    - On subsequent runs: only creates notifications for truly NEW movements
    - Auto-analyzes new movements with Claude AI
    """
    print(f"\n[MONITOR] Verificando atualizacoes... {_dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    state = _load_monitor_state()
    notifications = _load_notifications()
    known = state.get("known_movements", {})
    last_imports = state.get("last_imports", {})
    is_first_run = state.get("first_run_done") is not True
    new_count = 0

    # Step 1: Fetch all processes
    processes = monitor_fetch_all_processes()
    if not processes:
        print("  [MONITOR] Nenhum processo encontrado")
        return 0

    # Save processes cache
    _save_json_file(PROCESSES_CACHE_FILE, processes)

    # Only active processes (with inbox_atual = being monitored by LegalMail)
    active_processes = [p for p in processes if p.get("inbox_atual")]
    print(f"  [MONITOR] {len(processes)} processos total, {len(active_processes)} ativos")

    if is_first_run:
        print(f"  [MONITOR] PRIMEIRA EXECUCAO - registrando processos e notificando movimentacoes recentes")

    # Step 2: Check which processes have new updates
    checked = 0
    # Only check processes updated in the last 2 days (avoids wasting rate limit on old ones)
    cutoff_date = (_dt.datetime.now() - _dt.timedelta(days=2)).strftime("%Y-%m-%d")

    for proc in active_processes:
        idprocesso = str(proc.get("idprocessos", ""))
        numero = proc.get("numero_processo", "?")
        tribunal = proc.get("tribunal", "?")
        proc_last_import = proc.get("last_import") or ""

        if not idprocesso:
            continue

        # Skip if last_import hasn't changed since our last check
        stored_import = last_imports.get(idprocesso, "")
        if stored_import and proc_last_import == stored_import:
            continue  # No new data for this process

        # If we don't have stored data, only check recently updated processes
        if not stored_import:
            import_date = proc_last_import[:10] if proc_last_import else ""
            if import_date < cutoff_date:
                # Old process — just store last_import and skip
                last_imports[idprocesso] = proc_last_import
                continue

        # Update stored last_import
        last_imports[idprocesso] = proc_last_import

        time.sleep(2)  # Rate limit safety
        autos = monitor_fetch_autos(idprocesso)
        if not autos:
            continue

        # Get current movement IDs
        current_ids = [str(m.get("idmovimentacoes", "")) for m in autos]
        known_ids = set(known.get(idprocesso, []))

        # Find movements we haven't seen before
        new_movements = [m for m in autos if str(m.get("idmovimentacoes", "")) not in known_ids]

        # Filter: only movements from last 2 days (avoid flooding with old ones)
        mov_cutoff = (_dt.datetime.now() - _dt.timedelta(days=2)).strftime("%Y-%m-%d")
        new_movements = [m for m in new_movements
                         if (m.get("data_movimentacao", "") or "")[:10] >= mov_cutoff]

        # Sort by date (most recent last)
        new_movements.sort(key=lambda m: m.get("data_movimentacao", "") or "")

        if new_movements:
            print(f"  [MONITOR] {len(new_movements)} nova(s) movimentacao(oes) em {numero}")

            for mov in new_movements:
                mov_id = str(mov.get("idmovimentacoes", ""))
                titulo = mov.get("titulo", "Movimentacao")
                data_mov = mov.get("data_movimentacao", "")

                # Create notification
                notif = {
                    "type": "intimacao",
                    "source": "monitor_auto",
                    "timestamp": _dt.datetime.now().isoformat(),
                    "numero_processo": numero,
                    "tribunal": tribunal,
                    "sistema": proc.get("sistema_tribunal", ""),
                    "polo_ativo": proc.get("poloativo_nome", ""),
                    "polo_passivo": proc.get("polopassivo_nome", ""),
                    "classe": proc.get("nome_classe", ""),
                    "idprocesso": idprocesso,
                    "idmovimentacoes": mov_id,
                    "titulo_movimentacao": titulo,
                    "data_movimentacao": data_mov,
                    "documentos": [{
                        "tipo": "movement",
                        "title": titulo,
                        "movement_date": data_mov,
                        "idmovimentacoes": mov_id,
                    }],
                    "analyzed": False,
                    "analysis": None,
                }

                # Fetch movement text (for analysis later)
                texto = monitor_fetch_movement_text(mov_id)
                if texto:
                    notif["texto_movimentacao"] = texto[:8000]
                print(f"    -> {titulo[:80]}")

                notifications.append(notif)
                new_count += 1

        # Always update known movements (even on first run)
        known[idprocesso] = current_ids
        checked += 1

        # Don't check too many in one run to stay within rate limits
        if checked >= 50:
            print(f"  [MONITOR] Limite de verificacoes atingido, continuando no proximo ciclo")
            break

    # Save state
    state["known_movements"] = known
    state["last_imports"] = last_imports
    state["last_check"] = _dt.datetime.now().isoformat()
    state["processes_total"] = len(processes)
    state["processes_active"] = len(active_processes)
    state["first_run_done"] = True
    _save_monitor_state(state)
    _save_notifications(notifications)

    if is_first_run:
        print(f"  [MONITOR] Primeira execucao concluida: {checked} processos registrados")
    else:
        print(f"  [MONITOR] Concluido: {new_count} nova(s) notificacao(oes), {checked} processos verificados")
    return new_count


def _monitor_loop():
    """Background loop that runs monitoring periodically."""
    global _monitor_running
    # Wait a bit for app to start
    time.sleep(10)
    print(f"\n[MONITOR] Monitor automático iniciado (intervalo: {MONITOR_INTERVAL_MINUTES} min)")

    while _monitor_running:
        try:
            monitor_check_updates()
        except Exception as e:
            import traceback
            print(f"  [MONITOR] Erro no ciclo: {e}")
            traceback.print_exc()

        # Sleep in small increments so we can stop cleanly
        for _ in range(MONITOR_INTERVAL_MINUTES * 60):
            if not _monitor_running:
                break
            time.sleep(1)

    print("[MONITOR] Monitor automático parado")


def start_monitor():
    """Start the background monitoring thread."""
    global _monitor_thread, _monitor_running
    if _monitor_running:
        return False
    _monitor_running = True
    _monitor_thread = threading.Thread(target=_monitor_loop, daemon=True)
    _monitor_thread.start()
    return True


def stop_monitor():
    """Stop the background monitoring thread."""
    global _monitor_running
    _monitor_running = False


@app.route("/api/legalmail/monitor/status", methods=["GET"])
@require_admin
def legalmail_monitor_status():
    """Get monitor status and stats."""
    try:
        state = _load_monitor_state()
        notifications = _load_notifications()
        if not isinstance(notifications, list):
            notifications = []
        pending = [n for n in notifications if isinstance(n, dict) and not n.get("analyzed")]
        urgent = [n for n in notifications if isinstance(n, dict) and (n.get("analysis") or {}).get("urgencia") == "alta"]
        return jsonify({
            "running": _monitor_running,
            "interval_minutes": MONITOR_INTERVAL_MINUTES,
            "last_check": state.get("last_check"),
            "processes_total": state.get("processes_total", 0),
            "processes_active": state.get("processes_active", 0),
            "notifications_total": len(notifications),
            "notifications_pending": len(pending),
            "notifications_urgent": len(urgent),
            "auto_analyze": state.get("auto_analyze", True),
        })
    except Exception as e:
        return jsonify({"error": f"Status error: {e}", "running": _monitor_running}), 500


@app.route("/api/legalmail/monitor/start", methods=["POST"])
@require_admin
def legalmail_monitor_start():
    """Start background monitoring."""
    started = start_monitor()
    return jsonify({"status": "ok" if started else "already_running", "running": _monitor_running})


@app.route("/api/legalmail/monitor/stop", methods=["POST"])
@require_admin
def legalmail_monitor_stop():
    """Stop background monitoring."""
    stop_monitor()
    return jsonify({"status": "ok", "running": False})


@app.route("/api/legalmail/monitor/check-now", methods=["POST"])
@require_admin
def legalmail_monitor_check_now():
    """Run a manual check immediately (in a thread to not block).

    Pass ?rescan=true to clear known movements and re-detect recent ones.
    This is useful after first run or when updates were missed.
    """
    rescan = request.args.get("rescan", "").lower() == "true"
    if rescan:
        state = _load_monitor_state()
        state["known_movements"] = {}
        state["last_imports"] = {}
        # Keep first_run_done=True so it generates notifications (not silent first run)
        state["first_run_done"] = True
        _save_monitor_state(state)
        print("[MONITOR] Re-scan: known_movements limpo, vai re-detectar movimentações recentes")

    def _run():
        try:
            monitor_check_updates()
        except Exception as e:
            print(f"  [MONITOR] Erro: {e}")
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    msg = "Re-scan iniciado (detectando movimentações de ontem/hoje)" if rescan else "Verificação iniciada em background"
    return jsonify({"status": "ok", "message": msg})


@app.route("/api/legalmail/monitor/config", methods=["POST"])
@require_admin
def legalmail_monitor_config():
    """Update monitor configuration."""
    global MONITOR_INTERVAL_MINUTES
    data = request.get_json()
    state = _load_monitor_state()
    if "interval_minutes" in data:
        MONITOR_INTERVAL_MINUTES = max(5, int(data["interval_minutes"]))
    if "auto_analyze" in data:
        state["auto_analyze"] = bool(data["auto_analyze"])
    _save_monitor_state(state)
    return jsonify({"status": "ok", "interval_minutes": MONITOR_INTERVAL_MINUTES, "auto_analyze": state["auto_analyze"]})


@app.route("/api/legalmail/processo/autos", methods=["GET"])
def legalmail_processo_autos():
    """Get movements/autos for a specific process."""
    idprocesso = request.args.get("idprocesso", "")
    if not idprocesso:
        return jsonify({"error": "idprocesso obrigatório"}), 400
    resp = legalmail_request("get", f"/process/autos?idprocesso={idprocesso}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


def _load_notifications():
    return _load_json_file(NOTIFICATIONS_FILE, [])


def _save_notifications(notifications):
    _save_json_file(NOTIFICATIONS_FILE, notifications)


@app.route("/api/legalmail/processos", methods=["GET"])
def legalmail_listar_processos():
    """List all processes in the workspace."""
    offset = request.args.get("offset", 0, type=int)
    limit = request.args.get("limit", 50, type=int)
    resp = legalmail_request("get", f"/process/all?offset={offset}&limit={limit}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/processo/detalhe", methods=["GET"])
def legalmail_detalhe_processo():
    """Get details of a specific process."""
    numero = request.args.get("numero_processo", "")
    idprocesso = request.args.get("idprocesso", "")
    params = ""
    if numero:
        params = f"numero_processo={numero}"
    elif idprocesso:
        params = f"idprocesso={idprocesso}"
    else:
        return jsonify({"error": "Informe numero_processo ou idprocesso"}), 400
    resp = legalmail_request("get", f"/process/detail?{params}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/processo/movimentacao-url", methods=["GET"])
def legalmail_movimentacao_url():
    """Get S3 URL for a movement document."""
    idmovimentacoes = request.args.get("idmovimentacoes", "")
    if not idmovimentacoes:
        return jsonify({"error": "idmovimentacoes obrigatório"}), 400
    resp = legalmail_request("get", f"/process/movement/url?idmovimentacoes={idmovimentacoes}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/importar", methods=["POST"])
def legalmail_importar_processo():
    """Import processes for monitoring.

    Body: { "pedidos": [{"numero": "...", "tribunal": "TRF-3", "sistema": "pje", "classe_id": "143"}] }
    """
    data = request.get_json()
    pedidos = data.get("pedidos", [])
    if not pedidos:
        return jsonify({"error": "Nenhum pedido informado"}), 400

    # Add certificado_id to each pedido if not present
    for p in pedidos:
        if "certificado_id" not in p:
            p["certificado_id"] = LEGALMAIL_CERT_ID

    resp = legalmail_request("post", "/imports/request", json={"pedidos": pedidos})
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/importar/status", methods=["GET"])
def legalmail_importar_status():
    """Check import status by hash or process number."""
    hash_val = request.args.get("hash", "")
    numero = request.args.get("numero", "")
    params = ""
    if hash_val:
        params = f"hash={hash_val}"
    elif numero:
        params = f"numero={numero}"
    else:
        return jsonify({"error": "Informe hash ou numero"}), 400
    resp = legalmail_request("get", f"/imports/status?{params}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/importar/classes", methods=["GET"])
def legalmail_importar_classes():
    """Search import classes by term."""
    term = request.args.get("term", "")
    resp = legalmail_request("get", f"/imports/classes?term={term}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/importar/sistemas", methods=["GET"])
def legalmail_importar_sistemas():
    """Get allowed systems for a process number."""
    numero = request.args.get("numero", "")
    if not numero:
        return jsonify({"error": "numero obrigatório"}), 400
    resp = legalmail_request("get", f"/imports/allowed-systems?numero={numero}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/webhook/configurar", methods=["POST"])
def legalmail_configurar_webhook():
    """Register webhook endpoint for notifications.

    Body: { "endpoint": "https://your-server.com/api/legalmail/webhook", "key_endpoint": "optional-secret" }
    """
    data = request.get_json()
    endpoint_url = data.get("endpoint", "")
    key_endpoint = data.get("key_endpoint", "")
    if not endpoint_url:
        return jsonify({"error": "endpoint obrigatório"}), 400

    params = f"endpoint={endpoint_url}"
    if key_endpoint:
        params += f"&key_endpoint={key_endpoint}"
    resp = legalmail_request("post", f"/workspace/notifications/endpoint?{params}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/webhook", methods=["POST"])
def legalmail_webhook_receiver():
    """Receive webhook notifications from LegalMail.

    LegalMail sends new intimações/movimentações here automatically.
    We store them and can trigger analysis.
    """
    # Validate webhook secret if configured
    if WEBHOOK_SECRET:
        incoming = request.args.get("secret", "") or request.headers.get("X-Webhook-Secret", "")
        if not hmac.compare_digest(incoming, WEBHOOK_SECRET):
            return jsonify({"error": "forbidden"}), 403

    payload = request.get_json(force=True, silent=True)
    if not payload:
        return jsonify({"status": "ignored", "reason": "empty payload"}), 200

    notifications = _load_notifications()

    params = payload.get("params", [])
    import datetime
    timestamp = datetime.datetime.now().isoformat()

    for item in params:
        # Check if it's a protocol event
        if item.get("event") == "protocolo_finalizado":
            notifications.append({
                "type": "protocolo_finalizado",
                "timestamp": timestamp,
                "data": item,
                "analyzed": False,
            })
            continue

        # It's an intimation/movement notification
        notification = {
            "type": "intimacao",
            "timestamp": timestamp,
            "numero_processo": item.get("numero_processo", ""),
            "tribunal": item.get("tribunal", ""),
            "sistema": item.get("sistema_tribunal", ""),
            "polo_ativo": item.get("polo_ativo", ""),
            "polo_passivo": item.get("polo_passivo", ""),
            "classe": item.get("classe_processo", ""),
            "inbox_id": item.get("inbox_id", ""),
            "inbox_uri": item.get("inbox_uri", ""),
            "documentos": item.get("documento", []),
            "analyzed": False,
            "analysis": None,
        }
        notifications.append(notification)

    _save_notifications(notifications)
    print(f"  [WEBHOOK] Recebidas {len(params)} notificações")
    return jsonify({"status": "ok", "received": len(params)}), 200


@app.route("/api/legalmail/notificacoes", methods=["GET"])
def legalmail_listar_notificacoes():
    """List received webhook notifications."""
    notifications = _load_notifications()
    only_pending = request.args.get("pending", "false").lower() == "true"
    if only_pending:
        notifications = [n for n in notifications if not n.get("analyzed")]
    # Filter: only last 5 days
    cutoff = (_dt.datetime.now() - _dt.timedelta(days=5)).strftime("%Y-%m-%d")
    notifications = [n for n in notifications
                     if (n.get("data_movimentacao") or n.get("timestamp") or "")[:10] >= cutoff]
    # Sort by date (most recent first)
    notifications.sort(
        key=lambda n: n.get("data_movimentacao") or n.get("timestamp") or "",
        reverse=True
    )
    return jsonify(notifications)


@app.route("/api/legalmail/notificacoes/importar", methods=["POST"])
@require_admin
def legalmail_importar_notificacoes():
    """Import notifications from JSON array (replaces existing)."""
    data = request.get_json()
    if not isinstance(data, list):
        return jsonify({"error": "Esperado um array JSON"}), 400
    _save_notifications(data)
    return jsonify({"status": "ok", "imported": len(data)})


@app.route("/api/legalmail/notificacoes/limpar", methods=["POST"])
@require_admin
def legalmail_limpar_notificacoes():
    """Remove notifications older than 5 days from storage."""
    notifications = _load_notifications()
    total_before = len(notifications)
    cutoff = (_dt.datetime.now() - _dt.timedelta(days=5)).strftime("%Y-%m-%d")
    notifications = [n for n in notifications
                     if (n.get("data_movimentacao") or n.get("timestamp") or "")[:10] >= cutoff]
    _save_notifications(notifications)
    removed = total_before - len(notifications)
    return jsonify({"status": "ok", "removed": removed, "remaining": len(notifications)})


@app.route("/api/legalmail/notificacao/analisar", methods=["POST"])
@require_admin
def legalmail_analisar_intimacao():
    """Analyze an intimation using Claude AI.

    Body: { "index": 0 }  (index in notifications list - legacy)
    or: { "timestamp": "..." }  (find by timestamp - preferred)
    or: { "texto": "...", "numero_processo": "..." }  (direct text)
    """
    data = request.get_json()

    # Get text to analyze
    texto_intimacao = ""
    numero_processo = ""
    notif_index = data.get("index")
    notif_timestamp = data.get("timestamp")

    # Find notification by timestamp (preferred) or index (legacy)
    notif = None
    notifications = None
    if notif_timestamp:
        notifications = _load_notifications()
        for i, n in enumerate(notifications):
            if n.get("timestamp") == notif_timestamp:
                notif = n
                notif_index = i
                break
        if notif is None:
            return jsonify({"error": "Notificação não encontrada"}), 400
    elif notif_index is not None:
        notifications = _load_notifications()
        if notif_index < 0 or notif_index >= len(notifications):
            return jsonify({"error": "Índice inválido"}), 400
        notif = notifications[notif_index]

    if notif:
        numero_processo = notif.get("numero_processo", "")
        # Check if monitor already extracted text
        if notif.get("texto_movimentacao"):
            texto_intimacao = notif["texto_movimentacao"]

        # Collect all text from documents
        docs = notif.get("documentos", [])
        for doc in docs:
            if doc.get("tipo") == "movement" and not texto_intimacao:
                # Monitor movement - try to fetch text
                mov_id = doc.get("idmovimentacoes")
                if mov_id:
                    texto = monitor_fetch_movement_text(mov_id)
                    if texto:
                        texto_intimacao += f"\n--- {doc.get('title', 'Movimentação')} ({doc.get('movement_date', '')}) ---\n"
                        texto_intimacao += texto
            elif doc.get("tipo") == "text":
                texto_intimacao += f"\n--- {doc.get('title', 'Movimentação')} ({doc.get('movement_date', '')}) ---\n"
                texto_intimacao += doc.get("text", "")
            elif doc.get("tipo") == "pdf":
                # Download and extract text from PDF
                pdf_url = doc.get("link", "")
                if pdf_url:
                    try:
                        import requests as req_lib
                        pdf_resp = req_lib.get(pdf_url, timeout=30)
                        if pdf_resp.status_code == 200:
                            pdf_doc = fitz.open(stream=pdf_resp.content, filetype="pdf")
                            pdf_text = ""
                            for page in pdf_doc:
                                pdf_text += page.get_text()
                            pdf_doc.close()
                            texto_intimacao += f"\n--- {doc.get('title', 'Documento PDF')} ({doc.get('movement_date', '')}) ---\n"
                            texto_intimacao += pdf_text
                    except Exception as e:
                        texto_intimacao += f"\n[Erro ao baixar PDF: {e}]\n"
    else:
        texto_intimacao = data.get("texto", "")
        numero_processo = data.get("numero_processo", "")

    if not texto_intimacao:
        return jsonify({"error": "Nenhum texto para analisar"}), 400

    # Use Claude to analyze
    client = anthropic.Anthropic(timeout=120.0)
    analysis_prompt = f"""Você é um assistente jurídico experiente.

Analise a movimentação processual abaixo com atenção a:
- Há PRAZO para o advogado? (intimação, contestação, recurso, cumprimento, etc.)
- Calcule a data limite do prazo considerando dias úteis quando aplicável
- Urgência real: sentença/decisão de mérito = alta, intimação com prazo = alta, mero expediente = baixa
- Indique de forma CLARA e DIRETA o que o advogado precisa fazer
- Se for sentença: indique se foi procedente, improcedente ou parcialmente procedente

Processo: {numero_processo}

TEXTO:
{texto_intimacao[:8000]}

Responda APENAS com JSON válido:
{{
    "resumo": "Resumo claro e objetivo do que aconteceu (2-3 frases)",
    "tipo_movimentacao": "intimação|despacho|sentença|decisão|citação|audiência|perícia|expediente|outro",
    "prazo_dias": número de dias do prazo (0 se não há),
    "data_prazo": "AAAA-MM-DD" ou null,
    "urgencia": "alta|media|baixa",
    "acao_necessaria": "Ação específica que o advogado deve tomar (ex: 'Interpor recurso em 15 dias úteis' ou 'Nenhuma ação necessária, apenas ciência')",
    "tipo_peticao_sugerida": "contestação|recurso|apelação|agravo|manifestação|cumprimento|embargos|impugnação|outro|nenhuma",
    "resultado_merito": "procedente|improcedente|parcialmente_procedente|null (se não for sentença/decisão de mérito)",
    "observacoes": "Pontos relevantes para o advogado"
}}"""

    try:
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1500,
            messages=[{"role": "user", "content": analysis_prompt}]
        )
        response_text = response.content[0].text.strip()

        # Parse JSON from response (handle markdown code blocks)
        if "```json" in response_text:
            response_text = response_text.split("```json")[1].split("```")[0].strip()
        elif "```" in response_text:
            response_text = response_text.split("```")[1].split("```")[0].strip()

        analysis = json.loads(response_text)

        # Save analysis back to notification if index provided
        if notif_index is not None:
            notifications = _load_notifications()
            if notif_index < len(notifications):
                notifications[notif_index]["analyzed"] = True
                notifications[notif_index]["analysis"] = analysis
                _save_notifications(notifications)

        return jsonify({"status": "ok", "analysis": analysis})

    except json.JSONDecodeError as e:
        return jsonify({"error": f"Erro ao parsear resposta da IA: {e}", "raw": response_text}), 500
    except Exception as e:
        return jsonify({"error": f"Erro na análise: {e}"}), 500


@app.route("/api/legalmail/peticao-intermediaria", methods=["POST"])
def legalmail_criar_peticao_intermediaria():
    """Create an intermediate petition for an existing process.

    Body: {
        "idprocesso": 123,
        "texto_peticao": "optional - text to generate petition",
        "pdf_path": "optional - path to existing PDF",
        "tipo_peticao": "manifestação|contestação|recurso|...",
        "anexos": ["path1.pdf", "path2.pdf"]
    }
    """
    data = request.get_json()
    idprocesso = data.get("idprocesso")
    texto_peticao = data.get("texto_peticao", "")
    pdf_path = data.get("pdf_path", "")
    tipo_peticao = data.get("tipo_peticao", "Manifestação")
    anexos = data.get("anexos", [])

    if not idprocesso:
        return jsonify({"error": "idprocesso obrigatório"}), 400

    import time

    # Step 1: Create intermediate petition
    payload = {
        "fk_processo": int(idprocesso),
        "fk_certificado": LEGALMAIL_CERT_ID,
        "tutela_antecipada": 0,
        "custas_recolhidas": 0,
    }
    resp = legalmail_request("post", "/petition/intermediate", json=payload)
    if resp.status_code == 429:
        time.sleep(60)
        resp = legalmail_request("post", "/petition/intermediate", json=payload)

    if resp.status_code != 200:
        return jsonify({"error": f"Erro ao criar petição: {resp.status_code} {resp.text[:200]}"}), 400

    try:
        resp_data = resp.json()
    except Exception:
        return jsonify({"error": f"Resposta inválida: {resp.text[:200]}"}), 400

    idpeticoes = resp_data.get("idpeticoes")
    if not idpeticoes:
        return jsonify({"error": f"Sem idpeticoes na resposta: {resp_data}"}), 400

    result = {
        "idpeticoes": idpeticoes,
        "idprocesso": idprocesso,
    }

    # Step 2: If text provided, generate PDF with Claude
    if texto_peticao and not pdf_path:
        try:
            client_ai = anthropic.Anthropic(timeout=120.0)
            gen_resp = client_ai.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4000,
                messages=[{"role": "user", "content": f"Gere o texto completo da seguinte petição intermediária ({tipo_peticao}) para protocolar nos autos. O texto deve ser formal, técnico e pronto para protocolar:\n\n{texto_peticao}"}]
            )
            peticao_texto = gen_resp.content[0].text

            # Convert to PDF using PyMuPDF
            pdf_path = os.path.join(OUTPUT_DIR, f"peticao_intermediaria_{idpeticoes}.pdf")
            doc = fitz.open()
            page = doc.new_page()
            text_rect = fitz.Rect(72, 72, page.rect.width - 72, page.rect.height - 72)
            # Split text into chunks that fit on pages
            remaining = peticao_texto
            while remaining:
                rc = page.insert_textbox(text_rect, remaining, fontsize=11, fontname="helv")
                if rc >= 0:
                    break
                # Text overflowed - need more pages
                overflow_idx = len(remaining) - int(abs(rc))
                remaining = remaining[overflow_idx:]
                if remaining.strip():
                    page = doc.new_page()
                else:
                    break
            doc.save(pdf_path)
            doc.close()
            result["generated_pdf"] = pdf_path
        except Exception as e:
            result["warn_generation"] = f"Erro ao gerar PDF: {e}"

    # Step 3: Upload petition PDF
    if pdf_path and os.path.exists(pdf_path):
        with open(pdf_path, 'rb') as pf:
            resp = legalmail_request("post",
                f"/petition/file?idpeticoes={idpeticoes}&idprocessos={idprocesso}",
                files={"file": (os.path.basename(pdf_path), pf, "application/pdf")})
        if resp.status_code == 429:
            time.sleep(60)
            with open(pdf_path, 'rb') as pf:
                resp = legalmail_request("post",
                    f"/petition/file?idpeticoes={idpeticoes}&idprocessos={idprocesso}",
                    files={"file": (os.path.basename(pdf_path), pf, "application/pdf")})
        result["petition_upload"] = resp.status_code

    # Step 4: Upload attachments
    uploaded = 0
    for anexo_path in anexos:
        if os.path.exists(anexo_path):
            time.sleep(1)
            try:
                with open(anexo_path, 'rb') as af:
                    resp = legalmail_request("post",
                        f"/petition/attachments?idpeticoes={idpeticoes}&fk_documentos_tipos=1",
                        files={"file": (os.path.basename(anexo_path), af, "application/pdf")})
                if resp.status_code == 200:
                    uploaded += 1
                elif resp.status_code == 429:
                    time.sleep(60)
                    with open(anexo_path, 'rb') as af:
                        resp = legalmail_request("post",
                            f"/petition/attachments?idpeticoes={idpeticoes}&fk_documentos_tipos=1",
                            files={"file": (os.path.basename(anexo_path), af, "application/pdf")})
                    if resp.status_code == 200:
                        uploaded += 1
            except Exception as e:
                print(f"  [WARN] Erro anexo {anexo_path}: {e}")

    result["uploaded_attachments"] = uploaded

    # Step 5: Get available peças for send
    time.sleep(1)
    resp_types = legalmail_request("get", f"/petition/types?idpeticoes={idpeticoes}")
    if resp_types.status_code == 200:
        try:
            pecas = resp_types.json().get("pecas", [])
            result["pecas_disponiveis"] = pecas
        except Exception:
            pass

    result["status"] = "ok"
    result["url"] = f"https://app.legalmail.com.br/petitions/{idpeticoes}"
    return jsonify(result)


@app.route("/api/legalmail/peticao-intermediaria/protocolar", methods=["POST"])
def legalmail_protocolar_intermediaria():
    """Send/protocol an intermediate petition.

    Body: { "idpeticoes": 123, "idprocessos": 456, "fk_peca": 1, "data_protocolo": "2026-03-10" }
    """
    data = request.get_json()
    idpeticoes = data.get("idpeticoes")
    idprocessos = data.get("idprocessos")
    fk_peca = data.get("fk_peca")
    data_protocolo = data.get("data_protocolo", "")

    if not all([idpeticoes, idprocessos, fk_peca]):
        return jsonify({"error": "idpeticoes, idprocessos e fk_peca são obrigatórios"}), 400

    params = f"idpeticoes={idpeticoes}&idprocessos={idprocessos}&fk_peca={fk_peca}"
    if data_protocolo:
        params += f"&data_protocolo={data_protocolo}"

    resp = legalmail_request("post", f"/petition/intermediate/send?{params}")
    if resp.status_code == 200:
        return jsonify(resp.json())
    return jsonify({"error": resp.text[:300]}), resp.status_code


@app.route("/api/legalmail/analisar-todos", methods=["POST"])
def legalmail_analisar_todos_pendentes():
    """Analyze all pending (unanalyzed) notifications at once."""
    notifications = _load_notifications()
    pending = [(i, n) for i, n in enumerate(notifications)
               if not n.get("analyzed") and n.get("type") == "intimacao"]

    if not pending:
        return jsonify({"status": "ok", "message": "Nenhuma notificação pendente", "analyzed": 0})

    results = []
    for idx, notif in pending:
        # Collect text
        texto = ""
        for doc in notif.get("documentos", []):
            if doc.get("tipo") == "text":
                texto += f"\n{doc.get('title', '')}: {doc.get('text', '')}"
            elif doc.get("tipo") == "pdf" and doc.get("link"):
                try:
                    import requests as req_lib
                    pdf_resp = req_lib.get(doc["link"], timeout=30)
                    if pdf_resp.status_code == 200:
                        pdf_doc = fitz.open(stream=pdf_resp.content, filetype="pdf")
                        for page in pdf_doc:
                            texto += page.get_text()
                        pdf_doc.close()
                except Exception:
                    pass

        if not texto.strip():
            continue

        # Analyze with Claude
        try:
            client_ai = anthropic.Anthropic(timeout=120.0)
            analysis_prompt = f"""Analise esta intimação processual e responda APENAS com JSON:

Processo: {notif.get('numero_processo', '')}
Tribunal: {notif.get('tribunal', '')}

TEXTO:
{texto[:6000]}

JSON formato:
{{"resumo": "...", "tipo_movimentacao": "...", "prazo_dias": 0, "data_prazo": null, "urgencia": "alta|media|baixa", "acao_necessaria": "...", "tipo_peticao_sugerida": "...", "observacoes": "..."}}"""

            response = client_ai.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=1500,
                messages=[{"role": "user", "content": analysis_prompt}]
            )
            response_text = response.content[0].text.strip()
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0].strip()

            analysis = json.loads(response_text)
            notifications[idx]["analyzed"] = True
            notifications[idx]["analysis"] = analysis
            results.append({"index": idx, "processo": notif.get("numero_processo"), "analysis": analysis})
        except Exception as e:
            results.append({"index": idx, "processo": notif.get("numero_processo"), "error": str(e)})

    _save_notifications(notifications)
    return jsonify({"status": "ok", "analyzed": len(results), "results": results})


@app.route("/api/legalmail/analisar-consolidado", methods=["POST"])
def legalmail_analisar_consolidado():
    """Analyze all movements of a process together (consolidated analysis).

    Groups all notifications by process, sends everything to Claude for a single
    intelligent analysis, and auto-generates petition/manifestation if needed.

    Body: { "numero_processo": "5011520-58.2026.4.04.7000" }
    Or: {} to analyze ALL pending processes at once.
    """
    data = request.get_json() or {}
    target_processo = data.get("numero_processo", "")

    notifications = _load_notifications()
    if not isinstance(notifications, list):
        return jsonify({"error": "Sem notificações"}), 400

    # Group notifications by process
    groups = {}
    for i, n in enumerate(notifications):
        if not isinstance(n, dict):
            continue
        num = n.get("numero_processo", "")
        if not num:
            continue
        if target_processo and num != target_processo:
            continue
        if num not in groups:
            groups[num] = {
                "indices": [],
                "tribunal": n.get("tribunal", ""),
                "polo_ativo": n.get("polo_ativo", ""),
                "polo_passivo": n.get("polo_passivo", ""),
                "classe": n.get("nome_classe", "") or n.get("classe", ""),
                "idprocesso": n.get("idprocesso", ""),
                "movimentacoes": []
            }
        groups[num]["indices"].append(i)
        groups[num]["movimentacoes"].append({
            "data": n.get("data_movimentacao", ""),
            "titulo": n.get("titulo_movimentacao", ""),
            "texto": n.get("texto_movimentacao", "")[:3000],
        })

    if not groups:
        return jsonify({"status": "ok", "message": "Nenhum processo para analisar", "results": []})

    client_ai = anthropic.Anthropic(timeout=120.0)
    results = []

    for numero, group in groups.items():
        # Build consolidated text of ALL movements
        movs_text = ""
        for m in sorted(group["movimentacoes"], key=lambda x: x.get("data", "")):
            movs_text += f"\n--- {m['data']} | {m['titulo']} ---\n"
            if m['texto']:
                movs_text += m['texto'] + "\n"

        if not movs_text.strip():
            continue

        # Consolidated analysis prompt (Skill: advogado-analise-publicacoes)
        prompt = f"""Você é um especialista PhD em Direito Processual, advogado do escritório do Dr. José Roberto da Costa Junior (OAB/SP 378.163), com expertise em:
- Intimações e publicações judiciais (PJe, DJe, e-SAJ, PROJUDI, CRETA)
- Recursos trabalhistas (RO, RR, Agravo de Instrumento, Agravo Interno)
- Recursos cíveis e previdenciários (Apelação, Agravo de Instrumento, REsp, RE)
- BPC/LOAS (Lei 8.742/93, CF art. 203, Lei 8.213/91)
- Petições intercorrentes, manifestações sobre laudos periciais, embargos de declaração
- Impugnações, contrarrazões, petições de juntada

## PROCESSO EM ANÁLISE
PROCESSO: {numero}
TRIBUNAL: {group['tribunal']}
PARTE AUTORA: {group['polo_ativo']}
PARTE RÉ: {group['polo_passivo']}
CLASSE: {group['classe']}

## MOVIMENTAÇÕES (analisar CONSOLIDADAMENTE, não uma a uma)
{movs_text[:12000]}

## FLUXO OBRIGATÓRIO

### PASSO 1 — CLASSIFICAR O ATO JUDICIAL
Identifique:
1.1 Tipo: Despacho | Decisão interlocutória | Sentença | Acórdão | Intimação para cumprimento | Laudo pericial | Despacho de execução | Auto de penhora/avaliação
1.2 Área: Trabalhista (CLT, TRT, TST) | Previdenciário/BPC-LOAS (JEF, TRF, STJ) | Cível (TJSP, STJ, STF)
1.3 O que o juízo/tribunal determinou ou decidiu (resumo em 2-3 linhas)

### PASSO 2 — PRAZO PROCESSUAL
Tabela de referência:
| Ato | Prazo | Fundamento |
|-----|-------|------------|
| Recurso Ordinário Trabalhista | 8 dias úteis | Art. 895 CLT |
| Contrarrazões ao RO | 8 dias úteis | Art. 895 CLT |
| Recurso de Revista | 8 dias úteis | Art. 896 CLT |
| Agravo de Instrumento Trabalhista | 8 dias úteis | Art. 897 CLT |
| Embargos de Declaração (Trabalhista) | 5 dias úteis | Art. 897-A CLT |
| Embargos de Declaração (CPC) | 5 dias úteis | Art. 1.023 CPC |
| Apelação Cível | 15 dias úteis | Art. 1.003 CPC |
| Agravo de Instrumento Cível | 15 dias úteis | Art. 1.016 CPC |
| Contrarrazões Apelação/AI | 15 dias úteis | Art. 1.010 CPC |
| Recurso Especial | 15 dias úteis | Art. 1.029 CPC |
| Manifestação sobre laudo pericial | 15 dias | Art. 477 CPC |
| Manifestação geral (despacho) | 5 dias (trabalhista) / 15 dias (cível) | CLT/CPC |
| Impugnação à penhora | 15 dias | Art. 525 CPC |
| Embargos à execução (Trabalhista) | 5 dias | Art. 884 CLT |
Em caso de dúvida, adotar o MENOR prazo cabível e alertar.

### PASSO 3 — IDENTIFICAR A PEÇA ADEQUADA
DESPACHO/INTIMAÇÃO -> Petição de Juntada | Petição Intermediária | Resposta a Diligência
DECISÃO INTERLOCUTÓRIA DESFAVORÁVEL -> Agravo de Instrumento
SENTENÇA -> Recurso Ordinário (Trabalhista) | Apelação (Cível/Previdenciário)
ACÓRDÃO -> Embargos de Declaração | Recurso de Revista | Recurso Especial | Agravo Interno
LAUDO PERICIAL -> Manifestação/Impugnação ao Laudo | Quesitos Complementares
FASE DE EXECUÇÃO -> Impugnação à Penhora | Embargos à Execução | Petição de Cálculos

### PASSO 4 — SE PRECISA PETIÇÃO, REDIGIR COMPLETA
Estrutura obrigatória da peça:
- Endereçamento correto ao juízo/tribunal (EXMO(A). SR(A). DR(A)...)
- Qualificação da parte (já qualificado nos autos)
- Número do processo SEMPRE em destaque
- Fundamentação da tempestividade (prazo, data publicação, data limite)
- Fundamento legal expresso (artigos de lei, súmulas, jurisprudência)
- Pedidos claros e específicos
- Local, data por extenso e assinatura (José Roberto da Costa Junior, OAB/SP 378.163)

Endereçamentos:
- JEF: "EXMO(A). SR(A). DR(A). JUIZ(A) FEDERAL DO JUIZADO ESPECIAL FEDERAL DE [CIDADE/UF]"
- TRF: "EXMO(A). SR(A). DR(A). DESEMBARGADOR(A) FEDERAL RELATOR(A) DO EGRÉGIO TRIBUNAL REGIONAL FEDERAL DA [X]ª REGIÃO"
- Vara do Trabalho: "EXMO(A). SR(A). DR(A). JUIZ(A) DO TRABALHO DA [X]ª VARA DO TRABALHO DE [CIDADE/UF]"
- TRT: "EXMO(A). SR(A). DR(A). DESEMBARGADOR(A) PRESIDENTE DA [X]ª TURMA DO TRT DA [X]ª REGIÃO"

Regras absolutas:
- ZERO traços decorativos (sem —, sem –, sem bullets desnecessários)
- ZERO estilo "gerado por IA" (sem listas desnecessárias)
- Seções em algarismos romanos (I, II, III), subseções em árabe (1, 2, 3)
- Fundamentação jurídica robusta com artigos de lei
- Nomes das partes SEMPRE em caixa alta

## OUTPUT — RESPONDA APENAS COM JSON VÁLIDO (sem markdown, sem ```)
{{
    "resumo_consolidado": "Resumo claro e objetivo do que aconteceu no processo (3-5 frases). Indicar o tipo do ato, o que foi decidido/determinado, e a consequência prática.",
    "tipo_andamento": "sentença|decisão_interlocutoria|intimação|audiência|perícia|laudo_pericial|citação|despacho|acórdão|execução|expediente|outro",
    "resultado": "procedente|improcedente|parcialmente_procedente|null",
    "area_direito": "trabalhista|previdenciário|cível",
    "prazo_dias": 0,
    "prazo_fundamento": "Art. X do CPC/CLT (descrição breve)",
    "data_prazo": "AAAA-MM-DD ou null",
    "urgencia": "alta|media|baixa",
    "acao_necessaria": "Descrição clara, direta e específica da ação que o advogado deve tomar. Ex: 'Interpor Recurso Ordinário contra sentença improcedente' ou 'Juntar documentos conforme determinado' ou 'Manifestar sobre laudo pericial desfavorável'",
    "precisa_peticao": true,
    "tipo_peticao": "recurso_ordinario|apelação|agravo_instrumento|agravo_interno|embargos_declaração|recurso_revista|recurso_especial|manifestação|impugnação_laudo|contrarrazões|petição_juntada|petição_intermediária|impugnação_penhora|embargos_execução|nenhuma",
    "texto_peticao": "TEXTO COMPLETO E FORMAL da peça processual, pronta para protocolo. Incluir: endereçamento, qualificação, fundamentação de tempestividade, mérito com fundamentação jurídica (artigos, súmulas, jurisprudência), pedidos específicos, local/data/assinatura. Se precisa_peticao=false, usar null.",
    "observacoes": "Alertas importantes: prazos curtos, riscos processuais, providências paralelas necessárias, documentos a reunir"
}}"""

        try:
            response = client_ai.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                messages=[{"role": "user", "content": prompt}]
            )
            response_text = response.content[0].text.strip()

            # Parse JSON
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0].strip()

            analysis = json.loads(response_text)

            # If petition was generated, save as PDF
            pdf_filename = None
            if analysis.get("precisa_peticao") and analysis.get("texto_peticao"):
                try:
                    safe_num = re.sub(r'[^\d\-\.]', '', numero)
                    pdf_filename = f"peticao_{safe_num}_{_dt.datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
                    pdf_path = os.path.join(OUTPUT_DIR, pdf_filename)
                    _generate_petition_pdf(analysis["texto_peticao"], pdf_path, numero, analysis.get("tipo_peticao", ""))
                    analysis["pdf_filename"] = pdf_filename
                    analysis["pdf_url"] = f"/api/download/{pdf_filename}"
                    print(f"  [MONITOR] Petição gerada: {pdf_filename}")
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    print(f"  [MONITOR] Erro ao gerar PDF da petição: {e}")

            # Update all notifications of this process with consolidated analysis
            consolidated_data = {
                "resumo_consolidado": analysis.get("resumo_consolidado", ""),
                "tipo_andamento": analysis.get("tipo_andamento", ""),
                "area_direito": analysis.get("area_direito", ""),
                "resultado": analysis.get("resultado"),
                "prazo_dias": analysis.get("prazo_dias", 0),
                "prazo_fundamento": analysis.get("prazo_fundamento", ""),
                "data_prazo": analysis.get("data_prazo"),
                "urgencia": analysis.get("urgencia", "baixa"),
                "acao_necessaria": analysis.get("acao_necessaria", ""),
                "precisa_peticao": analysis.get("precisa_peticao", False),
                "tipo_peticao": analysis.get("tipo_peticao", "nenhuma"),
                "texto_peticao": analysis.get("texto_peticao"),
                "observacoes": analysis.get("observacoes", ""),
                "consolidado": True,
            }
            if pdf_filename:
                consolidated_data["pdf_filename"] = pdf_filename
                consolidated_data["pdf_url"] = f"/api/download/{pdf_filename}"

            for idx in group["indices"]:
                if idx < len(notifications):
                    notifications[idx]["analyzed"] = True
                    notifications[idx]["analysis"] = consolidated_data

            results.append({
                "numero_processo": numero,
                "polo_ativo": group["polo_ativo"],
                "movimentacoes": len(group["movimentacoes"]),
                "analysis": analysis,
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            results.append({
                "numero_processo": numero,
                "error": str(e),
            })

    _save_notifications(notifications)
    return jsonify({
        "status": "ok",
        "processos_analisados": len(results),
        "results": results
    })


@app.route("/api/legalmail/regenerar-peticao", methods=["POST"])
def legalmail_regenerar_peticao():
    """Regenerate a petition with optional instructions.

    Body: {
        "numero_processo": "5011520-58.2026.4.04.7000",
        "instrucoes": "Focar mais na questão da deficiência, usar jurisprudência do TRF-4"
    }
    """
    data = request.get_json() or {}
    numero = data.get("numero_processo", "")
    instrucoes = data.get("instrucoes", "")

    if not numero:
        return jsonify({"error": "numero_processo obrigatório"}), 400

    notifications = _load_notifications()
    # Collect all movements for this process
    movs_text = ""
    analysis_anterior = None
    indices = []
    proc_info = {}
    for i, n in enumerate(notifications):
        if not isinstance(n, dict) or n.get("numero_processo") != numero:
            continue
        indices.append(i)
        if not proc_info:
            proc_info = {
                "tribunal": n.get("tribunal", ""),
                "polo_ativo": n.get("polo_ativo", ""),
                "polo_passivo": n.get("polo_passivo", ""),
                "classe": n.get("classe", ""),
                "idprocesso": n.get("idprocesso", ""),
            }
        if n.get("analysis") and n["analysis"].get("consolidado"):
            analysis_anterior = n["analysis"]
        m_data = n.get("data_movimentacao", "")
        m_titulo = n.get("titulo_movimentacao", "")
        m_texto = n.get("texto_movimentacao", "")[:3000]
        movs_text += f"\n--- {m_data} | {m_titulo} ---\n{m_texto}\n"

    if not movs_text.strip():
        return jsonify({"error": "Nenhuma movimentação encontrada para este processo"}), 400

    tipo_peticao = analysis_anterior.get("tipo_peticao_sugerida", "manifestação") if analysis_anterior else "manifestação"

    client_ai = anthropic.Anthropic(timeout=120.0)
    prompt = f"""Você é um especialista PhD em Direito Processual do escritório do Dr. José Roberto da Costa Junior (OAB/SP 378.163).

Gere uma peça processual do tipo "{tipo_peticao}" para o processo abaixo. O texto deve ser COMPLETO, formal, técnico e pronto para protocolar.

## DADOS DO PROCESSO
PROCESSO: {numero}
TRIBUNAL: {proc_info.get('tribunal', '')}
PARTE AUTORA: {proc_info.get('polo_ativo', '')}
PARTE RÉ: {proc_info.get('polo_passivo', '')}
CLASSE: {proc_info.get('classe', '')}

## MOVIMENTAÇÕES RECENTES
{movs_text[:10000]}

{f'## INSTRUÇÕES ADICIONAIS DO ADVOGADO (PRIORIDADE MÁXIMA):{chr(10)}{instrucoes}' if instrucoes else ''}

## ESTRUTURA OBRIGATÓRIA DA PEÇA
1. Endereçamento correto ao juízo/tribunal:
   - JEF: "EXMO(A). SR(A). DR(A). JUIZ(A) FEDERAL DO JUIZADO ESPECIAL FEDERAL DE [CIDADE/UF]"
   - TRF: "EXMO(A). SR(A). DR(A). DESEMBARGADOR(A) FEDERAL RELATOR(A) DO EGRÉGIO TRIBUNAL REGIONAL FEDERAL DA [X]ª REGIÃO"
   - Vara do Trabalho: "EXMO(A). SR(A). DR(A). JUIZ(A) DO TRABALHO DA [X]ª VARA DO TRABALHO DE [CIDADE/UF]"
   - TRT: "EXMO(A). SR(A). DR(A). DESEMBARGADOR(A) PRESIDENTE DA [X]ª TURMA DO TRT DA [X]ª REGIÃO"
2. Qualificação: "[NOME], já qualificado(a) nos autos do Processo nº [nº]..."
3. Tempestividade: demonstrar prazo, data da publicação, data limite
4. Mérito com fundamentação jurídica robusta (artigos de lei, súmulas, jurisprudência aplicável)
   - BPC/LOAS: Lei 8.742/93, CF art. 203, Decreto 6.214/07, Tema 106 STF
   - Previdenciário: Lei 8.213/91, Decreto 3.048/99
   - Trabalhista: CLT, Súmulas TST
5. Pedidos claros, específicos e numerados
6. Fechamento: local, data por extenso, "José Roberto da Costa Junior - Advogado - OAB/SP 378.163"

## REGRAS DE REDAÇÃO
- Seções em algarismos romanos (I, II, III), subseções em árabe (1, 2, 3)
- Nomes das partes SEMPRE em caixa alta
- ZERO traços decorativos (sem —, sem –, sem bullets desnecessários)
- ZERO estilo "gerado por IA" (sem listas onde cabem parágrafos)
- Linguagem jurídica formal e técnica

Responda APENAS com JSON válido (sem markdown, sem ```):
{{
    "tipo_peticao": "{tipo_peticao}",
    "texto_peticao": "TEXTO COMPLETO DA PEÇA PROCESSUAL"
}}"""

    try:
        response = client_ai.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=16000,
            messages=[{"role": "user", "content": prompt}]
        )
        response_text = response.content[0].text.strip()
        if "```json" in response_text:
            response_text = response_text.split("```json")[1].split("```")[0].strip()
        elif "```" in response_text:
            response_text = response_text.split("```")[1].split("```")[0].strip()

        result = json.loads(response_text)

        # Save as PDF
        safe_num = re.sub(r'[^\d\-\.]', '', numero)
        pdf_filename = f"peticao_{safe_num}_{_dt.datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        pdf_path = os.path.join(OUTPUT_DIR, pdf_filename)
        _generate_petition_pdf(result.get("texto_peticao", ""), pdf_path, numero, result.get("tipo_peticao", ""))

        return jsonify({
            "status": "ok",
            "tipo_peticao": result.get("tipo_peticao", tipo_peticao),
            "texto_peticao": result.get("texto_peticao", ""),
            "pdf_filename": pdf_filename,
            "pdf_url": f"/api/download/{pdf_filename}",
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/legalmail/buscar-processo", methods=["GET"])
def legalmail_buscar_processo():
    """Busca processo por número ou nome da parte — usa cache local."""
    numero = request.args.get("numero", "").strip()
    nome = request.args.get("nome", "").strip()

    if not numero and not nome:
        return jsonify({"error": "Informe numero ou nome"}), 400

    # Carrega cache local do monitor (evita rate limit)
    cache_path = os.path.join(BASE_DIR, "processes_cache.json")
    if os.path.exists(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            todos = json.load(f)
    else:
        # Sem cache, busca primeira página da API
        resp = legalmail_request("get", "/process/all?offset=0&limit=50")
        todos = resp.json() if resp.status_code == 200 and isinstance(resp.json(), list) else []

    if numero:
        encontrados = [p for p in todos if (p.get("numero_processo") or "") == numero]
    elif nome:
        nome_lower = nome.lower()
        encontrados = [
            p for p in todos
            if nome_lower in (p.get("poloativo_nome") or "").lower()
            or nome_lower in (p.get("polopassivo_nome") or "").lower()
        ]
    else:
        encontrados = []

    # Enriquece com movimentações da API (com fallback nas notificações cacheadas)
    notificacoes = _load_notifications()
    for p in encontrados:
        idprocesso = p.get("idprocessos")
        numero_proc = p.get("numero_processo", "")
        movs = []
        try:
            resp_autos = legalmail_request("get", f"/process/autos?idprocesso={idprocesso}")
            if resp_autos.status_code == 200:
                autos = resp_autos.json()
                if isinstance(autos, list):
                    movs = autos[:10]
        except Exception:
            pass
        # Fallback: usa notificações cacheadas do monitor
        if not movs:
            movs = [
                {"titulo": n.get("titulo_movimentacao", ""), "data_movimentacao": n.get("data_movimentacao", "")}
                for n in notificacoes
                if n.get("numero_processo") == numero_proc or str(n.get("idprocesso")) == str(idprocesso)
            ][:10]
        p["movimentacoes"] = movs

    return jsonify({"processos": encontrados})


@app.route("/api/legalmail/gerar-mensagem-cliente", methods=["POST"])
def legalmail_gerar_mensagem_cliente():
    """Gera mensagem simples de WhatsApp explicando o andamento do processo ao cliente."""
    data = request.get_json() or {}
    processo = data.get("processo", {})
    movimentacoes = data.get("movimentacoes", [])
    instrucao_extra = data.get("instrucao", "")

    numero = processo.get("numero_processo", "")
    polo_ativo = processo.get("poloativo_nome", "")
    tribunal = processo.get("tribunal", "")
    classe = processo.get("nome_classe") or processo.get("abreviatura_classe", "")
    juizo = processo.get("juizo", "")
    data_prazo = processo.get("data_prazo", "")

    movs_texto = ""
    for m in movimentacoes[:5]:
        movs_texto += f"- {m.get('data_movimentacao', '')}: {m.get('titulo', '')}\n"

    prompt = f"""Você é um advogado que precisa enviar uma mensagem simples pelo WhatsApp para o cliente explicando o andamento do processo dele.

Dados do processo:
- Número: {numero}
- Cliente (polo ativo): {polo_ativo}
- Tribunal: {tribunal}
- Classe: {classe}
- Juízo: {juizo}
- Prazo: {data_prazo or 'Nenhum prazo imediato'}

Últimas movimentações:
{movs_texto or 'Nenhuma movimentação recente.'}

{f'Instrução adicional: {instrucao_extra}' if instrucao_extra else ''}

Escreva uma mensagem curta, clara e em linguagem simples (sem juridiquês) para o cliente. Use o primeiro nome dele. Explique o que está acontecendo no processo e o que ele precisa saber. Se houver prazo, mencione. Não inclua saudações formais, apenas a mensagem direta. Não inclua títulos, cabeçalhos ou marcações como "# Mensagem para WhatsApp" ou similares. Nunca mencione valores financeiros do processo."""

    try:
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY não configurada"}), 500
        client_ai = anthropic.Anthropic(api_key=api_key, timeout=60.0)
        response = client_ai.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=600,
            messages=[{"role": "user", "content": prompt}]
        )
        mensagem = response.content[0].text.strip()
        return jsonify({"mensagem": mensagem})
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {str(e)}"}), 500


@app.route("/api/health")
@require_admin
def health_check():
    """Test API connectivity and DB status."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    result = {
        "api_key_set": bool(api_key),
        "database_url_set": bool(DATABASE_URL),
        "use_db": USE_DB,
    }
    # Test DB connection - try even if USE_DB is False to show the real error
    result["database_url_preview"] = DATABASE_URL[:50] + "..." if DATABASE_URL else "empty"
    if USE_DB:
        try:
            conn = _get_db()
            cur = conn.cursor()
            cur.execute("SELECT key, LENGTH(value::text) as size FROM kv_store")
            rows = cur.fetchall()
            result["db_status"] = "connected"
            result["db_keys"] = {row[0]: row[1] for row in rows}
            cur.close()
            conn.close()
        except Exception as e:
            result["db_status"] = f"error: {e}"
    else:
        # Try to connect now to show the actual error
        result["db_status"] = "disabled_at_startup"
        if DATABASE_URL:
            try:
                import psycopg2 as _pg2
                conn = _pg2.connect(DATABASE_URL, connect_timeout=5)
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.close()
                conn.close()
                result["db_test_now"] = "SUCCESS - connection works! But USE_DB was set to False at startup."
            except Exception as e:
                result["db_test_now"] = f"FAILED: {type(e).__name__}: {e}"
    # Test AI
    try:
        client_ai = anthropic.Anthropic(api_key=api_key, timeout=30.0)
        response = client_ai.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=5,
            messages=[{"role": "user", "content": "oi"}]
        )
        result["api_status"] = "ok"
        result["response"] = response.content[0].text[:50]
    except Exception as e:
        result["api_status"] = "error"
        result["error"] = f"{type(e).__name__}: {str(e)}"
    return jsonify(result)


@app.route("/monitoramento")
def monitoramento_page():
    """Page for process monitoring and intimation analysis."""
    return render_template("monitoramento.html")


# ==================== WHATSAPP BOT (ConversApp) ====================

CONVERSAPP_API_TOKEN = os.environ.get("CONVERSAPP_API_TOKEN", "")  # Bearer token: pn_xxxx
CONVERSAPP_API_BASE = "https://api.wts.chat"
CONVERSAPP_CHANNEL_ID = os.environ.get("CONVERSAPP_CHANNEL_ID", "5395dbba-34f9-42a5-852f-77e5e11a7c94")

# ElevenLabs TTS
ELEVENLABS_API_KEY = os.environ.get("ELEVENLABS_API_KEY", "")
ELEVENLABS_VOICE_ID = os.environ.get("ELEVENLABS_VOICE_ID", "33B4UnXyTNbgLmdEDh5P")

# Business hours (Brazil timezone)
BOT_HORA_INICIO = 8   # 8h
BOT_HORA_FIM = 21     # 21h (temporário para teste)

# Conversation state per phone number
# Persisted in DB if available, otherwise in-memory
_SESSIONS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "whatsapp_sessions.json")
_PAUSED_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "paused_phones.json")
_sessions_file_lock = threading.Lock()
_paused_file_lock = threading.Lock()

def _save_sessions():
    """Persist sessions to disk."""
    try:
        serializable = {}
        for phone, sess in _whatsapp_sessions.items():
            s = dict(sess)
            # Remove non-serializable items
            for key in list(s.keys()):
                if callable(s[key]) or key.startswith('_lock'):
                    del s[key]
            serializable[phone] = s
        _safe_json_save(_SESSIONS_FILE, serializable, lock=_sessions_file_lock)
    except Exception as e:
        print(f"[WARN] Falha ao salvar sessões: {e}")

def _load_sessions():
    """Load sessions from disk."""
    try:
        data = _safe_json_load(_SESSIONS_FILE, lock=_sessions_file_lock)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}

def _save_paused():
    """Persist paused phones to disk."""
    try:
        _safe_json_save(_PAUSED_FILE, dict(_paused_phones), lock=_paused_file_lock)
    except Exception:
        pass

def _load_paused():
    """Load paused phones from disk."""
    try:
        data = _safe_json_load(_PAUSED_FILE, lock=_paused_file_lock)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}

_whatsapp_sessions = _load_sessions()
_whatsapp_locks = {}  # Per-phone threading locks to prevent race conditions
_whatsapp_locks_lock = threading.Lock()  # Lock for accessing _whatsapp_locks
_processed_msg_ids = set()  # Track processed message IDs for deduplication
_processed_msg_ids_lock = threading.Lock()  # Lock for thread-safe dedup
_session_last_activity = {}  # Track last activity per phone for cleanup
_paused_phones = _load_paused()  # Phones where Ana is paused (human takeover) - {phone: timestamp}
MICHELLE_USER_ID = "95f8cfa9-89e1-4ef6-af46-511651ba492f"


def conversapp_complete_session(session_id):
    """Complete/close a ConversApp session."""
    try:
        resp = conversapp_request("put", f"/chat/v1/session/{session_id}/complete", json={})
        if resp.status_code == 200:
            print(f"[CONVERSAPP] Sessão {session_id} concluída")
            return True
        else:
            print(f"[CONVERSAPP] Erro ao concluir sessão {session_id}: {resp.status_code} {resp.text[:200]}")
            return False
    except Exception as e:
        print(f"[CONVERSAPP] Erro ao concluir sessão: {e}")
        return False


def conversapp_transfer_session(session_id, user_id=None):
    """Transfer a ConversApp session to another agent."""
    user_id = user_id or MICHELLE_USER_ID
    try:
        resp = conversapp_request("put", f"/chat/v1/session/{session_id}/transfer", json={"userId": user_id})
        if resp.status_code == 200:
            print(f"[CONVERSAPP] Sessão {session_id} transferida para {user_id}")
            return True
        else:
            print(f"[CONVERSAPP] Erro ao transferir sessão {session_id}: {resp.status_code} {resp.text[:200]}")
            return False
    except Exception as e:
        print(f"[CONVERSAPP] Erro ao transferir sessão: {e}")
        return False


# ========== FOLLOW-UP INTELIGENTE DE DOCUMENTOS ==========

import datetime as _dt_followup

FOLLOWUP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "followup_queue.json")
_followup_file_lock = threading.Lock()

def _followup_load():
    """Load follow-up queue from disk (thread-safe, with backup fallback)."""
    return _safe_json_load(FOLLOWUP_FILE, lock=_followup_file_lock)

def _followup_save(queue):
    """Save follow-up queue to disk (thread-safe, atomic with backup)."""
    _safe_json_save(FOLLOWUP_FILE, queue, lock=_followup_file_lock)

def _followup_add(phone, nome, docs_pendentes, data_prometida=None, session_id=None, contexto=""):
    """Add or update a follow-up entry for a client."""
    queue = _followup_load()
    phone_clean = re.sub(r'[^\d]', '', str(phone))
    hoje = _dt_followup.date.today().isoformat()

    entry = queue.get(phone_clean, {})
    entry.update({
        "nome": nome,
        "phone": phone_clean,
        "docs_pendentes": docs_pendentes,
        "data_prometida": data_prometida,
        "session_id": session_id,
        "contexto": contexto[:500],
        "criado_em": entry.get("criado_em", hoje),
        "atualizado_em": hoje,
        "tentativas": entry.get("tentativas", 0),
        "ultimo_followup": entry.get("ultimo_followup"),
        "status": "pendente",
    })
    queue[phone_clean] = entry
    _followup_save(queue)
    print(f"[FOLLOWUP] Adicionado/atualizado: {phone_clean} ({nome}) - docs: {docs_pendentes}, data: {data_prometida}")
    return entry

def _followup_remove(phone):
    """Remove a client from follow-up queue."""
    queue = _followup_load()
    phone_clean = re.sub(r'[^\d]', '', str(phone))
    removed = queue.pop(phone_clean, None)
    if removed:
        _followup_save(queue)
        print(f"[FOLLOWUP] Removido: {phone_clean}")
    return removed

def _followup_detect_date(text):
    """Detect date promises in conversation text. Returns ISO date string or None."""
    hoje = _dt_followup.date.today()
    text_lower = text.lower().strip()

    # "hoje"
    if text_lower in ("hoje", "hoje mesmo", "agora", "já vou enviar", "ja vou enviar"):
        return hoje.isoformat()

    # "depois de amanhã" (check before "amanhã")
    if "depois de amanhã" in text_lower or "depois de amanha" in text_lower:
        return (hoje + _dt_followup.timedelta(days=2)).isoformat()

    # "amanhã"
    if "amanhã" in text_lower or "amanha" in text_lower:
        return (hoje + _dt_followup.timedelta(days=1)).isoformat()

    # "dia 15", "dia 20"
    m = re.search(r'dia\s+(\d{1,2})', text_lower)
    if m:
        dia = int(m.group(1))
        if 1 <= dia <= 31:
            try:
                data = hoje.replace(day=dia)
                if data < hoje:
                    mes = hoje.month + 1
                    ano = hoje.year
                    if mes > 12:
                        mes = 1
                        ano += 1
                    data = hoje.replace(year=ano, month=mes, day=dia)
                return data.isoformat()
            except ValueError:
                pass

    # "segunda", "terça", etc.
    dias_semana = {
        "segunda": 0, "terça": 1, "terca": 1, "quarta": 2,
        "quinta": 3, "sexta": 4, "sábado": 5, "sabado": 5, "domingo": 6
    }
    for nome_dia, num_dia in dias_semana.items():
        if nome_dia in text_lower:
            dias_ate = (num_dia - hoje.weekday()) % 7
            if dias_ate == 0:
                dias_ate = 7  # Next week
            return (hoje + _dt_followup.timedelta(days=dias_ate)).isoformat()

    # "semana que vem", "próxima semana"
    if "semana que vem" in text_lower or "próxima semana" in text_lower or "proxima semana" in text_lower:
        return (hoje + _dt_followup.timedelta(days=7)).isoformat()

    return None

def _followup_detect_docs(text):
    """Detect document mentions in text. Returns list of doc names."""
    text_lower = text.lower()
    docs_map = {
        "laudo": "laudo médico",
        "laudo médico": "laudo médico",
        "laudo medico": "laudo médico",
        "cad único": "CadÚnico",
        "cad unico": "CadÚnico",
        "cadúnico": "CadÚnico",
        "cadunico": "CadÚnico",
        "cadastro único": "CadÚnico",
        "comprovante de residência": "comprovante de residência",
        "comprovante de residencia": "comprovante de residência",
        "comprovante residência": "comprovante de residência",
        "comprovante residencia": "comprovante de residência",
        "comprovante de endereço": "comprovante de residência",
        "comprovante de endereco": "comprovante de residência",
        "rg": "RG",
        "identidade": "RG",
        "cpf": "CPF",
        "certidão de nascimento": "certidão de nascimento",
        "certidao de nascimento": "certidão de nascimento",
        "receita": "receita médica",
        "exame": "exame médico",
        "declaração": "declaração",
        "declaracao": "declaração",
        "contrato": "contrato",
        "procuração": "procuração",
        "procuracao": "procuração",
    }
    found = set()
    for keyword, doc_name in docs_map.items():
        if keyword in text_lower:
            found.add(doc_name)
    return list(found)

def _followup_read_conversation(session_id):
    """Read conversation history from ConversApp session."""
    try:
        msgs = []
        resp = conversapp_request("get", f"/chat/v1/session/{session_id}/message?pageSize=50")
        if resp.status_code == 200:
            data = resp.json()
            for msg in data.get("items", []):
                direction = "cliente" if msg.get("direction") != "TO_HUB" else "ana"
                text = msg.get("text") or ""
                msg_type = msg.get("type") or "TEXT"
                if text:
                    msgs.append({"de": direction, "texto": text, "tipo": msg_type})
                elif msg_type in ("IMAGE", "PHOTO"):
                    msgs.append({"de": direction, "texto": "[imagem enviada]", "tipo": msg_type})
        return msgs
    except Exception as e:
        print(f"[FOLLOWUP] Erro ao ler conversa: {e}")
        return []

def _followup_generate_message(entry, conversation_msgs):
    """Use Claude to generate a personalized follow-up message based on conversation context."""
    try:
        nome = entry.get("nome", "")
        docs = entry.get("docs_pendentes", [])
        tentativas = entry.get("tentativas", 0)
        data_prometida = entry.get("data_prometida", "")
        contexto = entry.get("contexto", "")

        # Build conversation summary (last 15 messages)
        conv_text = ""
        for msg in conversation_msgs[-15:]:
            de = "Cliente" if msg["de"] == "cliente" else "Ana"
            conv_text += f"{de}: {msg['texto'][:200]}\n"

        urgencia = "gentil"
        if tentativas == 1:
            urgencia = "lembrete amigável"
        elif tentativas >= 2:
            urgencia = "mais direto, mostrando importância"

        result = ai_chat(
            messages=[{"role": "user", "content": "Gere a mensagem de follow-up."}],
            system=f"""Você é Ana, do pós-venda da JRC Advocacia. Precisa enviar uma mensagem de follow-up para um cliente sobre documentos pendentes.

CONTEXTO:
- Nome do cliente: {nome}
- Documentos pendentes: {', '.join(docs) if docs else 'documentos gerais'}
- Data que o cliente prometeu enviar: {data_prometida or 'não especificou'}
- Número de vezes que já foi cobrado: {tentativas}
- Tom da mensagem: {urgencia}

HISTÓRICO DA ÚLTIMA CONVERSA:
{conv_text or contexto or 'Sem histórico disponível'}

REGRAS:
- Mensagem curta (2-3 linhas máximo)
- Humanizada, como se fosse uma pessoa real mandando mensagem
- NÃO use emojis
- NÃO pareça automático/robótico
- Se o cliente prometeu uma data, mencione de forma natural
- Se já cobrou antes, varie o texto (não repita a mesma mensagem)
- Chame pelo primeiro nome
- Não pressione demais, seja acolhedora

Escreva APENAS a mensagem, sem explicações.""",
            max_tokens=300,
        )
        return (result or "").strip()
    except Exception as e:
        print(f"[FOLLOWUP] Erro ao gerar mensagem: {e}")
        nome_primeiro = (entry.get("nome") or "").split()[0] if entry.get("nome") else ""
        return f"Oi {nome_primeiro}! Tudo bem? Passando pra lembrar sobre os documentos pendentes do seu processo. Quando puder, manda por aqui que eu já encaminho pro escritório!"


def _get_phone_lock(phone):
    """Get or create a threading lock for a specific phone number."""
    with _whatsapp_locks_lock:
        if phone not in _whatsapp_locks:
            _whatsapp_locks[phone] = threading.Lock()
        return _whatsapp_locks[phone]


def _cleanup_old_sessions():
    """Remove sessions inactive for more than 24 hours."""
    from datetime import datetime, timedelta
    cutoff = datetime.now() - timedelta(hours=24)
    phones_to_remove = []
    for phone, last_time in list(_session_last_activity.items()):
        if last_time < cutoff:
            phones_to_remove.append(phone)
    for phone in phones_to_remove:
        _whatsapp_sessions.pop(phone, None)
        _session_last_activity.pop(phone, None)
        with _whatsapp_locks_lock:
            _whatsapp_locks.pop(phone, None)
    if phones_to_remove:
        _save_sessions()
        print(f"[BOT] Limpou {len(phones_to_remove)} sessões inativas")
    # Also trim dedup set if too large
    with _processed_msg_ids_lock:
        if len(_processed_msg_ids) > 1000:
            _processed_msg_ids.clear()

BOT_SYSTEM_PROMPT = """Você é Ana, do pós-venda da JRC Advocacia.

Você é uma PESSOA REAL que trabalha no escritório. JAMAIS diga que é bot, assistente virtual, inteligência artificial ou automação. Você é a Ana e ponto final.

Seu objetivo é atender clientes que já contrataram o escritório, informar sobre andamento processual e responder dúvidas com clareza, calor humano e linguagem simples.

---

PERSONALIDADE

Tom acolhedor, paciente e direto. Fale como uma pessoa real, não como robô.
Linguagem simples, sem juridiquês. Explique tudo como se fosse para alguém que nunca entrou em um fórum.
Nunca demonstre impaciência. O cliente pode estar ansioso e vulnerável.
Sem emojis em mensagens de texto corrido e conversas.
Sem negrito fora do resumo do processo.

---

FLUXO DE ATENDIMENTO

PASSO 1 — Abertura

Quando o cliente enviar a primeira mensagem:
"[Saudação]! Aqui é a Ana, da JRC Advocacia. Tudo bem com você?"

Aguarda resposta do cliente.

Se bem -> "Que bom! Em que posso te ajudar?"
Se o cliente perguntar como a Ana está -> "Ótimo, obrigada! Em que posso te ajudar?"
Se mal ou ansioso -> "Sinto muito. Entendo que esse período de espera é difícil. Pode contar comigo. Em que posso te ajudar?"

PASSO 2 — Identificação

REGRA IMPORTANTE: Se o cliente JÁ indicou o tipo de caso (benefício, INSS, trabalhista, BPC, LOAS, aposentadoria, filho, etc.) NÃO repita a pergunta. Vá direto pedir o nome.

Só pergunte o tipo se REALMENTE não ficou claro. Exemplo:
- "quero saber do meu processo" -> pode perguntar o tipo
- "quero saber do benefício do meu filho" -> JÁ SABE que é INSS e é do filho, pedir nome direto
- "como tá o INSS?" -> JÁ SABE que é INSS, pedir nome direto
- "benefício do INSS" -> JÁ SABE, pedir nome direto

Pedir o nome:
- Se for para o próprio cliente -> "Pode me informar o seu nome completo, por gentileza?"
- Se for para outra pessoa -> "Pode me informar o nome completo de quem fez o pedido do benefício?"
- Se não sabe pra quem é -> "Pode me informar o nome completo de quem tem o processo?"

PASSO 3 — Aguardar nome

Após pedir o nome, APENAS aguarde o cliente responder. NÃO diga "vou consultar" ou "um momento" - o sistema faz isso automaticamente quando encontrar os dados.

PASSO 4 — Entrega do andamento judicial

Se encontrar mais de um processo, informar automaticamente o mais recente com base na data de abertura. Nunca perguntar ao cliente qual processo quer ver.

Usar sempre este formato (apenas aqui pode usar negrito e emojis):

📋 Andamento do seu processo

⚖️ *Processo nº:* [número]
🏛️ *Tribunal:* [nome]
📅 *Última movimentação:* [data]
✅ *O que aconteceu:* [tradução simples]
⏳ *Próximos passos:* [o que esperar]

Traduções obrigatórias:
"Petição inicial distribuída" -> "Seu processo foi aberto e registrado no sistema da Justiça Federal."
"Designada audiência de instrução" -> "Foi marcada uma audiência. Em breve o escritório vai te contatar com mais detalhes."
"Perícia médica agendada" -> "A Justiça marcou uma perícia médica para você. O escritório vai te avisar a data assim que confirmar."
"Sentença prolatada" -> "O juiz já deu a decisão no seu processo. O escritório está analisando e vai te informar o resultado."
"Trânsito em julgado" -> "O processo foi finalizado com decisão definitiva. O escritório vai entrar em contato para explicar os próximos passos."

Movimentações de segunda instância:
"Apelação interposta", "Remetido ao Tribunal", "Concluso ao relator", "Incluído em pauta de julgamento", "Acórdão publicado"
-> "Seu processo está na fase de recurso e aguarda julgamento pelos desembargadores. Essa etapa pode levar alguns meses, mas o escritório está acompanhando."

PASSO 5 — Encerramento

Após entregar a informação principal:
"Ficou alguma dúvida ou posso te ajudar com mais alguma coisa?"

Se não -> "Ótimo! Qualquer coisa é só chamar. Tenha um bom dia!" e adicione [ENCERRAR_SESSAO] no final da sua resposta.
Se sim -> responda e repita até encerrar naturalmente.

Quando o cliente encerrar a conversa (ex: "obrigado", "valeu", "ok", "tchau", "era só isso", "não, obrigada"), envie a despedida e adicione [ENCERRAR_SESSAO] no final.

Exemplos de encerramento:
- Cliente: "obrigada" -> "De nada! Qualquer coisa é só chamar. Tenha um ótimo dia! [ENCERRAR_SESSAO]"
- Cliente: "era só isso mesmo" -> "Ótimo! Estou por aqui se precisar. Tenha um bom dia! [ENCERRAR_SESSAO]"
- Cliente: "ok obrigado" -> "Que bom que pude ajudar! Qualquer dúvida é só chamar. [ENCERRAR_SESSAO]"

IMPORTANTE: O marcador [ENCERRAR_SESSAO] é invisível para o cliente, serve apenas como sinal interno. Só use quando o cliente CLARAMENTE encerrou a conversa.

---

REGRAS DE COMPORTAMENTO

SEMPRE:
Confirme o nome correto antes de consultar.
Traduza todos os termos jurídicos para linguagem simples.
Se o cliente parecer ansioso, valide: "Entendo, faz sentido estar ansioso nesse momento."
Se for boa notícia, comemore de forma natural e contida.

NUNCA:
Nunca revele que é bot, IA, assistente virtual ou automação.
Nunca dê opinião jurídica ou fale sobre o mérito do caso.
Nunca informe prazos definitivos sem confirmação da equipe.
Nunca compartilhe dados de outros clientes.
Nunca prometa resultados.
Nunca mencione cancelamento de processo ao cliente.
Nunca fale mal de outros advogados, do INSS ou do sistema judicial.
Nunca ofereça ligação ou transferência proativamente.
Nunca prometa que alguém vai ligar.

---

QUANDO ESCALAR PARA A EQUIPE HUMANA

Encaminhe apenas quando:
O cliente mencionar que recebeu uma carta, citação ou intimação em casa.
O cliente falar em desistir do processo.
Houver reclamação sobre o atendimento ou cobrança indevida.
A situação envolver urgência médica ou risco de vida.
O processo tiver decisão negativa.
Não encontrar nada no sistema.
A dúvida não puder ser resolvida pelo chat.

Mensagem padrão:
"Vou te transferir para a Michelle do nosso pós-venda, ela vai te ajudar melhor com isso!" e adicione [TRANSFERIR_MICHELLE] no final.

IMPORTANTE: O marcador [TRANSFERIR_MICHELLE] é invisível para o cliente, serve apenas como sinal interno. Só use quando realmente precisar escalar.

---

RESPOSTAS PARA SITUAÇÕES COMUNS

Quando vai receber:
"O pagamento acontece após a sentença ser executada e o INSS incluir no sistema de pagamentos. Esse processo pode levar alguns meses após a decisão final. O escritório vai te avisar assim que tiver uma previsão."

Se vai ganhar:
"Essa análise só o Dr. José Roberto consegue fazer com precisão, pois depende de muitos detalhes do seu caso. O que posso te dizer é que o escritório só entra com processos que considera viável."

Demora no processo:
"Entendo a ansiedade. Processos de BPC/LOAS costumam demorar entre 1 e 3 anos pela sobrecarga da Justiça Federal, não por falta de acompanhamento. Vou verificar agora se há alguma movimentação recente para você."

Sobre a perícia:
"A perícia médica é um passo muito importante. O juiz designa um perito independente para avaliar sua condição de saúde. É fundamental comparecer e levar todos os documentos médicos. O escritório vai te orientar sobre como se preparar."

Honorários/valores:
"Isso o escritório conversa direto com você."

---

PERGUNTAS COMUNS APÓS INFORMAR MOVIMENTAÇÃO

"Ganhei ou perdi?"
"Essa informação ainda não consigo te confirmar por aqui. O escritório vai te contatar assim que a análise estiver pronta."

"Quanto deu?" ou "Qual o valor?"
"O valor exato ainda está sendo calculado. Assim que estiver definido o escritório te avisa."

"Vai demorar muito?"
"Depende muito da fase em que o processo está e da fila da Justiça. Infelizmente não tenho como te dar uma previsão exata, mas posso te dizer que o escritório está acompanhando de perto."

"Quando vou receber?"
"O pagamento acontece após a decisão ser finalizada e o INSS incluir no sistema. Esse prazo pode variar bastante. O escritório vai te avisar assim que tiver uma previsão concreta."

"O que acontece agora?"
"O processo segue o trâmite normal da Justiça. Cada etapa tem seu prazo e o escritório está acompanhando tudo. Tem mais alguma dúvida que posso te ajudar?"

"Preciso fazer alguma coisa?"
"Por enquanto não. Caso precise de algo da sua parte, o escritório vai te contatar com as orientações necessárias."

REGRA GERAL PARA PERGUNTAS SOBRE RESULTADO E VALORES:
Nunca confirmar resultado, valores ou prazos.
Nunca criar expectativa positiva ou negativa.
Responder de forma acolhedora, curta e encerrar o assunto com naturalidade.
Não mencionar equipe ou advogado a menos que seja o único caminho.

---

SAUDAÇÃO CONFORME HORÁRIO
6h-12h: "Bom dia"
12h-18h: "Boa tarde"
18h-24h/0h-6h: "Boa noite"
"""

BOT_MSG_FORA_HORARIO = """Oi! Aqui é a Ana, da JRC Advocacia.

Agora já saí do escritório, nosso horário é de segunda a sexta, das 8h às 19h.

Pode deixar sua mensagem que amanhã cedo já te respondo!"""


def conversapp_request(method, endpoint, **kwargs):
    """Make authenticated request to Helena CRM API."""
    url = f"{CONVERSAPP_API_BASE}{endpoint}"
    headers = kwargs.pop("headers", {})
    headers["Authorization"] = f"Bearer {CONVERSAPP_API_TOKEN}"
    headers["Content-Type"] = "application/json"
    return getattr(requests, method)(url, headers=headers, timeout=30, **kwargs)


def elevenlabs_tts(text):
    """Convert text to speech using ElevenLabs API. Returns audio bytes (mp3) or None."""
    if not ELEVENLABS_API_KEY:
        print("[TTS] ElevenLabs API key não configurada")
        return None
    try:
        # Clean text: remove emojis, markdown bold, etc for cleaner speech
        import re as _re_tts
        clean = _re_tts.sub(r'[*_~`]', '', text)  # Remove markdown
        clean = _re_tts.sub(r'[\U0001F300-\U0001F9FF\U00002702-\U000027B0\U0000FE00-\U0000FEFF]', '', clean)  # Remove emojis
        clean = clean.strip()
        if not clean or len(clean) < 5:
            return None

        resp = requests.post(
            f"https://api.elevenlabs.io/v1/text-to-speech/{ELEVENLABS_VOICE_ID}",
            headers={
                "xi-api-key": ELEVENLABS_API_KEY,
                "Content-Type": "application/json",
            },
            json={
                "text": clean,
                "model_id": "eleven_multilingual_v2",
                "voice_settings": {
                    "stability": 0.5,
                    "similarity_boost": 0.75,
                }
            },
            timeout=30,
        )
        if resp.status_code == 200 and len(resp.content) > 1000:
            print(f"[TTS] Áudio gerado: {len(resp.content)} bytes")
            return resp.content
        else:
            print(f"[TTS] Erro: {resp.status_code} {resp.text[:200]}")
            return None
    except Exception as e:
        print(f"[TTS] Erro: {e}")
        return None


def conversapp_send_audio(phone, audio_bytes, session_id=None):
    """Send audio message via ConversApp. Uploads to S3 then sends."""
    if not CONVERSAPP_API_TOKEN or not audio_bytes:
        return False
    try:
        # Step 1: Get upload URL from ConversApp
        upload_resp = conversapp_request("post", "/core/v1/file/upload",
            json={"fileName": "audio_ana.mp3", "contentType": "audio/mpeg"})
        if upload_resp.status_code != 200:
            print(f"[TTS] Erro ao obter URL de upload: {upload_resp.status_code}")
            return False
        upload_data = upload_resp.json()
        upload_url = upload_data.get("urlUpload")
        file_key = upload_data.get("keyS3")

        if not upload_url:
            print("[TTS] URL de upload não recebida")
            return False

        # Step 2: Upload audio to S3
        s3_resp = requests.put(upload_url, data=audio_bytes,
            headers={"Content-Type": "audio/mpeg"}, timeout=30)
        if s3_resp.status_code not in (200, 201):
            print(f"[TTS] Erro ao fazer upload S3: {s3_resp.status_code}")
            return False

        # Step 3: Get the public URL (remove query params from upload URL)
        file_url = upload_url.split("?")[0]

        # Step 4: Send audio message
        if session_id:
            msg_resp = conversapp_request("post", f"/chat/v1/session/{session_id}/message",
                json={"fileUrl": file_url, "fileName": "audio.mp3"})
        else:
            msg_resp = conversapp_request("post", "/chat/v1/message/send",
                json={
                    "from": CONVERSAPP_CHANNEL_ID,
                    "to": phone,
                    "body": {"fileUrl": file_url, "fileName": "audio.mp3"}
                })

        if msg_resp.status_code in (200, 201):
            print(f"[TTS] Áudio enviado para {phone}")
            return True
        else:
            print(f"[TTS] Erro ao enviar áudio: {msg_resp.status_code} {msg_resp.text[:200]}")
            return False
    except Exception as e:
        print(f"[TTS] Erro: {e}")
        return False


def conversapp_get_contact(phone):
    """Get contact data from ConversApp by phone number."""
    try:
        # Normalize phone: ensure it has country code
        phone_clean = re.sub(r'[^\d+]', '', str(phone)).lstrip('+')
        if not phone_clean.startswith('55'):
            phone_clean = '55' + phone_clean
        resp = conversapp_request("get", f"/core/v1/contact/phonenumber/+{phone_clean}?IncludeDetails=CustomFields,Tags")
        if resp.status_code == 200:
            data = resp.json()
            print(f"[CONVERSAPP] Contato encontrado: {data.get('name', '?')}")
            return data
        else:
            print(f"[CONVERSAPP] Contato não encontrado para +{phone_clean}: {resp.status_code}")
            return None
    except Exception as e:
        print(f"[CONVERSAPP] Erro ao buscar contato: {e}")
        return None


def conversapp_update_contact(contact_id, updates):
    """Update contact fields in ConversApp.

    updates dict can contain:
    - customFields: {"key": "value"} for CPF, processo, tipo, status
    - tagIds: ["tag-id-1", ...] for labels
    - annotation: "text" for internal notes
    """
    try:
        # Always specify which fields we're updating
        fields = ["CustomFields"]
        if "tagIds" in updates:
            fields.append("Tags")
        if "annotation" in updates:
            fields.append("Annotation")

        body = {"fields": fields, **updates}
        resp = conversapp_request("put", f"/core/v1/contact/{contact_id}", json=body)
        if resp.status_code in (200, 201):
            print(f"[CONVERSAPP] Contato {contact_id} atualizado: {list(updates.keys())}")
            return True
        else:
            print(f"[CONVERSAPP] Erro ao atualizar contato: {resp.status_code} {resp.text[:200]}")
            return False
    except Exception as e:
        print(f"[CONVERSAPP] Erro ao atualizar: {e}")
        return False


def conversapp_get_custom_fields():
    """List all custom field definitions to get their keys."""
    try:
        resp = conversapp_request("get", "/core/v1/contact/custom-field")
        if resp.status_code == 200:
            return resp.json()
        return []
    except Exception:
        return []


def conversapp_get_tags():
    """List all available tags."""
    try:
        resp = conversapp_request("get", "/core/v1/tag")
        if resp.status_code == 200:
            return resp.json()
        return []
    except Exception:
        return []


def _conversapp_auto_fill(phone, session):
    """Auto-fill ConversApp contact fields after finding process/INSS data."""
    try:
        contact = conversapp_get_contact(phone)
        if not contact:
            print(f"[CONVERSAPP] Auto-fill: contato não encontrado para {phone}")
            return

        contact_id = contact.get("id")
        if not contact_id:
            return

        custom_fields = {}
        existing_fields = contact.get("customFields") or {}

        proc = session.get("processo")
        gmail = session.get("gmail_resultado")

        # ConversApp field keys (from /core/v1/contact/custom-field)
        KEY_PROCESSO = "n-de-processo"
        KEY_TIPO = "tipo-inss-ou-judicia"
        KEY_STATUS = "status-"
        KEY_CPF = "086-344-453-99"

        if proc:
            numero = proc.get("numero_processo", "")
            if numero and not existing_fields.get(KEY_PROCESSO):
                custom_fields[KEY_PROCESSO] = numero
            if not existing_fields.get(KEY_TIPO):
                custom_fields[KEY_TIPO] = "Judicial"
            status = proc.get("inbox_atual", "")
            if status:
                custom_fields[KEY_STATUS] = status

        elif gmail:
            protocolo = gmail.get("protocolo", "")
            if protocolo and not existing_fields.get(KEY_PROCESSO):
                custom_fields[KEY_PROCESSO] = protocolo
            if not existing_fields.get(KEY_TIPO):
                custom_fields[KEY_TIPO] = "INSS"
            status_inss = gmail.get("status_inss", "")
            if status_inss:
                custom_fields[KEY_STATUS] = status_inss

        if custom_fields:
            conversapp_update_contact(contact_id, {"customFields": custom_fields})
            print(f"[CONVERSAPP] Auto-fill para {phone}: {custom_fields}")
    except Exception as e:
        print(f"[CONVERSAPP] Erro auto-fill: {e}")


def _conversapp_load_context(phone):
    """Load contact data from ConversApp to pre-fill session context.
    Returns dict with known data or None."""
    try:
        contact = conversapp_get_contact(phone)
        if not contact:
            return None

        custom_fields = contact.get("customFields") or {}
        nome = contact.get("name") or ""

        # Check if we have useful data (keys from ConversApp custom fields)
        processo_num = custom_fields.get("n-de-processo")
        tipo = custom_fields.get("tipo-inss-ou-judicia")
        cpf = custom_fields.get("086-344-453-99")

        if not processo_num and not cpf:
            return None  # No useful data to pre-fill

        context = {
            "nome": nome,
            "processo_num": processo_num,
            "tipo": tipo,
            "cpf": cpf,
            "contact_id": contact.get("id"),
        }

        # Get tags
        tags = contact.get("tags") or []
        tag_names = [t.get("name", "") for t in tags] if isinstance(tags, list) else []
        context["tags"] = tag_names

        print(f"[CONVERSAPP] Contexto carregado para {phone}: nome={nome}, processo={processo_num}, tipo={tipo}")
        return context
    except Exception as e:
        print(f"[CONVERSAPP] Erro ao carregar contexto: {e}")
        return None


def whatsapp_send_message(phone, text, session_id=None):
    """Send a WhatsApp message via Helena CRM API."""
    if not CONVERSAPP_API_TOKEN:
        print(f"[WHATSAPP] (sem token) Msg para {phone}: {text[:100]}...")
        return False
    try:
        # If we have a session, reply in it
        if session_id:
            resp = conversapp_request("post", f"/chat/v1/session/{session_id}/message",
                                  json={"text": text})
        else:
            # Send new message (needs channel + phone)
            payload = {
                "channelId": CONVERSAPP_CHANNEL_ID,
                "number": phone,
                "text": text,
            }
            resp = conversapp_request("post", "/chat/v1/message/send", json=payload)

        if resp.status_code in (200, 201):
            print(f"[WHATSAPP] Enviado para {phone}")
            return True
        else:
            print(f"[WHATSAPP] Erro ao enviar: {resp.status_code} {resp.text[:200]}")
            return False
    except Exception as e:
        print(f"[WHATSAPP] Erro: {e}")
        return False


def whatsapp_buscar_processo(nome=None, cpf=None):
    """Search for a process by client name or CPF in the cached process list."""
    # Load process cache
    todos = _load_json_file(PROCESSES_CACHE_FILE)
    if not todos:
        # Try fetching first page from API
        try:
            resp = legalmail_request("get", "/process/all?offset=0&limit=50")
            todos = resp.json() if resp.status_code == 200 and isinstance(resp.json(), list) else []
        except Exception:
            todos = []

    encontrados = []
    if cpf:
        cpf_clean = re.sub(r'\D', '', cpf)
        for p in todos:
            polo_doc = re.sub(r'\D', '', p.get("poloativo_cpf", "") or "")
            if polo_doc and polo_doc == cpf_clean:
                encontrados.append(p)
    if not encontrados and nome:
        nome_lower = nome.lower().strip()
        palavras = [p for p in nome_lower.split() if len(p) > 2]  # ignore "de", "da", etc.

        # 1) Try full name match (substring in either direction)
        for p in todos:
            polo = (p.get("poloativo_nome") or "").lower()
            if polo and nome_lower and (nome_lower in polo or polo in nome_lower):
                encontrados.append(p)

        # 2) If nothing found, try matching ANY word from the name
        if not encontrados and palavras:
            for p in todos:
                polo = (p.get("poloativo_nome") or "").lower()
                if polo and any(palavra in polo for palavra in palavras):
                    encontrados.append(p)

        # If too many, try to narrow down by matching MORE words
        if len(encontrados) > 5 and len(palavras) > 1:
            # Score by how many words match
            scored = []
            for p in encontrados:
                polo = (p.get("poloativo_nome") or "").lower()
                score = sum(1 for palavra in palavras if palavra in polo)
                scored.append((score, p))
            scored.sort(key=lambda x: x[0], reverse=True)
            best_score = scored[0][0]
            encontrados = [p for s, p in scored if s == best_score]
        elif len(encontrados) > 5:
            # Single word - try startswith match
            primeiro_nome = palavras[0] if palavras else nome_lower
            strict = [p for p in encontrados
                      if (p.get("poloativo_nome") or "").lower().startswith(primeiro_nome)]
            if strict:
                encontrados = strict

    return encontrados[:5]  # Max 5 results


def whatsapp_get_movimentacoes(idprocesso):
    """Get recent movements for a process."""
    movs = []
    try:
        resp = legalmail_request("get", f"/process/autos?idprocesso={idprocesso}")
        if resp.status_code == 200:
            autos = resp.json()
            if isinstance(autos, list):
                movs = autos[:5]
    except Exception:
        pass
    # Fallback: notifications cache
    if not movs:
        notificacoes = _load_notifications()
        movs = [
            {"titulo": n.get("titulo_movimentacao", ""), "data_movimentacao": n.get("data_movimentacao", "")}
            for n in notificacoes
            if str(n.get("idprocesso")) == str(idprocesso)
        ][:5]
    return movs


# Gmail integration for INSS administrative status (via Gmail API OAuth2)
# Credentials stored in env vars to avoid committing secrets
GMAIL_REFRESH_TOKEN = os.environ.get("GMAIL_REFRESH_TOKEN", "")
GMAIL_CLIENT_ID = os.environ.get("GMAIL_CLIENT_ID", "")
GMAIL_CLIENT_SECRET = os.environ.get("GMAIL_CLIENT_SECRET", "")

def _get_gmail_service():
    """Get authenticated Gmail API service using env var credentials."""
    if not GMAIL_REFRESH_TOKEN or not GMAIL_CLIENT_ID:
        # Fallback: try local token file
        token_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gmail_token.json")
        if os.path.exists(token_file):
            with open(token_file) as f:
                token_data = json.load(f)
            refresh_token = token_data.get('refresh_token')
            client_id = token_data.get('client_id')
            client_secret = token_data.get('client_secret')
        else:
            print("[GMAIL] Credenciais não configuradas (GMAIL_REFRESH_TOKEN)")
            return None
    else:
        refresh_token = GMAIL_REFRESH_TOKEN
        client_id = GMAIL_CLIENT_ID
        client_secret = GMAIL_CLIENT_SECRET

    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    creds = Credentials(
        token=None,
        refresh_token=refresh_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=client_id,
        client_secret=client_secret,
        scopes=["https://www.googleapis.com/auth/gmail.readonly"]
    )

    # Always refresh to get a valid access token
    creds.refresh(Request())

    return build('gmail', 'v1', credentials=creds)


def whatsapp_buscar_gmail_inss(nome_cliente):
    """Search Gmail for INSS notifications about a client.

    Uses Gmail API to search jrcandamentos@gmail.com for emails
    from noreply@inss.gov.br containing the client's name.
    Returns dict with: nome, protocolo, servico, data, status, corpo
    """
    try:
        service = _get_gmail_service()
        if not service:
            return None

        nome_busca = nome_cliente.strip()

        # Try full name first, then progressively shorter combinations
        # NEVER search single words alone (privacy risk - could match wrong client)
        tentativas = [nome_busca]
        palavras = [p for p in nome_busca.split() if len(p) > 2]
        if palavras and len(palavras) < len(nome_busca.split()):
            tentativas.append(" ".join(palavras))
        # Try first + second name (minimum 2 words for privacy)
        if len(palavras) >= 2:
            tentativas.append(f"{palavras[0]} {palavras[1]}")
            # Try first + last
            if len(palavras) >= 3:
                tentativas.append(f"{palavras[0]} {palavras[-1]}")

        messages = []
        query_used = ""
        for tentativa in tentativas:
            query = f'from:noreply@inss.gov.br "{tentativa}"'
            results = service.users().messages().list(
                userId='me', q=query, maxResults=5
            ).execute()
            messages = results.get('messages', [])
            if messages:
                query_used = tentativa
                print(f"[GMAIL] Encontrado com busca '{tentativa}'")
                break

        if not messages:
            print(f"[GMAIL] Nenhum e-mail do INSS para '{nome_busca}' (tentou: {tentativas})")
            return None

        # Get most recent email
        msg = service.users().messages().get(
            userId='me', id=messages[0]['id'], format='full'
        ).execute()

        # Extract headers
        headers = msg.get('payload', {}).get('headers', [])
        subject = next((h['value'] for h in headers if h['name'] == 'Subject'), '')
        date = next((h['value'] for h in headers if h['name'] == 'Date'), '')

        # Get body text (snippet is good enough for status extraction)
        corpo = msg.get('snippet', '')

        # Also try to get full body
        payload = msg.get('payload', {})
        body_data = payload.get('body', {}).get('data', '')
        if not body_data and payload.get('parts'):
            for part in payload['parts']:
                if part.get('mimeType') == 'text/plain':
                    body_data = part.get('body', {}).get('data', '')
                    break
                elif part.get('mimeType') == 'text/html' and not body_data:
                    body_data = part.get('body', {}).get('data', '')

        if body_data:
            import base64
            corpo = base64.urlsafe_b64decode(body_data).decode('utf-8', errors='replace')

        info = {
            "corpo": corpo[:2000],
            "data_email": date,
            "assunto": subject,
            "nome_cliente": nome_cliente,
        }

        corpo_lower = corpo.lower()

        # Extract protocolo from subject or body
        proto_match = re.search(r'(?:requerimento|protocolo)[:\s]*(\d[\d./-]+)', subject + ' ' + corpo, re.IGNORECASE)
        if proto_match:
            info["protocolo"] = proto_match.group(1).strip()

        # Extract servico (clean HTML first to avoid matching wrong field)
        corpo_texto = re.sub(r'<[^>]+>', ' ', corpo)
        corpo_texto = re.sub(r'\s+', ' ', corpo_texto)
        serv_match = re.search(r'servi[çc]o\s*:\s*(.+?)(?:\s{2,}|Data|Unidade|Status|$)', corpo_texto, re.IGNORECASE)
        if serv_match:
            info["servico"] = serv_match.group(1).strip()

        # Extract status from subject or body
        for status_kw in ["exigência", "exigencia", "indeferido", "deferido",
                          "em análise", "em analise", "cancelado",
                          "concluído", "concluido", "cumprida"]:
            if status_kw in (subject + ' ' + corpo).lower():
                info["status_inss"] = status_kw.replace("exigencia", "Exigência").replace("em analise", "Em análise").replace("concluido", "Concluído").capitalize()
                break

        print(f"[GMAIL] Encontrado: {subject[:80]}")
        return info

    except Exception as e:
        print(f"[GMAIL] Erro: {e}")
        traceback.print_exc()
        return None



# Traduções de movimentações judiciais para linguagem simples
_TRADUCOES_MOV = {
    "petição inicial distribuída": "Seu processo foi aberto e registrado no sistema da Justiça Federal.",
    "designada audiência": "Foi marcada uma audiência. Em breve o escritório vai te contatar com mais detalhes.",
    "perícia médica agendada": "A Justiça marcou uma perícia médica. O escritório vai te avisar a data assim que confirmar.",
    "sentença prolatada": "O juiz analisou o seu caso e proferiu a decisão. A sentença é o momento em que o juiz dá o resultado final do processo na primeira instância, dizendo se o pedido foi aceito ou não.",
    "sentença": "O juiz analisou o seu caso e proferiu a decisão. A sentença é o momento em que o juiz dá o resultado final do processo na primeira instância, dizendo se o pedido foi aceito ou não.",
    "trânsito em julgado": "O processo foi finalizado com decisão definitiva. O escritório vai entrar em contato para explicar os próximos passos.",
    "apelação interposta": "Seu processo está na fase de recurso e aguarda julgamento pelos desembargadores. Essa etapa pode levar alguns meses, mas o escritório está acompanhando.",
    "remetido ao tribunal": "Seu processo está na fase de recurso e aguarda julgamento pelos desembargadores. Essa etapa pode levar alguns meses, mas o escritório está acompanhando.",
    "concluso ao relator": "Seu processo está na fase de recurso e aguarda julgamento pelos desembargadores. Essa etapa pode levar alguns meses, mas o escritório está acompanhando.",
    "incluído em pauta": "Seu processo está na fase de recurso e aguarda julgamento pelos desembargadores. Essa etapa pode levar alguns meses, mas o escritório está acompanhando.",
    "acórdão publicado": "O acórdão é a decisão final do recurso na segunda instância. Isso significa que um grupo de juízes superiores, chamados desembargadores, se reuniu e votou sobre o seu caso. É como uma segunda análise da Justiça.",
    "decisão interlocutória": "O juiz tomou uma decisão parcial no seu processo. Isso não é o resultado final, é uma decisão sobre algum ponto específico que surgiu durante o andamento. O processo continua seguindo normalmente.",
    "julgamento antecipado": "O juiz decidiu analisar o seu caso sem precisar de audiência ou perícia, pois entendeu que já tinha informações suficientes para dar a decisão.",
    "resolução parcial": "O juiz resolveu uma parte do seu processo. Seus documentos, provas, tudo foi analisado, mas ainda restam outros pontos a serem decididos. O andamento continua normalmente para as questões restantes.",
}

# Traduções de status INSS
_TRADUCOES_INSS = {
    "exigência": "O INSS pediu alguns documentos adicionais. O escritório já está cuidando disso.",
    "em análise": "O INSS está analisando o pedido. Assim que tiver alguma novidade, te informamos.",
    "deferido": "O benefício foi aprovado pelo INSS! O escritório vai te contatar com os próximos passos.",
    "indeferido": "Infelizmente o INSS negou o pedido, mas não se preocupe. O escritório vai analisar e te orientar sobre os próximos passos.",
    "concluído": "O processo no INSS foi concluído. O escritório vai te explicar os detalhes.",
    "cumprida": "A exigência foi cumprida e o processo segue em andamento no INSS.",
    "cancelado": None,  # Escalar para equipe
}


def _traduzir_movimentacao(titulo):
    """Traduz movimentação judicial para linguagem simples."""
    titulo_lower = titulo.lower()
    for chave, traducao in _TRADUCOES_MOV.items():
        if chave in titulo_lower:
            return traducao
    return titulo  # Sem tradução, retorna original


def _build_resultado_msg(session, processo_info):
    """Build results message: try Claude for natural explanation, code fallback if fails."""

    # Collect raw data for both Claude and fallback
    gmail_info = session.get("gmail_resultado")
    proc = session.get("processo")
    dados_raw = ""
    fallback_msg = ""

    # Gmail/INSS administrative result
    if gmail_info and "ANDAMENTO ADMINISTRATIVO" in processo_info:
        status = (gmail_info.get("status_inss") or "não identificado").lower()
        protocolo = gmail_info.get("protocolo", "não identificado")
        servico = gmail_info.get("servico", "")

        if "cancelado" in status:
            return ("Encontrei informações sobre o benefício, mas esse assunto a equipe do escritório "
                    "consegue te ajudar melhor. Vou avisar eles para entrar em contato com você.")

        traducao = _TRADUCOES_INSS.get(status, f"O status atual no INSS é: {gmail_info.get('status_inss', '-')}.")

        servico_upper = servico.upper()
        if "ASSISTENCIAL" in servico_upper or "BPC" in servico_upper or "DEFICI" in servico_upper:
            servico_friendly = "Benefício Assistencial (BPC/LOAS)"
        elif "APOSENTADORIA" in servico_upper:
            servico_friendly = "Aposentadoria"
        elif "AUXÍLIO" in servico_upper or "AUXILIO" in servico_upper:
            servico_friendly = "Auxílio por Incapacidade"
        else:
            servico_friendly = servico.title() if servico else ""

        nome = gmail_info.get("nome_cliente", "")
        dados_raw = f"""Dados do INSS para o cliente:
- Nome: {nome}
- Protocolo: {protocolo}
- Tipo: {servico_friendly or servico}
- Status: {gmail_info.get('status_inss', '-')}
- Tradução do status: {traducao}"""

        # Fallback message (code-built)
        fallback_msg = f"Encontrei as informações do {nome}!\n\n"
        fallback_msg += f"📋 *Andamento administrativo - INSS*\n\n"
        fallback_msg += f"📌 *Protocolo:* {protocolo}\n"
        if servico_friendly:
            fallback_msg += f"📎 *Tipo:* {servico_friendly}\n"
        fallback_msg += f"📊 *Status:* {gmail_info.get('status_inss', '-')}\n\n"
        fallback_msg += f"{traducao}\n\n"
        fallback_msg += "Ficou alguma dúvida ou posso te ajudar com mais alguma coisa?"

    # Judicial process result
    elif proc and "PROCESSO ENCONTRADO" in processo_info:
        numero = proc.get("numero_processo", "")
        tribunal = proc.get("tribunal", "")
        cliente = proc.get("poloativo_nome", "")

        # Use cached movimentações if available, otherwise fetch
        movs = session.get("_movimentacoes") or whatsapp_get_movimentacoes(proc.get("idprocessos"))
        movs_texto = ""
        for m in movs[:5]:
            data_m = (m.get("data_movimentacao") or "")[:10]
            titulo = m.get("titulo") or m.get("titulo_movimentacao", "")
            movs_texto += f"- {data_m}: {titulo}\n"

        dados_raw = f"""Dados do processo judicial:
- Número: {numero}
- Tribunal: {tribunal}
- Cliente: {cliente}
- Classe: {proc.get('nome_classe') or proc.get('abreviatura_classe', '')}
- Juízo: {proc.get('juizo', '')}
- Status: {proc.get('inbox_atual', 'Em andamento')}
Últimas movimentações:
{movs_texto or 'Nenhuma movimentação recente.'}"""

        ultima_data = ""
        traducao_mov = ""
        if movs:
            ultima_data = (movs[0].get("data_movimentacao") or "")[:10]
            titulo = movs[0].get("titulo") or movs[0].get("titulo_movimentacao", "")
            traducao_mov = _traduzir_movimentacao(titulo)

        fallback_msg = "Encontrei o seu processo!\n\n"
        fallback_msg += f"📋 *Andamento do processo*\n\n"
        fallback_msg += f"⚖️ *Processo nº:* {numero}\n"
        fallback_msg += f"🏛️ *Tribunal:* {tribunal}\n"
        if ultima_data:
            fallback_msg += f"📅 *Última movimentação:* {ultima_data}\n"
        if traducao_mov:
            fallback_msg += f"✅ *O que aconteceu:* {traducao_mov}\n"
        fallback_msg += "\nFicou alguma dúvida ou posso te ajudar com mais alguma coisa?"

    if not dados_raw:
        return "Não consegui localizar as informações no momento. Vou encaminhar para a equipe do escritório verificar. Eles vão entrar em contato com você."

    # Try AI with fallback (Anthropic -> GLM -> Gemini)
    try:
        resposta = ai_chat(
            messages=[{"role": "user", "content": f"Formate estes dados para o cliente:\n\n{dados_raw}"}],
            system="""Você é Ana, do pós-venda da JRC Advocacia. Formate uma mensagem de WhatsApp com os dados abaixo.

REGRAS:
- NÃO comece com saudação (sem "Oi", "Olá", nome do cliente, acenos). Já foi enviada uma mensagem antes dizendo "vou consultar". Comece DIRETO com os dados.
- Use emojis e negrito (*texto*) APENAS no resumo dos dados (cabeçalho)
- Depois do resumo, explique em linguagem simples o que está acontecendo e os próximos passos
- Tom acolhedor e direto, sem juridiquês
- Traduza TODAS as movimentações jurídicas para linguagem que qualquer pessoa entenda
- Termine com "Ficou alguma dúvida ou posso te ajudar com mais alguma coisa?"
- NÃO invente informações que não estão nos dados
- Máximo 10 linhas""",
            max_tokens=500,
        )
        resposta = (resposta or "").strip()
        if resposta and len(resposta) > 20:
            # Strip unwanted greetings that Claude may add despite instructions
            # Only match optional capitalized name (not any word like "Encontrei")
            resposta = re.sub(r'^(?:Oi|Olá|Ei|Hey)[,!]?\s*(?:[A-ZÀ-Ú][a-zà-ú]+)?[,!]?\s*[👋🤗😊]?\s*\n*', '', resposta, flags=re.UNICODE).strip()
            if resposta and len(resposta) > 20:
                return resposta
    except Exception as e:
        print(f"[BOT] Claude falhou ao formatar resultado, usando fallback: {e}")

    return fallback_msg


def _get_saudacao():
    """Return greeting based on current Brazil time."""
    from datetime import datetime, timedelta, timezone
    br_tz = timezone(timedelta(hours=-3))
    hora = datetime.now(br_tz).hour
    if 6 <= hora < 12:
        return "Bom dia"
    elif 12 <= hora < 18:
        return "Boa tarde"
    else:
        return "Boa noite"


def _is_horario_comercial():
    """Check if current time is within business hours (Brazil)."""
    from datetime import datetime, timedelta, timezone
    br_tz = timezone(timedelta(hours=-3))
    now = datetime.now(br_tz)
    # Monday=0, Sunday=6
    if now.weekday() >= 5:  # Weekend
        return False
    return BOT_HORA_INICIO <= now.hour < BOT_HORA_FIM


def whatsapp_processar_mensagem(phone, message):
    """Process incoming WhatsApp message using Claude AI for natural conversation."""
    msg = message.strip()
    msg_lower = msg.lower()

    # Check if Ana is paused for this phone (human takeover)
    phone_clean = re.sub(r'[^\d]', '', str(phone))
    if phone_clean in _paused_phones:
        # Auto-expire pause after 24 hours
        try:
            paused_at = _paused_phones[phone_clean]
            from datetime import datetime as _dt_check, timedelta as _td_check
            if isinstance(paused_at, str):
                paused_time = _dt_check.fromisoformat(paused_at.replace('Z', '+00:00')) if 'T' in paused_at else _dt_check.strptime(paused_at, '%Y-%m-%d %H:%M:%S.%f')
                if (_dt_check.now() - paused_time) > _td_check(hours=24):
                    _paused_phones.pop(phone_clean, None)
                    print(f"[BOT] Pausa expirada (24h) para {phone_clean} - Ana retomando")
                else:
                    print(f"[BOT] Ana pausada para {phone_clean} - atendimento humano")
                    return None
            else:
                print(f"[BOT] Ana pausada para {phone_clean} - atendimento humano")
                return None
        except Exception:
            print(f"[BOT] Ana pausada para {phone_clean} - atendimento humano")
            return None

    # Check business hours
    if not _is_horario_comercial():
        # Outside hours: send auto-reply only once per session
        session = _whatsapp_sessions.get(phone, {})
        if session.get("fora_horario_enviado"):
            return None  # Don't spam the same message
        session["fora_horario_enviado"] = True
        _whatsapp_sessions[phone] = session
        _save_sessions()
        return BOT_MSG_FORA_HORARIO

    # Get or create session
    session = _whatsapp_sessions.get(phone, {"historico": [], "processo": None})
    # Reset "fora_horario" flag when in business hours
    session.pop("fora_horario_enviado", None)

    # On first message of a new session, load context from ConversApp
    is_new_session = len(session.get("historico", [])) == 0
    if is_new_session and not session.get("_conversapp_loaded"):
        conversapp_ctx = _conversapp_load_context(phone)
        if conversapp_ctx:
            session["_conversapp_context"] = conversapp_ctx
            # If we have a process number, try to pre-load it
            proc_num = conversapp_ctx.get("processo_num")
            tipo = conversapp_ctx.get("tipo")
            cpf = conversapp_ctx.get("cpf")
            if proc_num and tipo == "Judicial":
                # Search by process number in LegalMail cache
                processos = whatsapp_buscar_processo(nome=conversapp_ctx.get("nome"))
                if processos:
                    # Find the one matching the stored number
                    for p in processos:
                        if p.get("numero_processo") == proc_num:
                            session["processo"] = p
                            break
                    if not session.get("processo"):
                        session["processo"] = processos[0]
            elif tipo == "INSS" and conversapp_ctx.get("nome"):
                session["contexto_inss"] = True
        session["_conversapp_loaded"] = True

    # Keep conversation history (last 10 messages for context)
    historico = session.get("historico", [])
    historico.append({"role": "user", "content": msg})
    if len(historico) > 20:
        historico = historico[-20:]
    session["historico"] = historico

    # Detect if client wants to check a process
    processo_info = ""
    processo = session.get("processo")

    # Check if message looks like a CPF
    digits = re.sub(r'\D', '', msg)
    is_cpf = len(digits) >= 9 and len(digits) <= 11

    # Check if client is providing identification (name or CPF)
    wants_processo = any(kw in msg_lower for kw in [
        "processo", "andamento", "ação", "acao", "como está", "como esta",
        "novidade", "atualização", "atualizacao", "movimentação", "movimentacao",
        "sentença", "sentenca", "audiência", "audiencia", "perícia", "pericia",
        "meu caso", "meu benefício", "meu beneficio"
    ])

    # If client previously asked about processo and is now giving name/CPF
    aguardando_id = session.get("aguardando_identificacao", False)

    # Check if message looks like a person's name (not a keyword/answer)
    palavras_nao_nome = {
        "inss", "sim", "não", "nao", "ok", "benefício", "beneficio",
        "trabalhista", "processo", "andamento", "loas", "bpc", "aposentadoria",
        "filho", "filha", "mãe", "mae", "pai", "marido", "esposa", "meu", "minha",
        "oi", "olá", "ola", "bom", "dia", "boa", "tarde", "noite", "obrigado", "obrigada",
        "é", "e", "do", "da", "de", "o", "a", "um", "uma", "pra", "para", "no", "na",
        "como", "está", "esta", "tá", "ta", "quero", "preciso", "saber", "ver",
        "judicial", "administrativo", "outro", "outra", "tipo", "qual",
        "tudo", "bem", "isso", "esse", "essa", "aqui", "ali", "lá", "la",
        "por", "favor", "por favor", "muito", "só", "so", "ainda",
        "quando", "onde", "porque", "mas", "que", "se", "com", "sem", "até", "ate",
        "?", "!", ".", "nada", "algo", "alguma", "coisa", "pode", "ajudar",
    }
    msg_parece_nome = not is_cpf and msg_lower not in palavras_nao_nome and len(msg.strip()) > 2
    # Also check: if ALL words are non-name keywords, it's not a name
    if msg_parece_nome:
        palavras_msg = set(msg_lower.split())
        if palavras_msg and palavras_msg.issubset(palavras_nao_nome):
            msg_parece_nome = False

    print(f"[BOT] aguardando_id={aguardando_id}, is_cpf={is_cpf}, msg_parece_nome={msg_parece_nome}, msg='{msg[:50]}'")

    # Detect INSS context from conversation history
    contexto_inss = session.get("contexto_inss", False)
    if not contexto_inss:
        # Only check USER messages (not bot messages that mention INSS in triaging)
        user_texto = " ".join(h.get("content", "") for h in historico if h.get("role") == "user").lower()
        if any(kw in user_texto for kw in ["inss", "benefício", "beneficio", "bpc", "loas"]):
            contexto_inss = True
            session["contexto_inss"] = True

    if aguardando_id and (is_cpf or msg_parece_nome):
        print(f"[BOT] Buscando processo para: '{msg}' (contexto_inss={contexto_inss})")

        gmail_info = None
        proc_encontrado = None

        # Search Gmail first if INSS context (more likely to be relevant)
        if contexto_inss and not is_cpf:
            print(f"[BOT] Contexto INSS, buscando Gmail primeiro...")
            gmail_info = whatsapp_buscar_gmail_inss(msg)
            print(f"[BOT] Gmail resultado: {bool(gmail_info)}")

        # If Gmail didn't find, try LegalMail
        if not gmail_info:
            print(f"[BOT] Buscando no LegalMail...")
            if is_cpf:
                processos = whatsapp_buscar_processo(cpf=msg)
            else:
                processos = whatsapp_buscar_processo(nome=msg)
            print(f"[BOT] LegalMail encontrou: {len(processos)} processos")

            if processos:
                nomes_unicos = set((p.get("poloativo_nome") or "").strip().upper() for p in processos)
                if len(nomes_unicos) == 1:
                    proc_encontrado = max(processos, key=lambda p: p.get("data_cadastro") or p.get("data_distribuicao") or "")
                elif len(processos) == 1:
                    proc_encontrado = processos[0]
                else:
                    nome_busca = msg.lower().strip()
                    for p in processos:
                        polo = (p.get("poloativo_nome") or "").lower()
                        if nome_busca == polo or (nome_busca in polo and len(nome_busca) > len(polo) * 0.6):
                            proc_encontrado = p
                            break

            # If LegalMail didn't find good match, try Gmail as fallback (if not tried yet)
            if not proc_encontrado and not contexto_inss:
                print(f"[BOT] LegalMail sem match, tentando Gmail como fallback...")
                gmail_info = whatsapp_buscar_gmail_inss(msg)
                print(f"[BOT] Gmail fallback resultado: {bool(gmail_info)}")

        # Set processo_info based on what was found (clear search state)
        if gmail_info:
            session["aguardando_identificacao"] = False
            session.pop("contexto_inss", None)
            gmail_resultado_limpo = {k: v for k, v in gmail_info.items() if k != 'corpo'}
            session["gmail_resultado"] = gmail_resultado_limpo
            processo_info = f"""
ANDAMENTO ADMINISTRATIVO ENCONTRADO NO GMAIL (e-mail do INSS):
- Cliente: {gmail_info.get('nome_cliente', '')}
- Protocolo: {gmail_info.get('protocolo', 'não identificado')}
- Serviço: {gmail_info.get('servico', 'não identificado')}
- Status INSS: {gmail_info.get('status_inss', 'não identificado')}
- Data do e-mail: {gmail_info.get('data_email', '')}
"""
        elif proc_encontrado:
            session["processo"] = proc_encontrado
            session["aguardando_identificacao"] = False
            session.pop("contexto_inss", None)
            movs = whatsapp_get_movimentacoes(proc_encontrado.get("idprocessos"))
            session["_movimentacoes"] = movs  # Cache for _build_resultado_msg
            movs_texto = ""
            for m in movs[:5]:
                data_m = (m.get("data_movimentacao") or "")[:10]
                titulo = m.get("titulo") or m.get("titulo_movimentacao", "")
                movs_texto += f"- {data_m}: {titulo}\n"
            processo_info = f"""
PROCESSO ENCONTRADO:
- Número: {proc_encontrado.get('numero_processo', '')}
- Cliente: {proc_encontrado.get('poloativo_nome', '')}
- Tribunal: {proc_encontrado.get('tribunal', '')}
- Classe: {proc_encontrado.get('nome_classe') or proc_encontrado.get('abreviatura_classe', '')}
- Juízo: {proc_encontrado.get('juizo', '')}
- Status: {proc_encontrado.get('inbox_atual', 'Em andamento')}
Últimas movimentações:
{movs_texto or 'Nenhuma movimentação recente.'}
"""
        elif processos:
            # Found processes but couldn't disambiguate - ask for full name
            nomes_lista = ", ".join(sorted(set((p.get("poloativo_nome") or "").strip() for p in processos)))
            processo_info = f"\nENCONTREI PROCESSOS MAS NÃO CONSEGUI IDENTIFICAR O CLIENTE CORRETO. Nomes encontrados: {nomes_lista}. Peça o nome completo ou sobrenome para identificar corretamente."
            session["aguardando_identificacao"] = True
        else:
            processo_info = "\nNENHUM PROCESSO ENCONTRADO nem no sistema judicial nem nos e-mails do INSS. Peça para tentar novamente com o nome completo como está no processo ou encaminhe para a equipe."
            session["aguardando_identificacao"] = False
            session.pop("contexto_inss", None)

    # If already has a process in session, include it
    elif processo and not processo_info:
        movs = whatsapp_get_movimentacoes(processo.get("idprocessos"))
        movs_texto = ""
        for m in movs[:5]:
            data_m = (m.get("data_movimentacao") or "")[:10]
            titulo = m.get("titulo") or m.get("titulo_movimentacao", "")
            movs_texto += f"- {data_m}: {titulo}\n"
        processo_info = f"""
PROCESSO DO CLIENTE (já identificado):
- Número: {processo.get('numero_processo', '')}
- Cliente: {processo.get('poloativo_nome', '')}
- Tribunal: {processo.get('tribunal', '')}
- Status: {processo.get('inbox_atual', 'Em andamento')}
Últimas movimentações:
{movs_texto or 'Nenhuma movimentação recente.'}
"""

    # If client wants to check process but hasn't identified yet
    # First check: do we already know this client from ConversApp?
    conversapp_ctx = session.get("_conversapp_context")
    if wants_processo and not processo and not processo_info and conversapp_ctx:
        ctx_nome = conversapp_ctx.get("nome", "")
        ctx_tipo = conversapp_ctx.get("tipo")
        ctx_proc_num = conversapp_ctx.get("processo_num")
        print(f"[BOT] Cliente conhecido do ConversApp: {ctx_nome}, tipo={ctx_tipo}, proc={ctx_proc_num}")

        if ctx_nome and len(ctx_nome) > 3:
            session["aguardando_identificacao"] = True
            gmail_info_ctx = None
            proc_ctx = None

            if ctx_tipo == "INSS":
                gmail_info_ctx = whatsapp_buscar_gmail_inss(ctx_nome)

            if not gmail_info_ctx:
                processos_ctx = whatsapp_buscar_processo(nome=ctx_nome)
                if processos_ctx:
                    # If we have the stored process number, find exact match
                    if ctx_proc_num:
                        for p in processos_ctx:
                            if p.get("numero_processo") == ctx_proc_num:
                                proc_ctx = p
                                break
                    if not proc_ctx:
                        nomes_u = set((p.get("poloativo_nome") or "").strip().upper() for p in processos_ctx)
                        if len(nomes_u) == 1:
                            proc_ctx = max(processos_ctx, key=lambda p: p.get("data_cadastro") or p.get("data_distribuicao") or "")
                        elif len(processos_ctx) == 1:
                            proc_ctx = processos_ctx[0]

                if not proc_ctx and ctx_tipo != "INSS":
                    gmail_info_ctx = whatsapp_buscar_gmail_inss(ctx_nome)

            if gmail_info_ctx:
                session["aguardando_identificacao"] = False
                session.pop("contexto_inss", None)
                gmail_resultado_limpo = {k: v for k, v in gmail_info_ctx.items() if k != 'corpo'}
                session["gmail_resultado"] = gmail_resultado_limpo
                processo_info = f"\nANDAMENTO ADMINISTRATIVO ENCONTRADO NO GMAIL (e-mail do INSS):\n- Cliente: {gmail_info_ctx.get('nome_cliente', '')}\n- Protocolo: {gmail_info_ctx.get('protocolo', 'não identificado')}\n- Serviço: {gmail_info_ctx.get('servico', 'não identificado')}\n- Status INSS: {gmail_info_ctx.get('status_inss', 'não identificado')}\n- Data do e-mail: {gmail_info_ctx.get('data_email', '')}"
            elif proc_ctx:
                session["processo"] = proc_ctx
                session["aguardando_identificacao"] = False
                session.pop("contexto_inss", None)
                movs = whatsapp_get_movimentacoes(proc_ctx.get("idprocessos"))
                session["_movimentacoes"] = movs
                movs_texto = ""
                for m in movs[:5]:
                    data_m = (m.get("data_movimentacao") or "")[:10]
                    titulo = m.get("titulo") or m.get("titulo_movimentacao", "")
                    movs_texto += f"- {data_m}: {titulo}\n"
                processo_info = f"\nPROCESSO ENCONTRADO:\n- Número: {proc_ctx.get('numero_processo', '')}\n- Cliente: {proc_ctx.get('poloativo_nome', '')}\n- Tribunal: {proc_ctx.get('tribunal', '')}\n- Status: {proc_ctx.get('inbox_atual', 'Em andamento')}\nÚltimas movimentações:\n{movs_texto or 'Nenhuma movimentação recente.'}"

    if wants_processo and not processo and not processo_info:
        # Try to extract a name embedded in the message (e.g., "processo do João Silva")
        nome_embutido = None
        nome_match = re.search(r'(?:processo|benefício|beneficio|andamento)\s+(?:do|da|de|del)\s+(.+)', msg, re.IGNORECASE)
        if nome_match:
            candidato = nome_match.group(1).strip().rstrip('?!.')
            # Check it's not just keywords
            palavras_cand = set(candidato.lower().split())
            if not palavras_cand.issubset(palavras_nao_nome) and len(candidato) > 3:
                nome_embutido = candidato

        if nome_embutido:
            # Name found in message - search directly
            print(f"[BOT] Nome embutido detectado: '{nome_embutido}'")
            session["aguardando_identificacao"] = True
            # Rerun search logic with extracted name
            user_texto = " ".join(h.get("content", "") for h in historico if h.get("role") == "user").lower()
            ctx_inss = any(kw in user_texto for kw in ["inss", "benefício", "beneficio", "bpc", "loas"])

            gmail_info_emb = None
            proc_emb = None

            if ctx_inss:
                gmail_info_emb = whatsapp_buscar_gmail_inss(nome_embutido)

            if not gmail_info_emb:
                processos_emb = whatsapp_buscar_processo(nome=nome_embutido)
                if processos_emb:
                    nomes_u = set((p.get("poloativo_nome") or "").strip().upper() for p in processos_emb)
                    if len(nomes_u) == 1:
                        proc_emb = max(processos_emb, key=lambda p: p.get("data_cadastro") or p.get("data_distribuicao") or "")
                    elif len(processos_emb) == 1:
                        proc_emb = processos_emb[0]

                if not proc_emb and not ctx_inss:
                    gmail_info_emb = whatsapp_buscar_gmail_inss(nome_embutido)

            if gmail_info_emb:
                session["aguardando_identificacao"] = False
                session.pop("contexto_inss", None)
                gmail_resultado_limpo = {k: v for k, v in gmail_info_emb.items() if k != 'corpo'}
                session["gmail_resultado"] = gmail_resultado_limpo
                processo_info = f"\nANDAMENTO ADMINISTRATIVO ENCONTRADO NO GMAIL (e-mail do INSS):\n- Cliente: {gmail_info_emb.get('nome_cliente', '')}\n- Protocolo: {gmail_info_emb.get('protocolo', 'não identificado')}\n- Serviço: {gmail_info_emb.get('servico', 'não identificado')}\n- Status INSS: {gmail_info_emb.get('status_inss', 'não identificado')}\n- Data do e-mail: {gmail_info_emb.get('data_email', '')}"
            elif proc_emb:
                session["processo"] = proc_emb
                session["aguardando_identificacao"] = False
                session.pop("contexto_inss", None)
                movs = whatsapp_get_movimentacoes(proc_emb.get("idprocessos"))
                session["_movimentacoes"] = movs
                movs_texto = ""
                for m in movs[:5]:
                    data_m = (m.get("data_movimentacao") or "")[:10]
                    titulo = m.get("titulo") or m.get("titulo_movimentacao", "")
                    movs_texto += f"- {data_m}: {titulo}\n"
                processo_info = f"\nPROCESSO ENCONTRADO:\n- Número: {proc_emb.get('numero_processo', '')}\n- Cliente: {proc_emb.get('poloativo_nome', '')}\n- Tribunal: {proc_emb.get('tribunal', '')}\n- Status: {proc_emb.get('inbox_atual', 'Em andamento')}\nÚltimas movimentações:\n{movs_texto or 'Nenhuma movimentação recente.'}"
            else:
                session["aguardando_identificacao"] = True
                processo_info = "\nO CLIENTE QUER CONSULTAR O PROCESSO. Se ele já indicou que é benefício, INSS, do filho, etc, NÃO pergunte novamente - vá direto pedir o nome completo. Só faça a triagem se realmente não ficou claro."
        else:
            session["aguardando_identificacao"] = True
            processo_info = "\nO CLIENTE QUER CONSULTAR O PROCESSO. Se ele já indicou que é benefício, INSS, do filho, etc, NÃO pergunte novamente - vá direto pedir o nome completo. Só faça a triagem se realmente não ficou claro."
    # If aguardando_id but message wasn't a name (was a keyword like "inss", "benefício")
    elif aguardando_id and not processo_info and not msg_parece_nome and not is_cpf:
        session["aguardando_identificacao"] = True
        processo_info = "\nO CLIENTE ESTÁ EM TRIAGEM. Ele respondeu algo sobre o tipo de processo/benefício. Agora peça o nome completo de quem tem o processo, por gentileza."

    # If we found process/gmail data, build results message directly (no AI needed)
    dados_encontrados = "PROCESSO ENCONTRADO" in processo_info or "ANDAMENTO ADMINISTRATIVO" in processo_info

    if dados_encontrados:
        msg_consulta = "Obrigada! Vou consultar agora o andamento. Um momento."
        historico.append({"role": "assistant", "content": msg_consulta})

        # Build results message directly in code (100% reliable, no AI call)
        msg_resultado = _build_resultado_msg(session, processo_info)
        historico.append({"role": "assistant", "content": msg_resultado})
        session["historico"] = historico
        _whatsapp_sessions[phone] = session
        _save_sessions()

        # Auto-fill ConversApp contact fields in background
        threading.Thread(target=_conversapp_auto_fill, args=(phone, session), daemon=True).start()

        return [msg_consulta, msg_resultado]

    # No data found - use Claude for conversation (triagem, saudação, etc.)
    saudacao = _get_saudacao()
    system_msg = f"""{BOT_SYSTEM_PROMPT}

HORÁRIO ATUAL: {saudacao} (usar esta saudação se for a primeira mensagem)
PRIMEIRA MENSAGEM DA CONVERSA: {"Sim" if len(historico) <= 1 else "Não"}
{processo_info}
{"O CLIENTE QUER CONSULTAR O PROCESSO. Se ele já indicou que é benefício, INSS, do filho, etc, NÃO pergunte novamente - vá direto pedir o nome completo. Só faça a triagem se realmente não ficou claro. NUNCA diga 'vou consultar' ou 'um momento' - você só pode dizer isso quando os dados já foram encontrados (isso é feito automaticamente pelo sistema)." if session.get("aguardando_identificacao") else ""}
"""

    # Build messages for Claude (with history for context)
    messages = []
    for h in historico[-10:]:
        messages.append({"role": h["role"], "content": h["content"]})

    try:
        resposta = ai_chat(
            messages=messages,
            system=system_msg,
            max_tokens=800,
        )
        if not resposta:
            resposta = "Desculpe, estou com uma instabilidade momentânea. Vou encaminhar para a equipe. Um momento!"
        resposta = resposta.strip()

        # Save assistant response in history (clean markers)
        resposta_limpa = resposta.replace("[ENCERRAR_SESSAO]", "").replace("[TRANSFERIR_MICHELLE]", "").strip()
        historico.append({"role": "assistant", "content": resposta_limpa})

        # Auto-detect if bot is asking for name -> set aguardando_identificacao
        resposta_lower = resposta.lower()
        if any(kw in resposta_lower for kw in ["nome completo", "nome de quem", "me informar o nome", "qual o nome", "qual é o nome", "pode me passar o nome", "me dizer o nome"]):
            session["aguardando_identificacao"] = True
            print(f"[BOT] Auto-detectado: bot pediu nome, setando aguardando_identificacao=True")

        # Auto-detect date promises and pending docs from client message for follow-up
        try:
            data_detectada = _followup_detect_date(msg)
            docs_detectados = _followup_detect_docs(msg)
            # Also check Ana's response for doc requests
            docs_ana = _followup_detect_docs(resposta_limpa)
            todos_docs = list(set(docs_detectados + docs_ana))

            # Keywords that indicate client is promising to send something
            promessa_envio = any(kw in msg_lower for kw in [
                "vou enviar", "vou mandar", "mando", "envio", "te mando",
                "vou pegar", "vou buscar", "vou conseguir", "vou providenciar",
            ])

            nome_cliente = ""
            proc = session.get("processo")
            if proc:
                nome_cliente = proc.get("poloativo_nome", "")
            elif session.get("_conversapp_context"):
                nome_cliente = session["_conversapp_context"].get("nome", "")

            sid_followup = session.get("conversapp_session_id")

            if data_detectada and (promessa_envio or todos_docs):
                # Client promised a specific date - reset attempts (client engaged)
                contexto_conv = " | ".join(h.get("content", "")[:100] for h in historico[-4:])
                _followup_add(phone, nome_cliente, todos_docs, data_prometida=data_detectada,
                             session_id=sid_followup, contexto=contexto_conv)
                # Reset attempts since client responded
                queue_temp = _followup_load()
                pc = re.sub(r'[^\d]', '', str(phone))
                if pc in queue_temp:
                    queue_temp[pc]["tentativas"] = 0
                    queue_temp[pc]["status"] = "pendente"
                    _followup_save(queue_temp)
                print(f"[BOT] Follow-up reagendado: {phone} -> data={data_detectada}, docs={todos_docs} (tentativas zeradas)")
            elif promessa_envio and not data_detectada:
                # Client promised but no specific date - reset and follow up in 3 days
                contexto_conv = " | ".join(h.get("content", "")[:100] for h in historico[-4:])
                _followup_add(phone, nome_cliente, todos_docs, data_prometida=None,
                             session_id=sid_followup, contexto=contexto_conv)
                # Reset attempts since client responded
                queue_temp = _followup_load()
                pc = re.sub(r'[^\d]', '', str(phone))
                if pc in queue_temp:
                    queue_temp[pc]["tentativas"] = 0
                    queue_temp[pc]["status"] = "pendente"
                    _followup_save(queue_temp)
                print(f"[BOT] Follow-up reagendado (sem data): {phone} -> docs={todos_docs} (tentativas zeradas)")
        except Exception as e:
            print(f"[BOT] Erro ao detectar follow-up: {e}")

        session["historico"] = historico
        _whatsapp_sessions[phone] = session
        _save_sessions()
        return resposta
    except Exception as e:
        print(f"[WHATSAPP] Erro IA: {e}")
        traceback.print_exc()
        return f"{_get_saudacao()}! Desculpe, estou com uma dificuldade técnica no momento. Por favor, tente novamente em alguns minutos ou entre em contato diretamente com o escritório."


from collections import deque as _deque
_webhook_log = _deque(maxlen=20)  # Thread-safe, auto-trimmed
_bot_debug_log = _deque(maxlen=20)


# ========== FOLLOW-UP ENDPOINTS ==========

# Tags que o sistema monitora para follow-up de documentos
FOLLOWUP_TAGS = {
    "a0f8ef0b-e237-41b1-9ed2-96f39597db1c": "Sem laudo",
    "a3e6bd4b-5786-4e1c-b61a-0dbfcf957f7c": "Falta a atualização do laudo",
    "ccc686c8-efa0-451a-8935-75181c4be14b": "Sem CAD ÚNICO",
    "fd305e71-bed1-4b2e-85b5-dd141b515eda": "Falta cad",
}

# Map tag names to document names for follow-up messages
TAG_TO_DOC = {
    "Sem laudo": "laudo médico",
    "Falta a atualização do laudo": "laudo médico atualizado",
    "Sem CAD ÚNICO": "inscrição no CadÚnico",
    "Falta cad": "inscrição no CadÚnico",
}


def _followup_buscar_contatos_por_tag(tag_id):
    """Fetch all contacts that have a specific tag via POST /core/v1/contact/filter."""
    contatos = []
    page = 1
    try:
        while True:
            resp = conversapp_request("post", "/core/v1/contact/filter", json={
                "tagIds": [tag_id],
                "pageSize": 50,
                "pageNumber": page,
            })
            if resp.status_code != 200:
                print(f"[FOLLOWUP] Erro ao buscar contatos por tag: {resp.status_code}")
                break
            data = resp.json()
            items = data.get("items", [])
            contatos.extend(items)
            if not data.get("hasMorePages", False):
                break
            page += 1
            if page > 20:  # Safety limit
                break
    except Exception as e:
        print(f"[FOLLOWUP] Erro ao buscar contatos: {e}")
    return contatos


def _followup_get_last_session(contact_id):
    """Get the most recent session for a contact."""
    try:
        resp = conversapp_request("get", f"/chat/v1/session?contactId={contact_id}&pageSize=1&orderBy=createdat&orderDirection=Descending")
        if resp.status_code == 200:
            items = resp.json().get("items", [])
            if items:
                return items[0]
    except Exception as e:
        print(f"[FOLLOWUP] Erro ao buscar sessão: {e}")
    return None


@app.route("/api/followup/fila", methods=["GET"])
def followup_fila():
    """View the follow-up queue."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    queue = _followup_load()
    return jsonify({"total": len(queue), "clientes": queue})


@app.route("/api/followup/adicionar", methods=["POST"])
def followup_adicionar():
    """Manually add a client to the follow-up queue."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    data = request.get_json() or {}
    phone = data.get("phone", "")
    nome = data.get("nome", "")
    docs = data.get("docs_pendentes", [])
    data_prometida = data.get("data_prometida")
    session_id = data.get("session_id")
    contexto = data.get("contexto", "")
    if not phone:
        return jsonify({"error": "Informe phone"}), 400
    entry = _followup_add(phone, nome, docs, data_prometida, session_id, contexto)
    return jsonify({"status": "ok", "entry": entry})


@app.route("/api/followup/remover", methods=["POST", "DELETE"])
def followup_remover():
    """Remove a client from the follow-up queue."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    phone = request.args.get("phone", "") or (request.get_json() or {}).get("phone", "")
    if not phone:
        return jsonify({"error": "Informe phone"}), 400
    removed = _followup_remove(phone)
    return jsonify({"status": "ok", "removido": bool(removed)})


@app.route("/api/followup/executar", methods=["POST", "GET"])
def followup_executar():
    """Run follow-up: scan ConversApp tags + local queue, send personalized messages."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401

    hoje = _dt_followup.date.today()
    queue = _followup_load()
    resultados = []
    contatos_processados = set()

    # ===== FASE 1: Buscar contatos por tag no ConversApp =====
    for tag_id, tag_nome in FOLLOWUP_TAGS.items():
        doc_nome = TAG_TO_DOC.get(tag_nome, "documentos pendentes")
        contatos = _followup_buscar_contatos_por_tag(tag_id)
        print(f"[FOLLOWUP] Tag '{tag_nome}': {len(contatos)} contatos encontrados")

        for contato in contatos:
            try:
                contact_id = contato.get("id", "")
                nome = contato.get("name") or ""
                phone_numbers = contato.get("phoneNumbers") or contato.get("phonenumber") or ""
                # Extract phone - can be string or list
                phone_raw = ""
                if isinstance(phone_numbers, list) and phone_numbers:
                    phone_raw = phone_numbers[0]
                elif isinstance(phone_numbers, str):
                    phone_raw = phone_numbers
                if not phone_raw:
                    phone_raw = contato.get("phoneNumber") or contato.get("phone") or ""

                phone_clean = re.sub(r'[^\d]', '', str(phone_raw))
                if not phone_clean or len(phone_clean) < 10:
                    continue

                # Skip if already processed in this run
                if phone_clean in contatos_processados:
                    # But collect the doc for this contact
                    if phone_clean in queue:
                        existing_docs = queue[phone_clean].get("docs_pendentes", [])
                        if doc_nome not in existing_docs:
                            existing_docs.append(doc_nome)
                            queue[phone_clean]["docs_pendentes"] = existing_docs
                    continue
                contatos_processados.add(phone_clean)

                # Check/create entry in local queue
                entry = queue.get(phone_clean)
                if not entry:
                    # New contact from tag - add to queue
                    # Get last session for conversation context
                    last_session = _followup_get_last_session(contact_id)
                    sid = last_session.get("sessionId") if last_session else None
                    entry = {
                        "nome": nome,
                        "phone": phone_clean,
                        "contact_id": contact_id,
                        "docs_pendentes": [doc_nome],
                        "data_prometida": None,
                        "session_id": sid,
                        "contexto": "",
                        "criado_em": hoje.isoformat(),
                        "atualizado_em": hoje.isoformat(),
                        "tentativas": 0,
                        "ultimo_followup": None,
                        "status": "pendente",
                        "tags": [tag_nome],
                    }
                    queue[phone_clean] = entry
                else:
                    # Existing entry - add this doc if not already there
                    existing_docs = entry.get("docs_pendentes", [])
                    if doc_nome not in existing_docs:
                        existing_docs.append(doc_nome)
                        entry["docs_pendentes"] = existing_docs
                    existing_tags = entry.get("tags", [])
                    if tag_nome not in existing_tags:
                        existing_tags.append(tag_nome)
                        entry["tags"] = existing_tags
                    queue[phone_clean] = entry

            except Exception as e:
                print(f"[FOLLOWUP] Erro ao processar contato: {e}")
                continue

    # ===== FASE 2: Processar fila e enviar follow-ups =====
    for phone_clean, entry in list(queue.items()):
        try:
            status = entry.get("status", "pendente")
            if status != "pendente":
                continue

            data_prometida = entry.get("data_prometida")
            ultimo_followup = entry.get("ultimo_followup")
            tentativas = entry.get("tentativas", 0)
            criado_em = entry.get("criado_em", hoje.isoformat())

            # Determine if we should send follow-up today
            enviar = False
            motivo = ""

            if data_prometida:
                data_p = _dt_followup.date.fromisoformat(data_prometida)
                if hoje >= data_p:
                    enviar = True
                    motivo = f"data prometida: {data_prometida}"
            else:
                ref_date = ultimo_followup or criado_em
                dias_desde = (hoje - _dt_followup.date.fromisoformat(ref_date)).days
                if tentativas == 0 and dias_desde >= 1:
                    # First contact: next day after tag was added
                    enviar = True
                    motivo = "primeira cobrança"
                elif tentativas == 1 and dias_desde >= 3:
                    enviar = True
                    motivo = "2º lembrete (+3 dias)"
                elif tentativas == 2 and dias_desde >= 4:
                    enviar = True
                    motivo = "3º lembrete (+4 dias)"
                elif tentativas >= 3 and dias_desde >= 5:
                    enviar = True
                    motivo = f"{tentativas+1}º lembrete (+5 dias)"

            # Don't follow up more than once per day
            if ultimo_followup == hoje.isoformat():
                enviar = False
                motivo = "já contatado hoje"

            # Max 5 attempts - escalate to Michelle
            if tentativas >= 5:
                entry["status"] = "escalar"
                queue[phone_clean] = entry
                resultados.append({
                    "phone": phone_clean,
                    "nome": entry.get("nome"),
                    "docs": entry.get("docs_pendentes"),
                    "acao": "escalar",
                    "motivo": "5 tentativas sem sucesso - escalar para Michelle"
                })
                continue

            if not enviar:
                resultados.append({
                    "phone": phone_clean,
                    "nome": entry.get("nome"),
                    "acao": "aguardar",
                    "motivo": motivo or "ainda não é hora",
                    "tentativas": tentativas
                })
                continue

            # Read conversation history for context
            conv_msgs = []
            sid = entry.get("session_id")
            if sid:
                conv_msgs = _followup_read_conversation(sid)

            # Generate personalized message
            mensagem = _followup_generate_message(entry, conv_msgs)

            # Send message
            phone_formatted = f"+55{phone_clean}" if not phone_clean.startswith("55") else f"+{phone_clean}"
            try:
                whatsapp_send_message(phone_formatted, mensagem)
                entry["tentativas"] = tentativas + 1
                entry["ultimo_followup"] = hoje.isoformat()
                if data_prometida:
                    entry["data_prometida"] = None
                queue[phone_clean] = entry
                resultados.append({
                    "phone": phone_clean,
                    "nome": entry.get("nome"),
                    "docs": entry.get("docs_pendentes"),
                    "acao": "mensagem_enviada",
                    "tentativa": tentativas + 1,
                    "motivo": motivo,
                    "mensagem": mensagem[:150]
                })
                print(f"[FOLLOWUP] Mensagem enviada para {phone_clean} ({entry.get('nome')}): {mensagem[:80]}")
                import time as _time_fup
                _time_fup.sleep(10)  # Space out messages to seem natural
            except Exception as e:
                resultados.append({"phone": phone_clean, "nome": entry.get("nome"), "acao": "erro", "motivo": str(e)})

        except Exception as e:
            resultados.append({"phone": phone_clean, "acao": "erro", "motivo": str(e)})

    _followup_save(queue)

    # Count stats
    enviados = sum(1 for r in resultados if r.get("acao") == "mensagem_enviada")
    aguardando = sum(1 for r in resultados if r.get("acao") == "aguardar")
    escalar = sum(1 for r in resultados if r.get("acao") == "escalar")

    return jsonify({
        "status": "ok",
        "data": hoje.isoformat(),
        "total_fila": len(queue),
        "resumo": {"enviados": enviados, "aguardando": aguardando, "escalar": escalar},
        "resultados": resultados
    })

@app.route("/api/whatsapp/pausar", methods=["POST", "GET"])
def whatsapp_pausar():
    """Pause Ana for a specific phone number (human takeover)."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    phone = request.args.get("phone", "") or (request.get_json() or {}).get("phone", "")
    if not phone:
        return jsonify({"error": "Informe ?phone=NUMERO"}), 400
    phone_clean = re.sub(r'[^\d]', '', str(phone))
    from datetime import datetime as _dt_pause
    _paused_phones[phone_clean] = str(_dt_pause.now())
    _save_paused()
    print(f"[BOT] Ana PAUSADA para {phone_clean}")
    return jsonify({"status": "pausado", "phone": phone_clean, "message": f"Ana não responde mais para {phone_clean}. Use /retomar para voltar."})


@app.route("/api/whatsapp/retomar", methods=["POST", "GET"])
def whatsapp_retomar():
    """Resume Ana for a specific phone number."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    phone = request.args.get("phone", "") or (request.get_json() or {}).get("phone", "")
    if not phone:
        return jsonify({"error": "Informe ?phone=NUMERO"}), 400
    phone_clean = re.sub(r'[^\d]', '', str(phone))
    removed = _paused_phones.pop(phone_clean, None)
    if removed:
        _save_paused()
        print(f"[BOT] Ana RETOMADA para {phone_clean}")
        return jsonify({"status": "retomado", "phone": phone_clean, "message": f"Ana voltou a responder para {phone_clean}."})
    return jsonify({"status": "ok", "phone": phone_clean, "message": f"Ana já estava ativa para {phone_clean}."})


@app.route("/api/whatsapp/pausados")
def whatsapp_pausados():
    """List all paused phones."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"pausados": _paused_phones})


@app.route("/api/whatsapp/webhook-log")
def whatsapp_webhook_log():
    """View last received webhook payloads for debugging."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(_webhook_log)

@app.route("/api/whatsapp/debug-log")
def whatsapp_debug_log():
    """View last bot processing logs for debugging."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(_bot_debug_log)

@app.route("/api/whatsapp/webhook", methods=["POST"])
def whatsapp_webhook():
    """Webhook receiver for ConversApp events.

    ConversApp sends MESSAGE_RECEIVED events when clients send messages.
    Configure webhook in ConversApp: Settings > Integrations > Webhooks
    URL: https://your-app.railway.app/api/whatsapp/webhook?secret=YOUR_WEBHOOK_SECRET
    Event: MESSAGE_RECEIVED
    """
    # Validate webhook origin
    if WEBHOOK_SECRET:
        incoming_secret = request.args.get("secret", "")
        if not hmac.compare_digest(incoming_secret, WEBHOOK_SECRET):
            print(f"[SECURITY] Webhook rejeitado - secret inválido de {request.remote_addr}")
            return jsonify({"error": "forbidden"}), 403

    # Rate limit webhooks: max 120/min (prevents abuse)
    ip = request.remote_addr or "unknown"
    if not _rate_limit_check(f"webhook:{ip}", max_requests=120, window_seconds=60):
        return jsonify({"error": "rate limit"}), 429

    data = request.get_json(force=False, silent=True) or {}
    if not data:
        return jsonify({"error": "invalid payload"}), 400

    # Store for debug (limit stored data size)
    import datetime as _dtlog
    _webhook_log.append({"ts": str(_dtlog.datetime.now()), "ip": ip, "data": data})
    # deque(maxlen=20) auto-trims

    # Log webhook (truncate to avoid PII leaking in logs)
    import json as _json
    print(f"[WHATSAPP] Webhook de {ip}: eventType={data.get('eventType', 'unknown')}")

    # ConversApp real payload format (NO "data" wrapper):
    # { "eventType": "MESSAGE_RECEIVED", "content": { "text": "...", "direction": "FROM_HUB",
    #   "sessionId": "...", "details": { "from": "+55...", "to": "+55..." } } }
    # But may also come wrapped: { "data": { "eventType": ..., "content": {...} } }
    if "eventType" in data or "content" in data:
        inner = data  # payload is flat (real ConversApp format)
    else:
        inner = data.get("data") or data  # wrapped format
    if isinstance(inner, list) and inner:
        inner = inner[0]

    # Event type
    event_type = inner.get("eventType") or data.get("event") or data.get("type") or ""

    # Only process MESSAGE_RECEIVED events
    if event_type and event_type not in ("MESSAGE_RECEIVED", "message.received"):
        return jsonify({"status": "ok", "action": f"skipped_{event_type}"})

    # Content
    content = inner.get("content") or inner
    if isinstance(content, list) and content:
        content = content[0]
    details = (content.get("details") or {}) if isinstance(content, dict) else {}

    # Skip outgoing messages FIRST (FROM_HUB = incoming from client, TO_HUB = outgoing)
    direction = content.get("direction") or inner.get("direction") or ""
    from_me = content.get("fromMe") or inner.get("fromMe")
    is_outbound = direction.upper() in ("TO_HUB", "OUTBOUND", "OUT", "SENT") or from_me in (True, "true")

    if is_outbound:
        # Check for agent commands in outgoing messages before skipping
        out_text = ""
        if isinstance(content, dict):
            out_text = (content.get("text") or content.get("body") or "").strip().lower()
        out_details = (content.get("details") or {}) if isinstance(content, dict) else {}
        # For outbound messages: details.from = bot number, details.to = client number
        # We also check sessionId to get the contact phone from session
        out_phone = ""
        if isinstance(out_details, dict):
            # In outbound, "to" is the client (who we're sending to)
            out_from = out_details.get("from") or ""
            out_to = out_details.get("to") or ""
            # The client phone is whichever is NOT our bot number
            pos_venda = os.environ.get("CONVERSAPP_POS_VENDA_NUMBERS", "+5519982268158,+5516988124636").split(",")
            if out_to and out_to not in pos_venda:
                out_phone = out_to
            elif out_from and out_from not in pos_venda:
                out_phone = out_from
            else:
                out_phone = out_to or out_from
        if not out_phone:
            out_phone = content.get("to") or inner.get("to") or ""
        out_phone_clean = re.sub(r'[^\d]', '', str(out_phone))

        if out_text in ("#parar", "#pausar", "#p") and out_phone_clean:
            _paused_phones[out_phone_clean] = str(__import__('datetime').datetime.now())
            _save_paused()
            print(f"[BOT] Comando #parar detectado - Ana PAUSADA para {out_phone_clean}")
            # Try to delete the command message so client doesn't see it
            cmd_msg_id = (content.get("id") or inner.get("id") or "") if isinstance(content, dict) else ""
            cmd_session_id = content.get("sessionId") or inner.get("sessionId") or ""
            if cmd_msg_id and cmd_session_id:
                try:
                    conversapp_request("delete", f"/chat/v1/message/{cmd_msg_id}")
                    print(f"[BOT] Mensagem de comando deletada: {cmd_msg_id}")
                except Exception:
                    pass
            return jsonify({"status": "ok", "action": "command_parar", "phone": out_phone_clean})

        elif out_text in ("#retomar", "#voltar", "#r") and out_phone_clean:
            _paused_phones.pop(out_phone_clean, None)
            _save_paused()
            print(f"[BOT] Comando #retomar detectado - Ana RETOMADA para {out_phone_clean}")
            cmd_msg_id = (content.get("id") or inner.get("id") or "") if isinstance(content, dict) else ""
            if cmd_msg_id:
                try:
                    conversapp_request("delete", f"/chat/v1/message/{cmd_msg_id}")
                    print(f"[BOT] Mensagem de comando deletada: {cmd_msg_id}")
                except Exception:
                    pass
            return jsonify({"status": "ok", "action": "command_retomar", "phone": out_phone_clean})

        return jsonify({"status": "ok", "action": "skipped_outbound"})

    # Filter by destination number (pós-venda) since ConversApp doesn't send channelId
    dest_number = (details.get("to") or "") if isinstance(details, dict) else ""
    POS_VENDA_NUMBERS = os.environ.get("CONVERSAPP_POS_VENDA_NUMBERS", "+5519982268158,+5516988124636").split(",")
    if dest_number and POS_VENDA_NUMBERS and dest_number not in POS_VENDA_NUMBERS:
        print(f"[WHATSAPP] Ignorando msg para outro número: {dest_number}")
        return jsonify({"status": "ok", "action": "wrong_number"})

    # Extract phone number from details.from or fallbacks
    phone = ""
    if isinstance(details, dict):
        phone = details.get("from") or ""
    if not phone:
        contact = content.get("contact") or inner.get("contact") or inner.get("contactDetails") or {}
        if isinstance(contact, dict):
            phone = contact.get("phonenumber") or contact.get("phone") or contact.get("number") or ""
    if not phone:
        phone = content.get("from") or inner.get("from") or inner.get("number") or ""
    phone = re.sub(r'[^\d+]', '', str(phone)).lstrip('+')

    # Extract message text (also check audio transcription)
    message = ""
    if isinstance(content, dict):
        message = content.get("text") or content.get("body") or ""
        # If audio/voice, use transcription
        if not message and isinstance(details, dict):
            transcription = details.get("transcription") or {}
            if isinstance(transcription, dict):
                message = transcription.get("text") or ""
    if not message:
        message = inner.get("text") or inner.get("body") or ""
    if not message and isinstance(content, str):
        message = content

    # Extract session ID for replying
    session_id = None
    if isinstance(content, dict):
        session_id = content.get("sessionId")
    if not session_id:
        session_id = inner.get("sessionId") or data.get("sessionId")

    # Check if session is active with a human agent - if so, Ana doesn't respond
    if session_id and phone and message:
        try:
            sess_resp = conversapp_request("get", f"/chat/v1/session/{session_id}")
            if sess_resp.status_code == 200:
                sess_data = sess_resp.json()
                sess_status = sess_data.get("status", "")
                sess_user = sess_data.get("userId")
                sess_bot = sess_data.get("botId")
                if sess_status == "IN_PROGRESS" and sess_user:
                    print(f"[WHATSAPP] Sessão ativa com atendente humano ({sess_user}), Ana não responde para {phone}")
                    return jsonify({"status": "ok", "action": "skipped_human_session"})
        except Exception as e:
            print(f"[WHATSAPP] Erro ao verificar sessão: {e}")
            # If can't check, proceed normally

    # Deduplication: skip if we already processed this message
    msg_id = (content.get("id") or inner.get("id") or data.get("id") or "") if isinstance(content, dict) else ""
    if not msg_id and phone and message:
        # Create synthetic ID from phone + message + timestamp (approximate dedup)
        import hashlib
        msg_id = hashlib.md5(f"{phone}:{message[:50]}:{str(data.get('timestamp', ''))}".encode()).hexdigest()
    with _processed_msg_ids_lock:
        if msg_id and msg_id in _processed_msg_ids:
            print(f"[WHATSAPP] Mensagem duplicada ignorada: {msg_id[:20]}")
            return jsonify({"status": "ok", "action": "duplicate"})
        if msg_id:
            _processed_msg_ids.add(msg_id)
            # Limit dedup set size to prevent memory growth
            if len(_processed_msg_ids) > 5000:
                # Remove oldest half (set is unordered, but this prevents unbounded growth)
                excess = list(_processed_msg_ids)[:2500]
                _processed_msg_ids.difference_update(excess)

    # Periodic cleanup of old sessions
    from datetime import datetime as _dt_cleanup
    _session_last_activity[phone] = _dt_cleanup.now()
    if len(_session_last_activity) > 50:
        _cleanup_old_sessions()

    # Detect message type (TEXT, AUDIO, IMAGE, etc.) for smart reply mode
    msg_type = ""
    if isinstance(content, dict):
        msg_type = (content.get("type") or "").upper()
    client_sent_audio = msg_type in ("AUDIO", "VOICE", "PTT")

    if phone and message:
        print(f"[WHATSAPP] De {phone}: {message[:100]} (tipo={msg_type})")

        # Store session_id and audio preference
        if session_id:
            sess = _whatsapp_sessions.get(phone, {"historico": []})
            sess["conversapp_session_id"] = session_id
            if client_sent_audio:
                sess["_responder_audio"] = True
            _whatsapp_sessions[phone] = sess

        # Process in background thread to not block webhook response
        def _process():
            phone_lock = _get_phone_lock(phone)
            phone_lock.acquire()  # Wait if another message from same phone is processing
            import datetime as _dtproc
            log_entry = {"ts": str(_dtproc.datetime.now()), "phone": phone, "msg": message[:100]}
            try:
                # Send "typing" indicator
                sid = _whatsapp_sessions.get(phone, {}).get("conversapp_session_id") or session_id
                if sid:
                    try:
                        conversapp_request("post", f"/chat/v1/session/{sid}/typing", json={})
                    except Exception:
                        pass

                # Generate AI response
                resposta = whatsapp_processar_mensagem(phone, message)
                import time as _time

                # Handle two-message response (list: [consulta_msg, results_msg])
                if isinstance(resposta, list) and len(resposta) == 2:
                    consulta_msg, results_msg = resposta
                    log_entry["resposta"] = f"[2msgs] {consulta_msg[:80]} | {results_msg[:120]}"

                    if consulta_msg:
                        # Send "vou consultar" with normal 30s delay first
                        elapsed = (_dtproc.datetime.now() - _dtproc.datetime.fromisoformat(log_entry["ts"])).total_seconds()
                        delay = max(0, 12 - elapsed)
                        if delay > 0:
                            for _ in range(int(delay // 5)):
                                if sid:
                                    try:
                                        conversapp_request("post", f"/chat/v1/session/{sid}/typing", json={})
                                    except Exception:
                                        pass
                                _time.sleep(5)
                            remaining = delay % 5
                            if remaining > 0:
                                _time.sleep(remaining)
                        whatsapp_send_message(phone, consulta_msg, session_id=sid)

                    if results_msg:
                        # Wait 20-30s before sending results (simulate looking up)
                        for _ in range(5):
                            if sid:
                                try:
                                    conversapp_request("post", f"/chat/v1/session/{sid}/typing", json={})
                                except Exception:
                                    pass
                            _time.sleep(5)
                        # Smart mode: if client sent audio, respond with audio
                        should_audio = _whatsapp_sessions.get(phone, {}).get("_responder_audio", False)
                        if should_audio:
                            audio_data = elevenlabs_tts(results_msg)
                            if audio_data:
                                conversapp_send_audio(phone, audio_data, session_id=sid)
                                log_entry["audio"] = True
                            else:
                                whatsapp_send_message(phone, results_msg, session_id=sid)
                        else:
                            whatsapp_send_message(phone, results_msg, session_id=sid)
                        log_entry["enviado"] = True
                        # Reset audio flag after two-message response
                        sess = _whatsapp_sessions.get(phone, {})
                        sess.pop("_responder_audio", None)
                        _whatsapp_sessions[phone] = sess

                    log_entry["session_id"] = sid

                elif resposta:
                    # Detect session action markers before sending
                    acao_encerrar = False
                    acao_transferir = False
                    if isinstance(resposta, str):
                        if "[ENCERRAR_SESSAO]" in resposta:
                            acao_encerrar = True
                            resposta = resposta.replace("[ENCERRAR_SESSAO]", "").strip()
                        if "[TRANSFERIR_MICHELLE]" in resposta:
                            acao_transferir = True
                            resposta = resposta.replace("[TRANSFERIR_MICHELLE]", "").strip()

                    log_entry["resposta"] = resposta[:200] if isinstance(resposta, str) else str(resposta)[:200]
                    # Single message - normal 30s delay
                    elapsed = (_dtproc.datetime.now() - _dtproc.datetime.fromisoformat(log_entry["ts"])).total_seconds()
                    delay = max(0, 12 - elapsed)
                    if delay > 0:
                        for _ in range(int(delay // 5)):
                            if sid:
                                try:
                                    conversapp_request("post", f"/chat/v1/session/{sid}/typing", json={})
                                except Exception:
                                    pass
                            _time.sleep(5)
                        remaining = delay % 5
                        if remaining > 0:
                            _time.sleep(remaining)

                    log_entry["session_id"] = sid
                    # Smart mode: if client sent audio, respond with audio only
                    should_audio = _whatsapp_sessions.get(phone, {}).get("_responder_audio", False)
                    if should_audio and isinstance(resposta, str):
                        audio_data = elevenlabs_tts(resposta)
                        if audio_data:
                            conversapp_send_audio(phone, audio_data, session_id=sid)
                            log_entry["audio"] = True
                        else:
                            # Fallback to text if TTS fails
                            whatsapp_send_message(phone, resposta, session_id=sid)
                    else:
                        whatsapp_send_message(phone, resposta, session_id=sid)
                    log_entry["enviado"] = True
                    # Reset audio flag after responding
                    sess = _whatsapp_sessions.get(phone, {})
                    sess.pop("_responder_audio", None)
                    _whatsapp_sessions[phone] = sess

                    # Execute session actions after sending the message
                    if sid and acao_transferir:
                        _time.sleep(2)  # Small delay before transferring
                        conversapp_transfer_session(sid)
                        log_entry["acao"] = "transferir_michelle"
                        # Clear local session
                        _whatsapp_sessions.pop(phone, None)
                        _save_sessions()
                        print(f"[BOT] Sessão transferida para Michelle: {phone}")
                    elif sid and acao_encerrar:
                        _time.sleep(2)  # Small delay before closing
                        conversapp_complete_session(sid)
                        log_entry["acao"] = "encerrar_sessao"
                        # Clear local session
                        _whatsapp_sessions.pop(phone, None)
                        _save_sessions()
                        print(f"[BOT] Sessão encerrada: {phone}")
            except Exception as e:
                log_entry["erro"] = str(e)
                print(f"[WHATSAPP] Erro: {e}")
                traceback.print_exc()
            finally:
                phone_lock.release()  # Always release lock
            _bot_debug_log.append(log_entry)
            # deque(maxlen=20) auto-trims

        threading.Thread(target=_process, daemon=True).start()
    elif phone and not message:
        # Client sent image/audio/document without text
        print(f"[WHATSAPP] Mídia sem texto de {phone} (tipo={msg_type})")

        # Check pause (human takeover)
        phone_clean_media = re.sub(r'[^\d]', '', str(phone))
        if phone_clean_media in _paused_phones:
            print(f"[WHATSAPP] Ana pausada para {phone}, ignorando mídia")
            return jsonify({"status": "ok", "action": "paused"})

        sid = None
        if isinstance(content, dict):
            sid = content.get("sessionId")
        if not sid:
            sid = inner.get("sessionId") or data.get("sessionId")

        # Check if session has human agent active
        if sid:
            try:
                sess_resp_media = conversapp_request("get", f"/chat/v1/session/{sid}")
                if sess_resp_media.status_code == 200:
                    sess_data_media = sess_resp_media.json()
                    if sess_data_media.get("status") == "IN_PROGRESS" and sess_data_media.get("userId"):
                        print(f"[WHATSAPP] Sessão com atendente humano, Ana ignora mídia de {phone}")
                        return jsonify({"status": "ok", "action": "skipped_human_session"})
            except Exception:
                pass

        # Check if it's an image - analyze with Claude Vision
        file_url = None
        if isinstance(content, dict):
            file_id = content.get("fileId")
            # Try multiple locations where ConversApp might put the file URL
            file_url = content.get("fileUrl") or content.get("mediaUrl") or content.get("url") or ""
            if not file_url:
                file_details = (details.get("file") or {}) if isinstance(details, dict) else {}
                if isinstance(file_details, dict):
                    file_url = file_details.get("url") or file_details.get("link") or ""
            if not file_url:
                file_url = details.get("mediaUrl") or details.get("fileUrl") or details.get("url") or ""
            # If no file URL found, try to get it from file ID
            if not file_url and file_id:
                try:
                    file_resp = conversapp_request("get", f"/core/v1/file/{file_id}")
                    if file_resp.status_code == 200:
                        file_data = file_resp.json()
                        file_url = file_data.get("url") or file_data.get("link") or ""
                except Exception:
                    pass

        is_image = msg_type in ("IMAGE", "PHOTO", "STICKER")

        if is_image and file_url and sid:
            # Analyze image with Claude Vision in background
            def _process_image():
                phone_lock = _get_phone_lock(phone)
                phone_lock.acquire()
                try:
                    # Send typing indicator
                    try:
                        conversapp_request("post", f"/chat/v1/session/{sid}/typing", json={})
                    except Exception:
                        pass

                    # Download image
                    img_resp = requests.get(file_url, timeout=15)
                    if img_resp.status_code != 200 or len(img_resp.content) < 500:
                        whatsapp_send_message(phone, "Recebi a imagem mas não consegui abrir. Pode enviar novamente?", session_id=sid)
                        return

                    # Detect content type
                    img_content_type = img_resp.headers.get("Content-Type", "image/jpeg")
                    if "png" in img_content_type:
                        media_type = "image/png"
                    elif "webp" in img_content_type:
                        media_type = "image/webp"
                    else:
                        media_type = "image/jpeg"

                    img_b64 = base64.b64encode(img_resp.content).decode("utf-8")

                    # Analyze with Claude Vision (fallback: generic ack)
                    try:
                        client_ai = anthropic.Anthropic(timeout=120.0)
                    except Exception:
                        client_ai = None

                    _img_system = """Você é Ana, do pós-venda da JRC Advocacia. Um cliente enviou uma imagem/documento.

Analise a imagem e responda de forma curta e direta:

1. IDENTIFIQUE o tipo de documento (laudo médico, comprovante de residência, RG, CPF, receita, exame, declaração, etc.)
2. AVALIE A QUALIDADE:
   - Está LEGÍVEL? (texto nítido, sem cortes)
   - Está CORTADO? (faltando partes do documento)
   - Está EMBAÇADO ou ESCURO? (difícil de ler)
   - Está DE CABEÇA PRA BAIXO ou TORTO?

SE O DOCUMENTO ESTÁ BOM:
"Recebi o [tipo do documento]! Está bem legível. Vou encaminhar para o escritório analisar."

SE O DOCUMENTO ESTÁ RUIM (cortado, ilegível, embaçado):
"Recebi o [tipo do documento], mas [problema específico]. Pode enviar novamente [instrução específica]?"

Exemplos de instrução:
- "tirando a foto mais de cima para pegar o documento inteiro?"
- "com mais luz para ficar mais nítido?"
- "sem cortar as bordas do documento?"

SE NÃO É UM DOCUMENTO (foto pessoal, meme, etc.):
"Recebi sua imagem! Se precisar enviar algum documento, pode mandar por aqui que eu encaminho pro escritório."

Seja curta e direta. Máximo 3 linhas."""

                    resposta = None
                    if client_ai:
                        try:
                            response = client_ai.messages.create(
                                model="claude-haiku-4-5-20251001",
                                max_tokens=400,
                                system=_img_system,
                                messages=[{
                                    "role": "user",
                                    "content": [
                                        {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                                        {"type": "text", "text": "Analise esta imagem enviada pelo cliente."}
                                    ]
                                }],
                            )
                            resposta = response.content[0].text.strip()
                        except Exception as _img_err:
                            print(f"[BOT] Claude Vision falhou: {_img_err}")

                    if not resposta:
                        # Fallback: generic ack (can't do vision without Anthropic)
                        resposta = "Recebi seu documento! Vou encaminhar para o escritório analisar. Se precisar enviar mais alguma coisa, pode mandar por aqui."

                    time.sleep(5)  # Small delay to seem natural
                    whatsapp_send_message(phone, resposta, session_id=sid)
                    print(f"[BOT] Imagem analisada para {phone}: {resposta[:100]}")

                    # Check if document was received OK - update follow-up queue
                    resposta_lower = resposta.lower()
                    doc_ok = "legível" in resposta_lower or "legivel" in resposta_lower or "encaminhar" in resposta_lower
                    if doc_ok:
                        docs_recebidos = _followup_detect_docs(resposta)
                        if docs_recebidos:
                            queue = _followup_load()
                            phone_fup = re.sub(r'[^\d]', '', str(phone))
                            if phone_fup in queue:
                                pendentes = queue[phone_fup].get("docs_pendentes", [])
                                for doc in docs_recebidos:
                                    if doc in pendentes:
                                        pendentes.remove(doc)
                                queue[phone_fup]["docs_pendentes"] = pendentes
                                if not pendentes:
                                    queue[phone_fup]["status"] = "completo"
                                    print(f"[FOLLOWUP] Todos os docs recebidos de {phone_fup}!")
                                else:
                                    print(f"[FOLLOWUP] Doc recebido de {phone_fup}, ainda faltam: {pendentes}")
                                _followup_save(queue)

                except Exception as e:
                    print(f"[WHATSAPP] Erro ao analisar imagem: {e}")
                    traceback.print_exc()
                    try:
                        whatsapp_send_message(phone, "Recebi o documento! Vou encaminhar para o escritório analisar.", session_id=sid)
                    except Exception:
                        pass
                finally:
                    phone_lock.release()

            threading.Thread(target=_process_image, daemon=True).start()

        elif sid:
            try:
                if msg_type in ("AUDIO", "VOICE", "PTT"):
                    whatsapp_send_message(phone, "Recebi seu áudio! Infelizmente não consegui ouvir. Pode me escrever por texto?", session_id=sid)
                else:
                    whatsapp_send_message(phone, "Recebi o arquivo! Vou encaminhar para o escritório.", session_id=sid)
            except Exception:
                pass
    else:
        print(f"[WHATSAPP] Sem texto processável. phone={phone}, content_type={type(content).__name__}")

    return jsonify({"status": "ok"})


@app.route("/api/whatsapp/test", methods=["POST"])
def whatsapp_test():
    """Test the bot logic without WhatsApp - send a message and get the response."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    data = request.get_json() or {}
    phone = data.get("phone", "test")
    message = data.get("message", "")
    if not message:
        return jsonify({"error": "Informe 'message'"}), 400

    resposta = whatsapp_processar_mensagem(phone, message)
    return jsonify({"response": resposta, "session": _whatsapp_sessions.get(phone, {})})


@app.route("/api/whatsapp/status")
@require_admin
def whatsapp_status():
    """Check WhatsApp bot configuration status."""
    return jsonify({
        "helena_token": bool(CONVERSAPP_API_TOKEN),
        "active_sessions": len(_whatsapp_sessions),
        "gmail_configured": bool(GMAIL_REFRESH_TOKEN and GMAIL_CLIENT_ID),
    })


@app.route("/api/whatsapp/test-gmail")
def whatsapp_test_gmail():
    """Test Gmail search for a name."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    nome = request.args.get("nome", "")
    if not nome:
        return jsonify({"error": "Informe ?nome=..."}), 400
    try:
        result = whatsapp_buscar_gmail_inss(nome)
        if result:
            # Remove corpo longo para facilitar leitura
            result["corpo"] = result.get("corpo", "")[:300]
        return jsonify({"found": bool(result), "result": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/whatsapp/conversapp-fields")
def whatsapp_conversapp_fields():
    """List ConversApp custom fields and tags (to discover key names)."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    fields = conversapp_get_custom_fields()
    tags = conversapp_get_tags()
    return jsonify({"custom_fields": fields, "tags": tags})


@app.route("/api/whatsapp/conversapp-contact")
def whatsapp_conversapp_contact():
    """Look up a contact in ConversApp by phone number."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    phone = request.args.get("phone", "")
    if not phone:
        return jsonify({"error": "Informe ?phone=..."}), 400
    contact = conversapp_get_contact(phone)
    return jsonify({"found": bool(contact), "contact": contact})


@app.route("/api/whatsapp/setup-webhook", methods=["POST"])
@require_admin
def whatsapp_setup_webhook():
    """Register webhook on ConversApp to receive messages.

    Call this once after deploy to configure ConversApp to send events here.
    Body: {"app_url": "https://your-app.railway.app"}
    """
    if not CONVERSAPP_API_TOKEN:
        return jsonify({"error": "CONVERSAPP_API_TOKEN não configurado"}), 400

    data = request.get_json() or {}
    app_url = data.get("app_url", "").rstrip("/")
    if not app_url:
        return jsonify({"error": "Informe app_url"}), 400

    webhook_url = f"{app_url}/api/whatsapp/webhook"

    # First, list available events
    try:
        resp = conversapp_request("get", "/core/v1/webhook/event")
        events = resp.json() if resp.status_code == 200 else []
        print(f"[HELENA] Eventos disponíveis: {events}")
    except Exception as e:
        events = []
        print(f"[HELENA] Erro ao listar eventos: {e}")

    # Find message-related events
    msg_events = []
    if isinstance(events, list):
        for ev in events:
            ev_name = ev.get("name") or ev.get("event") or str(ev)
            if any(kw in str(ev_name).lower() for kw in ["message", "mensagem", "received", "inbound"]):
                msg_events.append(ev_name)

    # Create webhook subscription
    try:
        payload = {
            "url": webhook_url,
            "events": msg_events if msg_events else ["message.received"],
        }
        resp = conversapp_request("post", "/core/v1/webhook/subscription", json=payload)
        result = resp.json() if resp.status_code in (200, 201) else resp.text[:300]
        return jsonify({
            "status": "ok" if resp.status_code in (200, 201) else "erro",
            "webhook_url": webhook_url,
            "events_subscribed": msg_events or ["message.received"],
            "available_events": events[:20] if events else "não listados",
            "helena_response": result,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Inicialização — roda tanto com gunicorn quanto python app.py
def _startup():
    if LEGALMAIL_API_KEY:
        # Se não há cache ainda, popula em background antes de iniciar o monitor
        if not os.path.exists(PROCESSES_CACHE_FILE):
            def _pre_cache():
                print("[STARTUP] Cache vazio, buscando processos...")
                procs = monitor_fetch_all_processes()
                if procs:
                    _save_json_file(PROCESSES_CACHE_FILE, procs)
                    print(f"[STARTUP] Cache populado com {len(procs)} processos")
            threading.Thread(target=_pre_cache, daemon=True).start()
        start_monitor()
        print(f"Monitor automático ativado (verifica a cada {MONITOR_INTERVAL_MINUTES} min)")
    else:
        print("AVISO: LEGALMAIL_API_KEY não configurada - monitor desativado")

# ========== FOLLOW-UP AUTOMÁTICO (DISTRIBUÍDO AO LONGO DO DIA) ==========
FOLLOWUP_ENABLED = os.environ.get("FOLLOWUP_ENABLED", "false").lower() == "true"
FOLLOWUP_MAX_DIA = int(os.environ.get("FOLLOWUP_MAX_DIA", "5"))  # Máximo de mensagens por dia
FOLLOWUP_HORA_INICIO = int(os.environ.get("FOLLOWUP_HORA_INICIO", "9"))  # Começa às 9h
FOLLOWUP_HORA_FIM = int(os.environ.get("FOLLOWUP_HORA_FIM", "17"))  # Termina às 17h

# Track daily sends to enforce limit
_followup_enviados_hoje = {"data": "", "count": 0}
_followup_fila_diaria = []  # Contacts to send today, processed one at a time
_followup_escaneado_hoje = False

import random as _random_fup

def _followup_scan_tags():
    """Scan ConversApp tags and build the daily send queue. Returns list of entries to send."""
    hoje = _dt_followup.date.today()
    queue = _followup_load()
    para_enviar = []

    # Scan tags and update queue
    for tag_id, tag_nome in FOLLOWUP_TAGS.items():
        doc_nome = TAG_TO_DOC.get(tag_nome, "documentos pendentes")
        contatos = _followup_buscar_contatos_por_tag(tag_id)
        print(f"[FOLLOWUP] Tag '{tag_nome}': {len(contatos)} contatos")

        for contato in contatos:
            try:
                contact_id = contato.get("id", "")
                nome = contato.get("name") or ""
                phone_raw = ""
                phone_numbers = contato.get("phoneNumbers") or contato.get("phonenumber") or ""
                if isinstance(phone_numbers, list) and phone_numbers:
                    phone_raw = phone_numbers[0]
                elif isinstance(phone_numbers, str):
                    phone_raw = phone_numbers
                if not phone_raw:
                    phone_raw = contato.get("phoneNumber") or contato.get("phone") or ""
                phone_clean = re.sub(r'[^\d]', '', str(phone_raw))
                if not phone_clean or len(phone_clean) < 10:
                    continue

                entry = queue.get(phone_clean)
                if not entry:
                    last_session = _followup_get_last_session(contact_id)
                    sid = last_session.get("sessionId") if last_session else None
                    entry = {
                        "nome": nome,
                        "phone": phone_clean,
                        "contact_id": contact_id,
                        "docs_pendentes": [doc_nome],
                        "data_prometida": None,
                        "session_id": sid,
                        "contexto": "",
                        "criado_em": hoje.isoformat(),
                        "atualizado_em": hoje.isoformat(),
                        "tentativas": 0,
                        "ultimo_followup": None,
                        "status": "pendente",
                        "tags": [tag_nome],
                    }
                    queue[phone_clean] = entry
                else:
                    existing_docs = entry.get("docs_pendentes", [])
                    if doc_nome not in existing_docs:
                        existing_docs.append(doc_nome)
                        entry["docs_pendentes"] = existing_docs
                    existing_tags = entry.get("tags", [])
                    if tag_nome not in existing_tags:
                        existing_tags.append(tag_nome)
                        entry["tags"] = existing_tags
                    queue[phone_clean] = entry
            except Exception:
                continue

    # Determine which entries need a message today
    for phone_clean, entry in queue.items():
        if entry.get("status") != "pendente":
            continue
        tentativas = entry.get("tentativas", 0)
        if tentativas >= 5:
            entry["status"] = "escalar"
            continue

        data_prometida = entry.get("data_prometida")
        ultimo_followup = entry.get("ultimo_followup")
        criado_em = entry.get("criado_em", hoje.isoformat())

        enviar = False
        if data_prometida:
            if hoje >= _dt_followup.date.fromisoformat(data_prometida):
                enviar = True
        else:
            ref_date = ultimo_followup or criado_em
            dias_desde = (hoje - _dt_followup.date.fromisoformat(ref_date)).days
            if tentativas == 0 and dias_desde >= 1:
                enviar = True
            elif tentativas == 1 and dias_desde >= 3:
                enviar = True
            elif tentativas == 2 and dias_desde >= 4:
                enviar = True
            elif tentativas >= 3 and dias_desde >= 5:
                enviar = True

        if ultimo_followup == hoje.isoformat():
            enviar = False

        if enviar:
            para_enviar.append(phone_clean)

    _followup_save(queue)
    # Shuffle to vary who gets contacted first each day
    _random_fup.shuffle(para_enviar)
    return para_enviar[:FOLLOWUP_MAX_DIA]  # Cap at max per day


def _followup_send_one(phone_clean):
    """Send one follow-up message to a single contact."""
    queue = _followup_load()
    entry = queue.get(phone_clean)
    if not entry:
        return None
    hoje = _dt_followup.date.today()

    try:
        # Read conversation for context
        conv_msgs = []
        sid = entry.get("session_id")
        if sid:
            conv_msgs = _followup_read_conversation(sid)

        # Generate personalized message
        mensagem = _followup_generate_message(entry, conv_msgs)

        # Send - alternate between text and audio
        phone_formatted = f"+55{phone_clean}" if not phone_clean.startswith("55") else f"+{phone_clean}"
        tentativas = entry.get("tentativas", 0)
        usar_audio = (tentativas % 2 == 1) and ELEVENLABS_API_KEY
        formato = "audio" if usar_audio else "texto"

        if usar_audio:
            audio_data = elevenlabs_tts(mensagem)
            if audio_data:
                conversapp_send_audio(phone_formatted, audio_data)
            else:
                whatsapp_send_message(phone_formatted, mensagem)
                formato = "texto (fallback)"
        else:
            whatsapp_send_message(phone_formatted, mensagem)

        # Update entry
        entry["tentativas"] = tentativas + 1
        entry["ultimo_followup"] = hoje.isoformat()
        if entry.get("data_prometida"):
            entry["data_prometida"] = None
        queue[phone_clean] = entry
        _followup_save(queue)

        print(f"[FOLLOWUP] ✓ Enviado ({formato}) para {entry.get('nome', phone_clean)}: {mensagem[:80]}")
        return {"phone": phone_clean, "nome": entry.get("nome"), "mensagem": mensagem[:150], "tentativa": entry["tentativas"], "formato": formato}
    except Exception as e:
        print(f"[FOLLOWUP] ✗ Erro ao enviar para {phone_clean}: {e}")
        return None


def _followup_timer_loop():
    """Background thread that distributes follow-up messages throughout the day."""
    import time as _time_timer
    global _followup_fila_diaria, _followup_escaneado_hoje, _followup_enviados_hoje

    print(f"[FOLLOWUP] Timer iniciado - máx {FOLLOWUP_MAX_DIA} msgs/dia, {FOLLOWUP_HORA_INICIO}h-{FOLLOWUP_HORA_FIM}h")

    while True:
        try:
            # Use Brazil timezone (UTC-3) for hour checks
            _br_tz = _dt_followup.timezone(_dt_followup.timedelta(hours=-3))
            agora = _dt_followup.datetime.now(_br_tz)
            hoje_str = agora.date().isoformat()
            hora = agora.hour

            # Reset daily counters at midnight
            if _followup_enviados_hoje["data"] != hoje_str:
                _followup_enviados_hoje = {"data": hoje_str, "count": 0}
                _followup_fila_diaria = []
                _followup_escaneado_hoje = False

            # Scan tags once per day at HORA_INICIO
            if hora >= FOLLOWUP_HORA_INICIO and not _followup_escaneado_hoje:
                print(f"[FOLLOWUP] Escaneando tags do ConversApp...")
                with app.app_context():
                    _followup_fila_diaria = _followup_scan_tags()
                _followup_escaneado_hoje = True
                print(f"[FOLLOWUP] Fila do dia: {len(_followup_fila_diaria)} contatos para enviar (máx {FOLLOWUP_MAX_DIA})")

            # Send one message if there are contacts in today's queue
            if (FOLLOWUP_HORA_INICIO <= hora < FOLLOWUP_HORA_FIM
                    and _followup_fila_diaria
                    and _followup_enviados_hoje["count"] < FOLLOWUP_MAX_DIA):

                phone_clean = _followup_fila_diaria.pop(0)
                with app.app_context():
                    result = _followup_send_one(phone_clean)
                if result:
                    _followup_enviados_hoje["count"] += 1
                    restantes = len(_followup_fila_diaria)
                    print(f"[FOLLOWUP] {_followup_enviados_hoje['count']}/{FOLLOWUP_MAX_DIA} enviados hoje, {restantes} restantes na fila")

                # Wait 30-60 minutes before next message (random to seem human)
                intervalo = _random_fup.randint(30, 60) * 60
                print(f"[FOLLOWUP] Próxima mensagem em ~{intervalo // 60} minutos")
                _time_timer.sleep(intervalo)
            else:
                # Nothing to send or outside hours - check again in 15 min
                _time_timer.sleep(900)

        except Exception as e:
            print(f"[FOLLOWUP] Erro no timer: {e}")
            traceback.print_exc()
            _time_timer.sleep(3600)

if FOLLOWUP_ENABLED:
    threading.Thread(target=_followup_timer_loop, daemon=True).start()
    print(f"[FOLLOWUP] Follow-up ATIVADO - máx {FOLLOWUP_MAX_DIA} msgs/dia entre {FOLLOWUP_HORA_INICIO}h-{FOLLOWUP_HORA_FIM}h")
else:
    print("[FOLLOWUP] Follow-up DESATIVADO - ative com FOLLOWUP_ENABLED=true")


# ========== SALÁRIO MATERNIDADE - ACOMPANHAMENTO AUTOMÁTICO ==========

MATERNIDADE_ENABLED = os.environ.get("MATERNIDADE_ENABLED", "false").lower() == "true"
MATERNIDADE_TAG_ID = "664ffa40-4b5b-4d9e-9f3d-e2e709d1195a"  # Tag "Salário maternidade"
MATERNIDADE_GPS_TAG_ID = os.environ.get("MATERNIDADE_GPS_TAG_ID", "f9d9c6a8-59e9-4e4b-bf9a-d2a689dbddd7")  # Tag "Gerar GPS"
MATERNIDADE_KEY_NIT = os.environ.get("MATERNIDADE_KEY_NIT", "nit")  # Key do campo NIT no ConversApp
MATERNIDADE_KEY_DATA_PARTO = os.environ.get("MATERNIDADE_KEY_DATA_PARTO", "data-prevista-do-par")  # Key do campo data parto
MATERNIDADE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "maternidade_queue.json")
MATERNIDADE_MAX_DIA = int(os.environ.get("MATERNIDADE_MAX_DIA", "5"))  # Máx mensagens por dia
_maternidade_file_lock = threading.Lock()


def _maternidade_load():
    """Load maternity tracking queue from disk (thread-safe, with backup fallback)."""
    return _safe_json_load(MATERNIDADE_FILE, lock=_maternidade_file_lock)

def _maternidade_save(queue):
    """Save maternity tracking queue to disk (thread-safe, atomic with backup)."""
    _safe_json_save(MATERNIDADE_FILE, queue, lock=_maternidade_file_lock)


def _maternidade_generate_engagement(entry, tipo="engajamento"):
    """Generate a personalized engagement message for a maternity client."""
    try:
        nome = entry.get("nome", "")
        data_parto = entry.get("data_parto", "")
        nit = entry.get("nit", "")
        tentativas_eng = entry.get("engajamentos_enviados", 0)
        fase = entry.get("fase", "gestacao")  # gestacao, pre_parto, pos_parto

        if tipo == "gps_aviso":
            return None

        if tipo == "pre_gps":
            primeiro_nome = nome.split()[0] if nome else ""
            meses_rest = ""
            if data_parto:
                try:
                    dias = (_dt_followup.date.fromisoformat(data_parto) - _dt_followup.date.today()).days
                    if dias > 30:
                        meses_rest = f"faltam cerca de {dias // 30} meses"
                    elif dias > 0:
                        meses_rest = f"faltam {dias} dias"
                except Exception:
                    pass
            client_ai = anthropic.Anthropic(timeout=120.0)
            response = client_ai.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=200,
                system=f"""Você é Ana, do pós-venda da JRC Advocacia. Precisa avisar uma gestante que o escritório vai gerar a guia de pagamento do INSS (GPS) pra ela pagar e garantir o salário maternidade.

Nome: {primeiro_nome}
Parto previsto: {data_parto} ({meses_rest})

IMPORTANTE: Com apenas 1 contribuição antes do parto ela já garante o benefício. Quanto antes pagar, melhor.

A mensagem deve:
- Ser curta (2-3 linhas)
- Perguntar como ela está
- Avisar que vão enviar a guia de pagamento do INSS pra ela
- Explicar que é importante pagar o quanto antes pra garantir o benefício
- Dizer que a Michelle vai enviar a guia
- Sem emojis
- Humanizada
- NÃO dizer "o parto está se aproximando" se ainda faltam muitos meses

Escreva APENAS a mensagem.""",
                messages=[{"role": "user", "content": "Gere a mensagem."}],
            )
            return response.content[0].text.strip()

        if tipo == "gps_lembrete":
            primeiro_nome = nome.split()[0] if nome else ""
            return f"Oi {primeiro_nome}! Tudo bem com você e o bebê? Passando pra lembrar sobre a guia de pagamento do INSS que te enviamos. É importante pagar antes do vencimento pra garantir seu benefício. Qualquer dúvida me chama!"

        if tipo == "pos_parto":
            primeiro_nome = nome.split()[0] if nome else ""
            return f"Oi {primeiro_nome}! Como está o bebê? Quando puder, me envia a certidão de nascimento pra gente dar entrada no seu salário maternidade!"

        # Engagement during pregnancy
        client_ai = anthropic.Anthropic(timeout=120.0)

        meses_restantes = ""
        if data_parto:
            try:
                data_p = _dt_followup.date.fromisoformat(data_parto)
                dias = (data_p - _dt_followup.date.today()).days
                if dias > 0:
                    meses_restantes = f"Faltam aproximadamente {dias // 30} meses e {dias % 30} dias para o parto"
                else:
                    meses_restantes = f"O parto estava previsto para {data_parto} (já passou)"
            except Exception:
                pass

        response = client_ai.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=250,
            system=f"""Você é Ana, do pós-venda da JRC Advocacia. Precisa enviar uma mensagem de engajamento para uma gestante que é cliente do escritório para o benefício de salário maternidade.

CONTEXTO:
- Nome: {nome}
- Data prevista do parto: {data_parto or 'não informada'}
- {meses_restantes}
- Mensagens de engajamento já enviadas: {tentativas_eng}

OBJETIVO:
- Manter o vínculo com a cliente
- Mostrar que o escritório se importa
- Diminuir chance de inadimplência
- NÃO falar de valores, honorários ou pagamento do escritório
- Pode perguntar como ela está, como está a gestação, se está tudo bem

REGRAS:
- Mensagem curta (2-3 linhas máximo)
- Humanizada, como uma pessoa real
- NÃO use emojis
- Varie o texto a cada vez (não repita mensagens anteriores)
- Chame pelo primeiro nome
- Seja acolhedora e carinhosa
- Se for perto do parto, deseje tudo de bom

Escreva APENAS a mensagem.""",
            messages=[{"role": "user", "content": "Gere a mensagem de engajamento."}],
        )
        return response.content[0].text.strip()
    except Exception as e:
        print(f"[MATERNIDADE] Erro ao gerar mensagem: {e}")
        primeiro_nome = (entry.get("nome") or "").split()[0] if entry.get("nome") else ""
        return f"Oi {primeiro_nome}! Tudo bem com você? Passando pra saber como está a gestação. Qualquer dúvida sobre o benefício é só me chamar!"


def _maternidade_scan():
    """Scan ConversApp for maternity clients and update the tracking queue."""
    hoje = _dt_followup.date.today()
    queue = _maternidade_load()

    # Fetch all contacts with "salário maternidade" tag
    contatos = _followup_buscar_contatos_por_tag(MATERNIDADE_TAG_ID)
    print(f"[MATERNIDADE] Tag 'salário maternidade': {len(contatos)} contatos")

    for contato in contatos:
        try:
            contact_id = contato.get("id", "")
            nome = contato.get("name") or ""
            phone_raw = ""
            phone_numbers = contato.get("phoneNumbers") or contato.get("phonenumber") or ""
            if isinstance(phone_numbers, list) and phone_numbers:
                phone_raw = phone_numbers[0]
            elif isinstance(phone_numbers, str):
                phone_raw = phone_numbers
            if not phone_raw:
                phone_raw = contato.get("phoneNumber") or contato.get("phone") or ""
            phone_clean = re.sub(r'[^\d]', '', str(phone_raw))
            if not phone_clean or len(phone_clean) < 10:
                continue

            # Get custom fields
            custom_fields = contato.get("customFields") or {}
            nit = ""
            data_parto = ""
            if isinstance(custom_fields, dict):
                nit = custom_fields.get(MATERNIDADE_KEY_NIT, "") or ""
                data_parto_raw = custom_fields.get(MATERNIDADE_KEY_DATA_PARTO, "") or ""
                # Try to parse date (could be DD/MM/YYYY or YYYY-MM-DD)
                if data_parto_raw:
                    try:
                        if "/" in data_parto_raw:
                            parts = data_parto_raw.split("/")
                            if len(parts) == 3:
                                data_parto = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                        elif "-" in data_parto_raw:
                            data_parto = data_parto_raw[:10]
                    except Exception:
                        pass
            elif isinstance(custom_fields, list):
                for cf in custom_fields:
                    key = cf.get("key") or cf.get("name", "")
                    val = cf.get("value", "")
                    if key == MATERNIDADE_KEY_NIT:
                        nit = val
                    elif key == MATERNIDADE_KEY_DATA_PARTO:
                        data_parto_raw = val
                        try:
                            if "/" in str(data_parto_raw):
                                parts = str(data_parto_raw).split("/")
                                if len(parts) == 3:
                                    data_parto = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                            elif "-" in str(data_parto_raw):
                                data_parto = str(data_parto_raw)[:10]
                        except Exception:
                            pass

            # Get or create entry
            entry = queue.get(phone_clean)
            if not entry:
                last_session = _followup_get_last_session(contact_id)
                sid = last_session.get("sessionId") if last_session else None
                entry = {
                    "nome": nome,
                    "phone": phone_clean,
                    "contact_id": contact_id,
                    "nit": nit,
                    "data_parto": data_parto,
                    "session_id": sid,
                    "fase": "gestacao",
                    "criado_em": hoje.isoformat(),
                    "engajamentos_enviados": 0,
                    "ultimo_engajamento": None,
                    "gps_notificado": False,
                    "gps_enviada": False,
                    "certidao_pedida": False,
                    "status": "ativo",
                }
            else:
                # Update fields that might have changed
                if nit:
                    entry["nit"] = nit
                if data_parto:
                    entry["data_parto"] = data_parto
                entry["nome"] = nome or entry.get("nome", "")

            # Determine phase based on birth date and GPS status
            if data_parto:
                try:
                    data_p = _dt_followup.date.fromisoformat(data_parto)
                    dias_ate_parto = (data_p - hoje).days

                    if dias_ate_parto < -7:
                        # More than 7 days after due date - post-birth phase
                        entry["fase"] = "pos_parto"
                    elif not entry.get("gps_notificado"):
                        # GPS not generated yet - priority! Generate ASAP
                        entry["fase"] = "pre_parto"
                    else:
                        # GPS already notified - engagement until birth
                        entry["fase"] = "gestacao"
                except Exception:
                    pass

            queue[phone_clean] = entry
        except Exception as e:
            print(f"[MATERNIDADE] Erro ao processar contato: {e}")
            continue

    _maternidade_save(queue)
    return queue


def _maternidade_process():
    """Process maternity queue: send engagement, GPS notifications, birth certificate requests."""
    hoje = _dt_followup.date.today()
    queue = _maternidade_scan()
    enviados = 0
    resultados = []

    for phone_clean, entry in list(queue.items()):
        if enviados >= MATERNIDADE_MAX_DIA:
            break
        if entry.get("status") != "ativo":
            continue

        fase = entry.get("fase", "gestacao")
        ultimo_eng = entry.get("ultimo_engajamento")
        data_parto = entry.get("data_parto", "")
        nit = entry.get("nit", "")

        # Don't message more than once per day
        if ultimo_eng == hoje.isoformat():
            continue

        mensagem = None
        acao = ""

        # ===== PRE-PARTO: Notify team to generate GPS =====
        if fase == "pre_parto" and not entry.get("gps_notificado") and nit:
            # Tag contact with "Gerar GPS" if tag ID is configured
            if MATERNIDADE_GPS_TAG_ID:
                try:
                    contact_id = entry.get("contact_id")
                    if contact_id:
                        # Get current tags
                        contact_resp = conversapp_request("get", f"/core/v1/contact/{contact_id}")
                        if contact_resp.status_code == 200:
                            contact_data = contact_resp.json()
                            current_tags = contact_data.get("tagIds") or []
                            if MATERNIDADE_GPS_TAG_ID not in current_tags:
                                current_tags.append(MATERNIDADE_GPS_TAG_ID)
                                conversapp_request("put", f"/core/v1/contact/{contact_id}", json={"tagIds": current_tags})
                                print(f"[MATERNIDADE] Tag 'Gerar GPS' adicionada: {entry.get('nome')}")
                except Exception as e:
                    print(f"[MATERNIDADE] Erro ao adicionar tag GPS: {e}")

            # Send friendly message to the client about the payment guide
            phone_formatted = f"+55{phone_clean}" if not phone_clean.startswith("55") else f"+{phone_clean}"
            primeiro_nome = (entry.get("nome") or "").split()[0] if entry.get("nome") else ""
            msg_cliente = _maternidade_generate_engagement(entry, tipo="pre_gps")
            if not msg_cliente:
                msg_cliente = f"Oi {primeiro_nome}! Tudo bem? O parto está se aproximando e precisamos gerar a sua guia de pagamento do INSS pra garantir o salário maternidade. A Michelle do nosso escritório vai te enviar a guia em breve!"
            sid = entry.get("session_id")
            try:
                whatsapp_send_message(phone_formatted, msg_cliente, session_id=sid)
            except Exception:
                pass

            # Create INTERNAL NOTE in the conversation (only team sees it, not the client)
            if sid:
                try:
                    import time as _t_mat
                    _t_mat.sleep(3)
                    competencia = hoje.strftime("%m/%Y")
                    note_text = f"⚠️ GERAR GPS - Salário Maternidade\n\nCliente: {entry.get('nome')}\nNIT: {nit}\nCódigo: 1473 (Facultativo Simplificado)\nCompetência: {competencia}\nValor: R$178,31 (11% de R$1.621)\nParto previsto: {data_parto}\n\nGerar pelo site: meu.inss.gov.br > Emissão de GPS"
                    conversapp_request("post", f"/chat/v1/session/{sid}/note", json={"text": note_text})
                    print(f"[MATERNIDADE] Nota interna + mensagem enviadas: {entry.get('nome')}")

                    _t_mat.sleep(2)
                    # Transfer session to Michelle so she sees the note
                    conversapp_transfer_session(sid, MICHELLE_USER_ID)
                    print(f"[MATERNIDADE] Sessão transferida pra Michelle: {entry.get('nome')}")
                except Exception as e:
                    print(f"[MATERNIDADE] Erro ao notificar GPS: {e}")

            entry["gps_notificado"] = True
            acao = "gps_notificado"
            resultados.append({"phone": phone_clean, "nome": entry.get("nome"), "acao": acao})
            queue[phone_clean] = entry
            continue  # Don't send engagement on same cycle as GPS notification

        # ===== PÓS-PARTO: Pedir certidão de nascimento =====
        if fase == "pos_parto" and not entry.get("certidao_pedida"):
            mensagem = _maternidade_generate_engagement(entry, tipo="pos_parto")
            entry["certidao_pedida"] = True
            acao = "certidao_pedida"

        # ===== ENGAJAMENTO PERIÓDICO (a cada 15 dias) =====
        elif fase in ("gestacao", "pre_parto"):
            dias_desde = 999
            if ultimo_eng:
                dias_desde = (hoje - _dt_followup.date.fromisoformat(ultimo_eng)).days
            elif entry.get("criado_em"):
                dias_desde = (hoje - _dt_followup.date.fromisoformat(entry["criado_em"])).days

            if dias_desde >= 15:
                # Check if GPS was sent and needs payment follow-up
                if entry.get("gps_enviada") and not entry.get("gps_paga"):
                    mensagem = _maternidade_generate_engagement(entry, tipo="gps_lembrete")
                    acao = "gps_lembrete"
                else:
                    mensagem = _maternidade_generate_engagement(entry, tipo="engajamento")
                    acao = "engajamento"

        # Send message if we have one - alternate between text and audio
        if mensagem:
            phone_formatted = f"+55{phone_clean}" if not phone_clean.startswith("55") else f"+{phone_clean}"
            try:
                eng_count = entry.get("engajamentos_enviados", 0)
                # Alternate: even = text, odd = audio (starts with text)
                usar_audio = (eng_count % 2 == 1) and ELEVENLABS_API_KEY
                formato = "audio" if usar_audio else "texto"

                if usar_audio:
                    audio_data = elevenlabs_tts(mensagem)
                    if audio_data:
                        conversapp_send_audio(phone_formatted, audio_data)
                    else:
                        # Fallback to text if TTS fails
                        whatsapp_send_message(phone_formatted, mensagem)
                        formato = "texto (fallback)"
                else:
                    whatsapp_send_message(phone_formatted, mensagem)

                entry["ultimo_engajamento"] = hoje.isoformat()
                entry["engajamentos_enviados"] = eng_count + 1
                queue[phone_clean] = entry
                enviados += 1
                resultados.append({
                    "phone": phone_clean,
                    "nome": entry.get("nome"),
                    "fase": fase,
                    "acao": acao,
                    "formato": formato,
                    "mensagem": mensagem[:100]
                })
                print(f"[MATERNIDADE] Enviado ({formato}) para {entry.get('nome')}: {mensagem[:80]}")
            except Exception as e:
                print(f"[MATERNIDADE] Erro ao enviar: {e}")

    _maternidade_save(queue)
    return resultados


# Maternity endpoint for manual execution and monitoring
@app.route("/api/maternidade/fila", methods=["GET"])
def maternidade_fila():
    """View the maternity tracking queue."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    queue = _maternidade_load()
    return jsonify({"total": len(queue), "clientes": queue})


@app.route("/api/maternidade/executar", methods=["POST", "GET"])
def maternidade_executar():
    """Run maternity check manually."""
    if not _check_admin_token():
        return jsonify({"error": "unauthorized"}), 401
    resultados = _maternidade_process()
    queue = _maternidade_load()
    return jsonify({
        "status": "ok",
        "data": _dt_followup.date.today().isoformat(),
        "total_fila": len(queue),
        "resultados": resultados
    })


# Maternity timer (runs alongside follow-up timer)
def _maternidade_timer_loop():
    """Background thread for maternity engagement messages."""
    import time as _time_mat
    print(f"[MATERNIDADE] Timer iniciado - máx {MATERNIDADE_MAX_DIA} msgs/dia")
    escaneado_hoje = False
    ultimo_dia = ""

    while True:
        try:
            _br_tz = _dt_followup.timezone(_dt_followup.timedelta(hours=-3))
            agora = _dt_followup.datetime.now(_br_tz)
            hoje_str = agora.date().isoformat()
            hora = agora.hour

            # Reset daily
            if ultimo_dia != hoje_str:
                escaneado_hoje = False
                ultimo_dia = hoje_str

            # Run once per day, after follow-up (start at 10h to not overlap)
            if hora >= 10 and not escaneado_hoje:
                if _is_horario_comercial():
                    print(f"[MATERNIDADE] Executando check diário...")
                    escaneado_hoje = True
                    with app.app_context():
                        try:
                            resultados = _maternidade_process()
                            print(f"[MATERNIDADE] Resultado: {len(resultados)} ações")
                        except Exception as e:
                            print(f"[MATERNIDADE] Erro: {e}")
                            traceback.print_exc()

            # Check every 30 min
            _time_mat.sleep(1800)
        except Exception as e:
            print(f"[MATERNIDADE] Erro no timer: {e}")
            _time_mat.sleep(3600)

if MATERNIDADE_ENABLED:
    threading.Thread(target=_maternidade_timer_loop, daemon=True).start()
    print(f"[MATERNIDADE] Acompanhamento ATIVADO - máx {MATERNIDADE_MAX_DIA} msgs/dia")
else:
    print("[MATERNIDADE] Acompanhamento DESATIVADO - ative com MATERNIDADE_ENABLED=true")


# ========== REGISTER BLUEPRINTS ==========
from mayahub import mayahub_bp, MAYAHUB_API_KEY
from relatorios import relatorios_bp
app.register_blueprint(mayahub_bp)
app.register_blueprint(relatorios_bp)

if MAYAHUB_API_KEY:
    print(f"[MAYAHUB] Voice AI ATIVADO - Assistant ID: {os.environ.get('MAYAHUB_ASSISTANT_ID', '')}")
else:
    print("[MAYAHUB] Voice AI DESATIVADO - configure MAYAHUB_API_KEY")

print("[RELATORIOS] Pagina de relatorios disponivel em /relatorios")

try:
    _startup()
except Exception as e:
    print(f"[STARTUP ERROR] {e}")

if __name__ == "__main__":
    print(f"Diretório da aplicação: {BASE_DIR}")
    print(f"Timbrado: {TIMBRADO_PATH}")
    print(f"Output: {OUTPUT_DIR}")
    print("\nIniciando servidor em http://localhost:5000")
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", debug=False, port=port)
